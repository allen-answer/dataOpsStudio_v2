from __future__ import annotations

import io

from openpyxl import load_workbook

from app.infrastructure.result_export import write_xlsx_workbook
from app.services.lineage_batch_export import (
    _MAX_CELL_CHARS,
    SHEET_DYNAMIC_SQL,
    SHEET_FIELD_MAPPINGS,
    SHEET_FLOW_GROUPS,
    SHEET_OVERVIEW,
    SHEET_PARSE_ERRORS,
    SHEET_PROCEDURE_SEGMENTS,
    SHEET_RISKS,
    SHEET_SCRIPT_EDGES,
    SHEET_SCRIPTS,
    SHEET_STATEMENTS,
    SHEET_TABLE_FLOW,
    SHEET_TABLE_ROLES,
    SHEET_TARGET_INTEGRATION,
    SHEET_WARNINGS,
    batch_export_sheets,
)

_A_GROUP_SHEETS = [
    SHEET_OVERVIEW,
    SHEET_SCRIPTS,
    SHEET_SCRIPT_EDGES,
    SHEET_RISKS,
    SHEET_FLOW_GROUPS,
    SHEET_TABLE_ROLES,
    SHEET_TARGET_INTEGRATION,
]
_B_GROUP_SHEETS = [
    SHEET_PROCEDURE_SEGMENTS,
    SHEET_STATEMENTS,
    SHEET_PARSE_ERRORS,
    SHEET_DYNAMIC_SQL,
    SHEET_WARNINGS,
    SHEET_FIELD_MAPPINGS,
    SHEET_TABLE_FLOW,
]
_EXPECTED_SHEETS = _A_GROUP_SHEETS + _B_GROUP_SHEETS


def _full_report() -> dict[str, object]:
    return {
        "file_count": 3,
        "parsed": 2,
        "failed": 1,
        "skipped": {"non_sql": 4, "too_large": 1, "over_file_limit": 2},
        "table_edge_total": 5,
        "column_mapping_total": 7,
        "files": [
            {
                "source_ref": "etl/a.sql",
                "status": "parsed",
                "run_id": "run-1",
                "table_edge_count": 3,
                "column_mapping_count": 4,
                "parse_error_count": 0,
                "lenient_statement_count": 1,
                "tables_read": ["app.src", "app.dim"],
                "tables_written": ["app.mid"],
                "details": {
                    "procedure_segments": [
                        {
                            "procedure_name": "load_mid",
                            "procedure_kind": "PROCEDURE",
                            "segment_index": 0,
                            "line_start": 1,
                            "line_end": 9,
                            "parse_status": "ok",
                            "preceding_comment": "-- load",
                            "sql": "insert into app.mid select 1",
                        }
                    ],
                    "statements": [{"index": 0, "type": "INSERT"}],
                    "parse_errors": [
                        {
                            "statement_index": 2,
                            "error_type": "lineage_error",
                            "message": "OptimizeError",
                            "statement_type": "MERGE",
                            "unsupported": True,
                        }
                    ],
                    "dynamic_sql_segments": [
                        {
                            "statement_index": 3,
                            "confidence": "low",
                            "message": "execute immediate",
                            "sql": "EXECUTE IMMEDIATE v_sql",
                        }
                    ],
                    "warnings": [
                        {"code": "lenient_table_level", "message": "缺表降级", "statement_index": 0}
                    ],
                    "insert_mappings": [
                        {
                            "target_table": "app.mid",
                            "target_column": "id",
                            "source_table": "app.src",
                            "source_column": "id",
                            "transformation": "DIRECT",
                            "transformation_subtype": "DIRECT",
                            "statement_index": 0,
                        }
                    ],
                    "graph_edges": [
                        {
                            "source_table": "app.src",
                            "target_table": "app.mid",
                            "statement_index": 0,
                            "inferred": False,
                        }
                    ],
                },
            },
            {
                "source_ref": "etl/bad.sql",
                "status": "failed",
                "error": "undecodable_or_empty",
            },
        ],
        "script_edges": [
            # worker 落的真实键:source_file/target_file/tables(list)
            {"source_file": "etl/a.sql", "target_file": "etl/b.sql", "tables": ["app.mid"]},
        ],
        "semantic_view": {
            "targets": [
                {
                    "table": "app.mid",
                    "primary_role": "target",
                    "roles": ["target", "intermediate"],
                    "refresh_mode": "truncate_insert",
                    "counts": {"insert": 2, "truncate": 1},
                    "source_tables": ["app.src", "app.dim"],
                    "source_files": ["etl/a.sql"],
                    "titles": ["load mid"],
                }
            ],
            "table_roles": [
                {"table": "app.dim", "primary_role": "dimension", "roles": ["dimension"]},
            ],
            "observations": ["app.mid 全量重刷"],
            "risks": [
                {"level": "high", "type": "cycle", "message": "环依赖 app.a -> app.b"},
            ],
        },
    }


def test_returns_fourteen_named_sheets_in_order() -> None:
    sheets = batch_export_sheets(_full_report())
    assert [name for name, _, _ in sheets] == _EXPECTED_SHEETS
    # B 组紧接 A 组之后
    assert [name for name, _, _ in sheets][7:] == _B_GROUP_SHEETS


def test_overview_single_row_maps_top_level_stats() -> None:
    sheets = dict((name, (cols, rows)) for name, cols, rows in batch_export_sheets(_full_report()))
    cols, rows = sheets[SHEET_OVERVIEW]
    assert [c.name for c in cols] == [
        "file_count",
        "parsed",
        "failed",
        "skipped_non_sql",
        "skipped_too_large",
        "skipped_over_file_limit",
        "table_edge_total",
        "column_mapping_total",
        "script_edge_count",
    ]
    assert len(rows) == 1
    assert rows[0].values == [3, 2, 1, 4, 1, 2, 5, 7, 1]


def test_scripts_sheet_lists_each_file_and_joins_table_lists() -> None:
    sheets = dict((name, (cols, rows)) for name, cols, rows in batch_export_sheets(_full_report()))
    _, rows = sheets[SHEET_SCRIPTS]
    assert len(rows) == 2
    first = rows[0].values
    assert first[0] == "etl/a.sql"
    assert first[1] == "parsed"
    assert first[2] == "run-1"
    assert first[3:7] == [3, 4, 0, 1]
    assert first[7] == "app.src; app.dim"
    assert first[8] == "app.mid"
    # failed file: missing counts default to 0, run_id/error handled as blanks/text
    failed = rows[1].values
    assert failed[0] == "etl/bad.sql"
    assert failed[1] == "failed"
    assert failed[2] is None
    assert failed[3:7] == [0, 0, 0, 0]
    assert failed[9] == "undecodable_or_empty"


def test_script_edges_sheet_rows() -> None:
    sheets = dict((name, (cols, rows)) for name, cols, rows in batch_export_sheets(_full_report()))
    cols, rows = sheets[SHEET_SCRIPT_EDGES]
    assert [c.name for c in cols] == ["producer_file", "consumer_file", "table"]
    assert rows[0].values == ["etl/a.sql", "etl/b.sql", "app.mid"]


def test_risks_sheet_includes_risks_then_observations() -> None:
    sheets = dict((name, (cols, rows)) for name, cols, rows in batch_export_sheets(_full_report()))
    _, rows = sheets[SHEET_RISKS]
    assert rows[0].values == ["high", "cycle", "环依赖 app.a -> app.b"]
    # observation appended as a type=observation row (level blank)
    assert rows[1].values == [None, "observation", "app.mid 全量重刷"]


def test_flow_groups_sheet_aggregates_by_target() -> None:
    sheets = dict((name, (cols, rows)) for name, cols, rows in batch_export_sheets(_full_report()))
    cols, rows = sheets[SHEET_FLOW_GROUPS]
    assert [c.name for c in cols] == ["target_table", "source_tables", "source_files"]
    assert rows[0].values == ["app.mid", "app.src; app.dim", "etl/a.sql"]


def test_table_roles_sheet_rows() -> None:
    sheets = dict((name, (cols, rows)) for name, cols, rows in batch_export_sheets(_full_report()))
    cols, rows = sheets[SHEET_TABLE_ROLES]
    assert [c.name for c in cols] == ["table", "primary_role", "roles"]
    assert rows[0].values == ["app.dim", "dimension", "dimension"]


def test_target_integration_expands_count_columns() -> None:
    sheets = dict((name, (cols, rows)) for name, cols, rows in batch_export_sheets(_full_report()))
    cols, rows = sheets[SHEET_TARGET_INTEGRATION]
    assert [c.name for c in cols] == [
        "table",
        "primary_role",
        "roles",
        "refresh_mode",
        "count_insert",
        "count_update",
        "count_merge",
        "count_delete",
        "count_truncate",
        "source_tables",
        "source_files",
        "titles",
    ]
    values = rows[0].values
    assert values[0] == "app.mid"
    assert values[3] == "truncate_insert"
    # counts: insert=2, update=0, merge=0, delete=0, truncate=1
    assert values[4:9] == [2, 0, 0, 0, 1]
    assert values[9] == "app.src; app.dim"
    assert values[11] == "load mid"


def test_procedure_segments_sheet_prefixes_file_name() -> None:
    by_name = {name: (cols, rows) for name, cols, rows in batch_export_sheets(_full_report())}
    cols, rows = by_name[SHEET_PROCEDURE_SEGMENTS]
    assert [c.name for c in cols] == [
        "file_name",
        "procedure_name",
        "procedure_kind",
        "segment_index",
        "line_start",
        "line_end",
        "parse_status",
        "preceding_comment",
        "sql",
    ]
    assert len(rows) == 1
    assert rows[0].values[0] == "etl/a.sql"
    assert rows[0].values[1] == "load_mid"
    assert rows[0].values[3:6] == [0, 1, 9]
    assert rows[0].values[8] == "insert into app.mid select 1"


def test_statements_sheet_maps_index_and_type() -> None:
    by_name = {name: (cols, rows) for name, cols, rows in batch_export_sheets(_full_report())}
    cols, rows = by_name[SHEET_STATEMENTS]
    assert [c.name for c in cols] == ["file_name", "statement_index", "type"]
    assert rows[0].values == ["etl/a.sql", 0, "INSERT"]


def test_parse_errors_sheet_rows() -> None:
    by_name = {name: (cols, rows) for name, cols, rows in batch_export_sheets(_full_report())}
    cols, rows = by_name[SHEET_PARSE_ERRORS]
    assert [c.name for c in cols] == [
        "file_name",
        "statement_index",
        "error_type",
        "message",
        "statement_type",
        "unsupported",
    ]
    assert rows[0].values[:5] == ["etl/a.sql", 2, "lineage_error", "OptimizeError", "MERGE"]
    assert rows[0].values[5] == "True"


def test_dynamic_sql_sheet_rows() -> None:
    by_name = {name: (cols, rows) for name, cols, rows in batch_export_sheets(_full_report())}
    cols, rows = by_name[SHEET_DYNAMIC_SQL]
    assert [c.name for c in cols] == [
        "file_name",
        "statement_index",
        "confidence",
        "message",
        "sql",
    ]
    assert rows[0].values == [
        "etl/a.sql",
        3,
        "low",
        "execute immediate",
        "EXECUTE IMMEDIATE v_sql",
    ]


def test_warnings_sheet_rows() -> None:
    by_name = {name: (cols, rows) for name, cols, rows in batch_export_sheets(_full_report())}
    cols, rows = by_name[SHEET_WARNINGS]
    assert [c.name for c in cols] == ["file_name", "code", "message", "statement_index"]
    assert rows[0].values == ["etl/a.sql", "lenient_table_level", "缺表降级", 0]


def test_field_mappings_sheet_rows() -> None:
    by_name = {name: (cols, rows) for name, cols, rows in batch_export_sheets(_full_report())}
    cols, rows = by_name[SHEET_FIELD_MAPPINGS]
    assert [c.name for c in cols] == [
        "file_name",
        "target_table",
        "target_column",
        "source_table",
        "source_column",
        "transformation",
        "transformation_subtype",
        "statement_index",
    ]
    assert rows[0].values == ["etl/a.sql", "app.mid", "id", "app.src", "id", "DIRECT", "DIRECT", 0]


def test_table_flow_sheet_rows() -> None:
    by_name = {name: (cols, rows) for name, cols, rows in batch_export_sheets(_full_report())}
    cols, rows = by_name[SHEET_TABLE_FLOW]
    assert [c.name for c in cols] == [
        "file_name",
        "source_table",
        "target_table",
        "statement_index",
        "inferred",
    ]
    assert rows[0].values == ["etl/a.sql", "app.src", "app.mid", 0, "False"]


def test_empty_report_yields_fourteen_sheets_with_only_headers() -> None:
    sheets = batch_export_sheets({})
    assert [name for name, _, _ in sheets] == _EXPECTED_SHEETS
    by_name = {name: rows for name, _, rows in sheets}
    # overview always emits a single zeroed summary row
    assert by_name[SHEET_OVERVIEW][0].values == [0, 0, 0, 0, 0, 0, 0, 0, 0]
    for name in _EXPECTED_SHEETS:
        if name == SHEET_OVERVIEW:
            continue
        assert by_name[name] == []


def test_files_without_details_leave_b_group_empty() -> None:
    # 旧报告 / worker 未持久化 details:B 组守空,不炸。
    report = {"files": [{"source_ref": "etl/a.sql", "status": "parsed"}]}
    by_name = {name: rows for name, _, rows in batch_export_sheets(report)}
    for name in _B_GROUP_SHEETS:
        assert by_name[name] == []


def test_none_report_is_tolerated() -> None:
    sheets = batch_export_sheets(None)
    assert [name for name, _, _ in sheets] == _EXPECTED_SHEETS


def test_missing_semantic_view_leaves_semantic_sheets_empty() -> None:
    report = {"file_count": 1, "parsed": 1, "failed": 0, "files": [], "script_edges": []}
    by_name = {name: rows for name, _, rows in batch_export_sheets(report)}
    assert by_name[SHEET_RISKS] == []
    assert by_name[SHEET_FLOW_GROUPS] == []
    assert by_name[SHEET_TABLE_ROLES] == []
    assert by_name[SHEET_TARGET_INTEGRATION] == []


def test_long_and_control_char_cells_are_sanitized() -> None:
    long_table = "app." + "x" * (_MAX_CELL_CHARS + 500)
    report = {
        "semantic_view": {
            "risks": [{"level": "low", "type": "note", "message": "line1\nline2\ttab"}],
            "table_roles": [{"table": long_table, "primary_role": "source_fact", "roles": []}],
        }
    }
    by_name = {name: rows for name, _, rows in batch_export_sheets(report)}
    message = by_name[SHEET_RISKS][0].values[2]
    assert "\n" not in message and "\t" not in message
    assert message == "line1 line2 tab"
    table_cell = by_name[SHEET_TABLE_ROLES][0].values[0]
    assert len(table_cell) <= _MAX_CELL_CHARS
    assert table_cell.endswith("…(truncated)")


def test_roundtrip_through_workbook_is_readable_and_formula_safe() -> None:
    report = _full_report()
    # inject a formula-like risk message to confirm write_xlsx_workbook prefix guard
    report["semantic_view"]["risks"].append(  # type: ignore[index]
        {"level": "low", "type": "note", "message": "=SUM(A1:A2)"}
    )
    stream = io.BytesIO()
    write_xlsx_workbook(
        stream=stream,
        sheets=batch_export_sheets(report),
        limit_bytes=8 * 1024 * 1024,
    )
    stream.seek(0)
    workbook = load_workbook(stream, read_only=True)
    assert workbook.sheetnames == _EXPECTED_SHEETS
    overview = workbook[SHEET_OVERVIEW]
    header = [cell.value for cell in next(overview.iter_rows())]
    assert header[0] == "file_count"
    data_row = [cell.value for cell in list(overview.iter_rows())[1]]
    assert data_row[0] == 3  # numeric cell preserved as number
    risks = workbook[SHEET_RISKS]
    messages = [row[2].value for row in list(risks.iter_rows())[1:]]
    assert "'=SUM(A1:A2)" in messages  # formula prefix neutralized
