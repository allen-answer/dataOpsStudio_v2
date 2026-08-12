from __future__ import annotations

import io
import json
from collections.abc import Callable, Iterator
from datetime import UTC, datetime, timedelta
from typing import Any, cast

import pytest
from fastapi.routing import APIRoute

from app.api.app import create_app
from app.api.routes import core as core_routes
from app.api.schemas import SqlGenerateResponse
from app.api.security import create_access_token, decode_access_token
from app.api.services import ApiServices
from app.dbclients.protocol import AdapterConnectionError
from app.domain.ai import AiContext, AiOptions, AiResponse, EgressLevel
from app.domain.ai_copilot import MAX_TABLES
from app.domain.compare_result import encode_compare_result_row
from app.domain.job import Job, JobKind
from app.domain.license import LicenseMode
from app.domain.result import (
    SQL_WORKSPACE_DEFAULT_MAX_ROWS,
    SQL_WORKSPACE_DEFAULT_PAGE_SIZE,
    SQL_WORKSPACE_MAX_ROWS,
    ResultRef,
)
from app.domain.schema import Column, ColumnType, Index, Row, Schema, Table
from app.domain.workflow import WorkflowSpec
from app.services.ai.default_gateway import DefaultAiGateway
from app.services.ai.errors import (
    AiGatewayError,
    BudgetExceededError,
    EgressBlockedError,
    ProviderError,
)
from tests._asgi_client import AsgiClient

pytestmark = pytest.mark.contract


def test_t4_api_route_surface_matches_contract() -> None:
    app = create_app(services=cast(ApiServices, _NoopServices()))

    routes = {
        (method, route.path)
        for route in app.routes
        if isinstance(route, APIRoute)
        for method in route.methods
    }

    assert ("POST", "/api/auth/login") in routes
    assert ("GET", "/api/projects") in routes
    assert ("GET", "/api/datasources") in routes
    assert ("POST", "/api/datasources") in routes
    assert ("GET", "/api/datasources/{datasource_id}") in routes
    assert ("PUT", "/api/datasources/{datasource_id}") in routes
    assert ("DELETE", "/api/datasources/{datasource_id}") in routes
    assert ("POST", "/api/datasources/{datasource_id}/test") in routes
    assert ("POST", "/api/sql/execute") in routes
    assert ("GET", "/api/version") in routes
    assert ("GET", "/api/sql/consoles") in routes
    assert ("POST", "/api/sql/consoles") in routes
    assert ("PATCH", "/api/sql/consoles/{console_id}") in routes
    assert ("DELETE", "/api/sql/consoles/{console_id}") in routes
    assert ("GET", "/api/sql/history") in routes
    assert ("GET", "/api/sql/templates") in routes
    assert ("POST", "/api/sql/templates") in routes
    assert ("PATCH", "/api/sql/templates/{template_id}") in routes
    assert ("DELETE", "/api/sql/templates/{template_id}") in routes
    assert ("POST", "/api/sql/templates/{template_id}/render") in routes
    assert ("POST", "/api/sql/explain") in routes
    assert ("POST", "/api/sql/preflight") in routes
    assert ("POST", "/api/sql/format") in routes
    assert ("POST", "/api/sql/expand-star") in routes
    assert ("GET", "/api/compare/tasks") in routes
    assert ("POST", "/api/compare/tasks") in routes
    assert ("GET", "/api/compare/tasks/{task_id}") in routes
    assert ("PATCH", "/api/compare/tasks/{task_id}") in routes
    assert ("DELETE", "/api/compare/tasks/{task_id}") in routes
    assert ("POST", "/api/compare/tasks/{task_id}/run") in routes
    assert ("GET", "/api/compare/runs/{run_id}/results") in routes
    assert ("GET", "/api/compare/runs/{run_id}/profile") in routes
    assert ("GET", "/api/compare/runs/{run_id}/diff-sql") in routes
    assert ("POST", "/api/compare/runs/{run_id}/ai-attribution") in routes
    assert ("POST", "/api/projects/{project_id}/compare/infer") in routes
    assert ("POST", "/api/projects/{project_id}/compare/pk-precheck") in routes
    assert ("GET", "/api/projects/{project_id}/compare/suggest-tasks") in routes
    assert ("POST", "/api/projects/{project_id}/compare/draft-task") in routes
    assert ("GET", "/api/projects/{project_id}/compare/runs-dashboard") in routes
    assert ("POST", "/api/projects/{project_id}/lineage/analyze") in routes
    assert ("GET", "/api/projects/{project_id}/lineage/subgraph") in routes
    assert ("GET", "/api/projects/{project_id}/lineage/impact") in routes
    assert ("POST", "/api/projects/{project_id}/lineage/ai-impact") in routes
    assert ("POST", "/api/projects/{project_id}/lineage/export") in routes

    assert ("PATCH", "/api/projects/{project_id}/lineage/edges/{edge_id}") in routes
    assert ("POST", "/api/projects/{project_id}/lineage/runs/{run_id}/ai-enrich") in routes
    assert ("GET", "/api/projects/{project_id}/lineage/runs/{run_id}/ai-enrich") in routes
    assert ("GET", "/api/projects/{project_id}/workflows") in routes
    assert ("POST", "/api/projects/{project_id}/workflows") in routes
    assert ("GET", "/api/projects/{project_id}/workflows/{workflow_id}") in routes
    assert ("PUT", "/api/projects/{project_id}/workflows/{workflow_id}") in routes
    assert ("DELETE", "/api/projects/{project_id}/workflows/{workflow_id}") in routes
    assert ("POST", "/api/projects/{project_id}/workflows/{workflow_id}/runs") in routes
    assert ("GET", "/api/projects/{project_id}/workflows/{workflow_id}/runs") in routes
    assert ("GET", "/api/projects/{project_id}/workflow-runs/{run_id}") in routes
    assert ("POST", "/api/projects/{project_id}/workflow-runs/{run_id}:cancel") in routes
    assert ("GET", "/api/datasources/{datasource_id}/metadata/schemas") in routes
    assert ("GET", "/api/datasources/{datasource_id}/metadata/tables") in routes
    assert ("GET", "/api/datasources/{datasource_id}/metadata/columns") in routes
    assert ("GET", "/api/datasources/{datasource_id}/metadata/indexes") in routes
    assert ("POST", "/api/datasources/{datasource_id}/ai/sql-generate") in routes
    assert ("POST", "/api/datasources/{datasource_id}/ai/slow-sql-diagnose") in routes
    assert ("GET", "/api/jobs") in routes
    assert ("GET", "/api/jobs/{job_id}") in routes
    assert ("GET", "/api/jobs/{job_id}/progress") in routes
    assert ("GET", "/api/jobs/{job_id}/result") in routes
    assert ("POST", "/api/jobs/{job_id}/pages") in routes
    assert ("POST", "/api/jobs/{job_id}/export") in routes
    assert ("POST", "/api/compare/runs/{run_id}/export") in routes
    assert ("GET", "/api/exports/{token}") in routes
    assert ("POST", "/api/jobs/{job_id}/cancel") in routes
    assert ("GET", "/api/account/security") in routes
    assert ("POST", "/api/account/password") in routes
    assert ("POST", "/api/account/mfa/enroll") in routes
    assert ("POST", "/api/account/mfa/verify") in routes
    assert ("POST", "/api/account/mfa/disable") in routes
    assert ("POST", "/api/account/recovery-codes/regenerate") in routes
    assert ("GET", "/api/admin/ai-config") in routes
    assert ("PUT", "/api/admin/ai-config") in routes
    assert ("POST", "/api/admin/ai-config/test") in routes
    assert ("GET", "/api/admin/system-settings") in routes
    assert ("PUT", "/api/admin/system-settings") in routes


def test_version_is_public_and_reports_sanitized_build_identity(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("DATAOPS_BUILD_VERSION", "2.0.1-test")
    monkeypatch.setenv("DATAOPS_BUILD_COMMIT", "abcdef0123456789")
    monkeypatch.setenv("DATAOPS_IMAGE_VERSION", "2.0.1-test-image")
    app = create_app(services=cast(ApiServices, _NoopServices()))

    response = AsgiClient(app).get("/api/version")

    assert response.status_code == 200
    assert response.json() == {
        "version": "2.0.1-test",
        "commit": "abcdef0123456789",
        "image_version": "2.0.1-test-image",
    }


@pytest.mark.parametrize(
    ("requested_max_rows", "expected_max_rows"),
    [
        (None, SQL_WORKSPACE_DEFAULT_MAX_ROWS),
        (5_000, 5_000),
    ],
)
def test_sql_execute_enqueues_validated_row_limit(
    requested_max_rows: int | None,
    expected_max_rows: int,
) -> None:
    services = _Services(
        _FakeEngine(
            [
                _datasource_row(),
                {"id": "project-1"},
            ]
        )
    )
    app = create_app(services=cast(ApiServices, services))
    body: dict[str, object] = {"sql": "SELECT * FROM users", "datasource_id": "ds-1"}
    if requested_max_rows is not None:
        body["max_rows"] = requested_max_rows

    response = AsgiClient(app).post(
        "/api/sql/execute",
        headers=_auth_headers(),
        json_body=body,
    )

    assert response.status_code == 202
    payload = services.job_backend.enqueued[0].payload
    assert payload["page_size"] == SQL_WORKSPACE_DEFAULT_PAGE_SIZE
    assert payload["max_result_rows"] == expected_max_rows
    assert payload["max_rows"] == expected_max_rows
    assert payload["pagination_mode"] == "unavailable"
    assert payload["pagination_reason"] == "top_level_order_by_required"


def test_sql_execute_separates_page_size_and_safety_limit() -> None:
    services = _Services(
        _FakeEngine(
            [
                _datasource_row(),
                {"id": "project-1"},
            ]
        )
    )
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/sql/execute",
        headers=_auth_headers(),
        json_body={
            "sql": "SELECT * FROM users ORDER BY id",
            "datasource_id": "ds-1",
            "page_size": 100,
            "max_result_rows": 5_000,
        },
    )

    assert response.status_code == 202
    payload = services.job_backend.enqueued[0].payload
    assert payload["page_size"] == 100
    assert payload["max_result_rows"] == 5_000
    assert payload["pagination_mode"] == "ordered_offset"
    assert payload["db_type"] == "mysql"


def test_job_progress_returns_manifest_version_without_fetching_rows() -> None:
    created_at = datetime(2026, 8, 12, 1, 0, tzinfo=UTC)
    started_at = created_at + timedelta(seconds=2)
    job_row = {
        "id": "job-progress",
        "kind": "sql_query",
        "status": "running",
        "owner_user_id": "user-1",
        "project_id": "project-1",
        "datasource_ids": ["ds-1"],
        "worker_id": "worker-a",
        "created_at": created_at,
        "started_at": started_at,
        "finished_at": None,
        "error_code": None,
        "payload": {
            "result_set_id": "rs-progress",
            "max_result_rows": 1000,
            "db_type": "mysql",
            "limit_pushdown": False,
            "limit_pushdown_reason": "aggregate_requires_full_input",
            "output_limit_applied": True,
            "query_shape": "aggregate",
        },
        "result_ref": {
            "backend": "local_fs",
            "uri": "resultsets/rs-progress/manifest.json",
            "metadata": {
                "timings": {"connect_ms": 12, "execute_first_row_ms": 30},
                "execution": {
                    "connect_started_at": "2026-08-12T01:00:02+00:00",
                    "connected_at": "2026-08-12T01:00:02.012+00:00",
                    "rows_read": 101,
                    "rows_returned": 100,
                    "effective_sql_hash": "a" * 64,
                },
            },
        },
    }
    result_store = _ResultStore(
        rows=[Row(values=["must-not-be-read"])],
        manifest={
            "columns": [{"name": "count", "type": "integer"}],
            "loaded_rows": 100,
            "result_version": 3,
            "first_batch_at": created_at.timestamp() + 3,
            "truncated": False,
            "has_more": True,
            "pagination_mode": "unavailable",
            "pagination_reason": "top_level_order_by_required",
        },
    )
    services = _Services(
        _FakeEngine([job_row, {"id": "project-1"}]),
        result_store=result_store,
    )
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).get(
        "/api/jobs/job-progress/progress",
        headers=_auth_headers(),
        params={"after_version": 2},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "running"
    assert payload["loaded_rows"] == 100
    assert payload["result_version"] == 3
    assert payload["has_new_result"] is True
    assert payload["first_batch_ready"] is True
    assert payload["retry_after_ms"] == 1000
    assert payload["timings"] == {
        "queue_ms": 2000,
        "connect_ms": 12,
        "execute_first_row_ms": 30,
        "fetch_ms": None,
        "spool_ms": None,
        "total_ms": None,
    }
    assert payload["execution"]["rows_read"] == 101
    assert payload["execution"]["rows_returned"] == 100
    assert payload["execution"]["limit_pushdown"] is False
    assert payload["execution"]["output_limit_applied"] is True
    assert payload["execution"]["query_shape"] == "aggregate"
    assert payload["execution"]["worker_id"] == "worker-a"
    assert "must-not-be-read" not in response.body.decode()


def test_sql_execute_rejects_conflicting_new_and_legacy_safety_limits() -> None:
    app = create_app(services=cast(ApiServices, _Services(_FakeEngine([]))))

    response = AsgiClient(app).post(
        "/api/sql/execute",
        headers=_auth_headers(),
        json_body={
            "sql": "SELECT * FROM users",
            "datasource_id": "ds-1",
            "max_result_rows": 100,
            "max_rows": 200,
        },
    )

    assert response.status_code == 422


def test_request_job_page_enqueues_ordered_stateless_continuation() -> None:
    root_row = _pagination_root_row(pagination_mode="ordered_offset")
    backend = _JobBackend()
    result_store = _ResultStore(
        manifest={
            "columns": [{"name": "id", "type": "integer"}],
            "loaded_rows": 100,
            "truncated": False,
            "has_more": True,
            "pagination_mode": "ordered_offset",
            "pagination_reason": "fresh_read_ordered_offset",
        }
    )
    services = _Services(
        _FakeEngine([root_row, {"id": "project-1"}]),
        job_backend=backend,
        result_store=result_store,
    )
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/jobs/job-root/pages",
        headers=_auth_headers(),
        json_body={"offset": 100},
    )

    assert response.status_code == 202
    page_job = backend.enqueued[0]
    assert page_job.payload["pagination_root_job_id"] == "job-root"
    assert page_job.payload["page_offset"] == 100
    assert page_job.payload["result_set_id"] == "rs-pages"
    assert page_job.payload["params"] == {"min_id": 1}
    assert page_job.datasource_ids == ["ds-1"]


def test_request_job_page_fails_closed_without_stable_order() -> None:
    services = _Services(
        _FakeEngine([_pagination_root_row(pagination_mode="unavailable"), {"id": "project-1"}]),
        result_store=_ResultStore(
            manifest={
                "loaded_rows": 100,
                "truncated": False,
                "has_more": True,
                "pagination_mode": "unavailable",
                "pagination_reason": "top_level_order_by_required",
            }
        ),
    )
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/jobs/job-root/pages",
        headers=_auth_headers(),
        json_body={"offset": 100},
    )

    assert response.status_code == 409
    assert response.json()["error"] == "stable_order_required"


@pytest.mark.parametrize("max_rows", [0, SQL_WORKSPACE_MAX_ROWS + 1])
def test_sql_execute_rejects_row_limit_outside_supported_range(max_rows: int) -> None:
    app = create_app(services=cast(ApiServices, _Services(_FakeEngine([]))))

    response = AsgiClient(app).post(
        "/api/sql/execute",
        headers=_auth_headers(),
        json_body={
            "sql": "SELECT * FROM users",
            "datasource_id": "ds-1",
            "max_rows": max_rows,
        },
    )

    assert response.status_code == 422


@pytest.mark.parametrize("page_size", [0, 1001])
def test_sql_execute_rejects_page_size_outside_supported_range(page_size: int) -> None:
    app = create_app(services=cast(ApiServices, _Services(_FakeEngine([]))))

    response = AsgiClient(app).post(
        "/api/sql/execute",
        headers=_auth_headers(),
        json_body={
            "sql": "SELECT * FROM users",
            "datasource_id": "ds-1",
            "page_size": page_size,
        },
    )

    assert response.status_code == 422


def test_login_uses_admin_configured_access_token_ttl() -> None:
    engine = _FakeEngine([_user_row_for_login()])
    services = _Services(engine, access_token_ttl_seconds=86_400)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/auth/login",
        json_body={"username": "admin", "password": "secret"},
    )

    assert response.status_code == 200
    token = response.json()["access_token"]
    claims = decode_access_token(token, secret=services.jwt_secret)
    assert claims.expires_at - claims.issued_at == 86_400


def test_admin_system_settings_roundtrip_updates_auth_and_license_flags() -> None:
    engine = _FakeEngine(
        [
            [],
            None,
            None,
            [
                _setting_row("auth.access_token_ttl_seconds", "86400"),
                _setting_row("license.enforcement_enabled", "false"),
            ],
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/admin/system-settings",
        headers=_auth_headers(),
        json_body={"access_token_ttl_seconds": 86_400, "license_enforcement_enabled": False},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["access_token_ttl_seconds"] == 86_400
    assert payload["license_enforcement_enabled"] is False
    assert any(audit["action"] == "admin_system_settings_update" for audit in services.audits)


def test_license_status_includes_enforcement_flag_from_system_settings() -> None:
    row = {
        "id": 1,
        "mode": "trial",
        "edition": None,
        "customer": None,
        "expires_at": _dt(1),
        "features": [],
        "repair_reason": None,
    }
    services = _Services(_FakeEngine([row]), license_enforcement_enabled=False)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).get("/api/license/status", headers=_auth_headers())

    assert response.status_code == 200
    assert response.json()["license_enforcement_enabled"] is False


def test_list_datasources_filters_by_authz_and_hides_secret_fields() -> None:
    engine = _FakeEngine(
        [
            [
                {
                    "id": "ds-visible",
                    "name": "warehouse",
                    "db_type": "mysql",
                    "host": "mysql.internal",
                    "port": 3306,
                    "environment": "prod",
                    "environment_verified": True,
                    "database_name": "app",
                    "created_at": _dt(1),
                    "username": "should-not-return",
                    "password_secret_ref": "secret-should-not-return",
                    "operation_policy": _policy(),
                },
                {
                    "id": "ds-no-database",
                    "name": "warehouse-no-db",
                    "db_type": "mysql",
                    "host": "mysql2.internal",
                    "port": 3307,
                    "environment": "sandbox",
                    "environment_verified": False,
                    "database_name": None,
                    "created_at": _dt(2),
                    "username": "should-not-return",
                    "password_secret_ref": "secret-should-not-return",
                    "operation_policy": _policy(allow_select=False),
                },
            ]
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/datasources",
        headers=_auth_headers(),
        params={"limit": 10, "offset": 0},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload == [
        {
            "id": "ds-visible",
            "name": "warehouse",
            "db_type": "mysql",
            "host": "mysql.internal",
            "port": 3306,
            "environment": "prod",
            "environment_verified": True,
            "database": "app",
            "operation_policy": _policy(),
            "created_at": "2026-01-01T00:00:01Z",
        },
        {
            "id": "ds-no-database",
            "name": "warehouse-no-db",
            "db_type": "mysql",
            "host": "mysql2.internal",
            "port": 3307,
            "environment": "sandbox",
            "environment_verified": False,
            "database": None,
            "operation_policy": _policy(allow_select=False),
            "created_at": "2026-01-01T00:00:02Z",
        },
    ]
    assert "password" not in response.body.decode("utf-8")
    assert "secret" not in response.body.decode("utf-8")
    assert "username" not in response.body.decode("utf-8")
    statement = engine.statements[-1]
    assert "project_members" in statement
    assert "owner_user_id" in statement
    assert "password_secret_ref" not in statement


def test_list_jobs_filters_paginates_and_hides_payload_and_raw_errors() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            [
                {
                    "id": "job-conn",
                    "kind": "test_connection",
                    "status": "failed",
                    "created_at": _dt(3),
                    "started_at": _dt(4),
                    "finished_at": _dt(5),
                    "error": "Access denied for user should-not-return",
                    "error_code": "connection_failed",
                    "payload": {"sql": "SELECT secret_should_not_return"},
                },
                {
                    "id": "job-sql",
                    "kind": "sql_query",
                    "status": "failed",
                    "created_at": _dt(2),
                    "started_at": None,
                    "finished_at": _dt(6),
                    "error": "Unknown column secret_schema.should_not_return",
                    "error_code": "sql_failed",
                    "payload": {"sql": "SELECT secret_should_not_return"},
                },
                {
                    "id": "job-timeout",
                    "kind": "sql_query",
                    "status": "failed",
                    "created_at": _dt(1),
                    "started_at": None,
                    "finished_at": _dt(7),
                    "error": "worker heartbeat timed out",
                    "error_code": "timeout",
                },
            ],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/jobs",
        headers=_auth_headers(),
        params={"project_id": "project-1", "status": "failed", "limit": 3, "offset": 0},
    )

    assert response.status_code == 200
    payload = response.json()
    assert [job["id"] for job in payload] == ["job-conn", "job-sql", "job-timeout"]
    assert [job["error"] for job in payload] == [
        "connection_failed",
        "sql_failed",
        "timeout",
    ]
    assert [job["error_code"] for job in payload] == [
        "connection_failed",
        "sql_failed",
        "timeout",
    ]
    body = response.body.decode("utf-8")
    assert "payload" not in body
    assert "SELECT secret_should_not_return" not in body
    assert "Access denied" not in body
    assert "Unknown column" not in body
    statement = engine.statements[-1]
    assert "project_members" in statement
    assert "jobs.owner_user_id" in statement
    assert "jobs.project_id" in statement
    assert "jobs.status" in statement
    assert "LIMIT" in statement
    assert "OFFSET" in statement


def test_list_jobs_classifies_cancelled_without_leaking_reason() -> None:
    engine = _FakeEngine(
        [
            [
                {
                    "id": "job-cancelled",
                    "kind": "sql_query",
                    "status": "cancelled",
                    "created_at": _dt(0),
                    "started_at": None,
                    "finished_at": _dt(8),
                    "error": "user requested cancel",
                    "error_code": None,
                    "payload": {"sql": "SELECT cancelled_should_not_return"},
                }
            ],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/jobs",
        headers=_auth_headers(),
        params={"status": "cancelled"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert [job["id"] for job in payload] == ["job-cancelled"]
    assert payload[0]["error"] == "cancelled"
    assert payload[0]["error_code"] is None
    body = response.body.decode("utf-8")
    assert "user requested cancel" not in body
    assert "SELECT cancelled_should_not_return" not in body


def test_get_job_response_includes_timestamps_for_terminal_jobs() -> None:
    engine = _FakeEngine(
        [
            {
                "id": "job-success",
                "project_id": "project-1",
                "kind": "sql_query",
                "status": "success",
                "created_at": _dt(1),
                "started_at": _dt(2),
                "finished_at": _dt(9),
                "error": None,
                "error_code": None,
                "result_ref": {
                    "backend": "local_fs",
                    "uri": "spool/rs-1",
                    "metadata": {
                        "timings": {
                            "connect_ms": 120,
                            "execute_first_row_ms": 340,
                            "fetch_ms": 560,
                            "spool_ms": 80,
                        }
                    },
                },
                "payload": {"result_set_id": "rs-1", "sql": "SELECT should_not_return"},
            },
            {"id": "project-1"},
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/jobs/job-success",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["created_at"] == "2026-01-01T00:00:01Z"
    assert payload["started_at"] == "2026-01-01T00:00:02Z"
    assert payload["finished_at"] == "2026-01-01T00:00:09Z"
    assert payload["timings"] == {
        "queue_ms": 1000,
        "connect_ms": 120,
        "execute_first_row_ms": 340,
        "fetch_ms": 560,
        "spool_ms": 80,
        "total_ms": 8000,
    }
    assert "SELECT should_not_return" not in response.body.decode("utf-8")


def test_list_jobs_project_filter_requires_authz() -> None:
    engine = _FakeEngine([None])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/jobs",
        headers=_auth_headers(),
        params={"project_id": "other-project"},
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"


def test_update_datasource_keeps_password_when_blank_and_returns_policy() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(password_secret_ref="secret-old"),
            {"id": "project-1"},
            _datasource_row(
                name="warehouse-renamed",
                password_secret_ref="secret-old",
                operation_policy=_policy(allow_select=False),
            ),
            {"id": "project-1"},
        ]
    )
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/datasources/ds-1",
        headers=_auth_headers(),
        json_body={
            "name": "warehouse-renamed",
            "password": "",
            "operation_policy": _policy(allow_select=False),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["name"] == "warehouse-renamed"
    assert payload["operation_policy"]["allow_select"] is False
    assert secret_store.stored == []
    assert secret_store.deleted == []
    assert "password" not in response.body.decode("utf-8")
    assert "secret-old" not in response.body.decode("utf-8")
    assert any("UPDATE datasources" in statement for statement in engine.statements)
    assert any(audit["action"] == "datasource_update" for audit in services.audits)


def test_update_datasource_clears_database_when_null_is_sent() -> None:
    cleared_row = _datasource_row()
    cleared_row["database_name"] = None
    engine = _FakeEngine(
        [
            _datasource_row(password_secret_ref="secret-old"),
            {"id": "project-1"},
            cleared_row,
            {"id": "project-1"},
        ]
    )
    services = _Services(engine, secret_store=_SecretStore())
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/datasources/ds-1",
        headers=_auth_headers(),
        json_body={"database": None},
    )

    assert response.status_code == 200
    assert response.json()["database"] is None
    update_stmt = next(stmt for stmt in engine.executed if "UPDATE datasources" in str(stmt))
    compiled = update_stmt.compile()
    assert "database_name" in compiled.params
    assert compiled.params["database_name"] is None


def test_update_datasource_rotates_password_when_new_value_present() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(password_secret_ref="secret-old"),
            {"id": "project-1"},
            _datasource_row(password_secret_ref="secret-new"),
            {"id": "project-1"},
        ]
    )
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/datasources/ds-1",
        headers=_auth_headers(),
        json_body={"password": "new-local-test-password"},
    )

    assert response.status_code == 200
    assert secret_store.stored == ["new-local-test-password"]
    assert secret_store.deleted == ["secret-old"]
    assert "new-local-test-password" not in response.body.decode("utf-8")
    assert "secret-new" not in response.body.decode("utf-8")


def test_delete_datasource_returns_409_with_job_references() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(password_secret_ref="secret-old"),
            {"id": "project-1"},
            [
                {
                    "id": "job-1",
                    "kind": "sql_query",
                    "status": "running",
                }
            ],
            [],
        ]
    )
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "DELETE",
        "/api/datasources/ds-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 409
    assert response.json() == {
        "error": "datasource_in_use",
        "message": "Datasource is referenced by existing jobs",
        "references": [{"job_id": "job-1", "kind": "sql_query", "status": "running"}],
    }
    assert secret_store.deleted == []
    assert not any("DELETE FROM datasources" in statement for statement in engine.statements)


def test_delete_datasource_ignores_terminal_test_connection_jobs() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(password_secret_ref="secret-old"),
            {"id": "project-1"},
            [],
            [],
        ]
    )
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "DELETE",
        "/api/datasources/ds-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 204
    assert secret_store.deleted == ["secret-old"]
    assert any("DELETE FROM datasources" in statement for statement in engine.statements)
    reference_query = next(statement for statement in engine.statements if "FROM jobs" in statement)
    assert "jobs.kind !=" in reference_query
    assert "jobs.status NOT IN" in reference_query


def test_delete_datasource_still_blocks_business_job_references() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(password_secret_ref="secret-old"),
            {"id": "project-1"},
            [
                {
                    "id": "job-query-success",
                    "kind": "sql_query",
                    "status": "success",
                },
            ],
            [],
        ]
    )
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "DELETE",
        "/api/datasources/ds-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 409
    assert response.json() == {
        "error": "datasource_in_use",
        "message": "Datasource is referenced by existing jobs",
        "references": [{"job_id": "job-query-success", "kind": "sql_query", "status": "success"}],
    }
    assert secret_store.deleted == []
    assert not any("DELETE FROM datasources" in statement for statement in engine.statements)


def test_metadata_endpoints_return_cached_contract_fields(monkeypatch: pytest.MonkeyPatch) -> None:
    adapter = _MetadataAdapter()
    adapter_kwargs: list[dict[str, object]] = []

    def build_adapter(
        conn_info: object,
        secret_store: object,
        **kwargs: object,
    ) -> _MetadataAdapter:
        del conn_info, secret_store
        adapter_kwargs.append(kwargs)
        return adapter

    monkeypatch.setattr(
        core_routes,
        "build_database_adapter",
        build_adapter,
    )
    engine = _FakeEngine(
        [
            _datasource_row(),
            {"id": "project-1"},
            _datasource_row(),
            {"id": "project-1"},
            _datasource_row(),
            {"id": "project-1"},
            _datasource_row(),
            {"id": "project-1"},
        ]
    )
    services = _Services(engine, metadata_probe_timeout_seconds=6)
    app = create_app(services=cast(ApiServices, services))
    client = AsgiClient(app)

    schemas_response = client.get(
        "/api/datasources/ds-1/metadata/schemas",
        headers=_auth_headers(),
        params={"refresh": "1"},
    )
    tables_response = client.get(
        "/api/datasources/ds-1/metadata/tables",
        headers=_auth_headers(),
        params={"schema": "app", "refresh": "1"},
    )
    columns_response = client.get(
        "/api/datasources/ds-1/metadata/columns",
        headers=_auth_headers(),
        params={"schema": "app", "table": "users", "refresh": "1"},
    )
    indexes_response = client.get(
        "/api/datasources/ds-1/metadata/indexes",
        headers=_auth_headers(),
        params={"schema": "app", "table": "users", "refresh": "1"},
    )

    assert schemas_response.status_code == 200
    assert schemas_response.json() == [{"name": "app"}]
    assert tables_response.status_code == 200
    assert tables_response.json() == [
        {"schema_name": "app", "name": "users", "table_type": "BASE TABLE"}
    ]
    assert columns_response.status_code == 200
    assert columns_response.json() == [
        {
            "name": "id",
            "type": "integer",
            "driver_type": "INT",
            "nullable": False,
            "primary_key": True,
            "comment": "primary key",
        },
        {
            "name": "name",
            "type": "string",
            "driver_type": "VARCHAR(64)",
            "nullable": True,
            "primary_key": False,
            "comment": None,
        },
    ]
    assert indexes_response.status_code == 200
    assert indexes_response.json() == [
        {"name": "PRIMARY", "columns": ["id"], "is_unique": True, "is_primary": True}
    ]
    assert {audit["action"] for audit in services.audits} >= {
        "metadata_schemas_list",
        "metadata_tables_list",
        "metadata_columns_list",
        "metadata_indexes_list",
    }
    assert adapter_kwargs == [
        {"connect_timeout_seconds": 6, "statement_timeout_seconds": 6},
        {"connect_timeout_seconds": 6, "statement_timeout_seconds": 6},
        {"connect_timeout_seconds": 6, "statement_timeout_seconds": 6},
        {"connect_timeout_seconds": 6, "statement_timeout_seconds": 6},
    ]


def test_metadata_probe_connection_failure_is_structured_error(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        core_routes,
        "build_database_adapter",
        lambda conn_info, secret_store, **kwargs: _FailingMetadataAdapter(),
    )
    engine = _FakeEngine([_datasource_row(), {"id": "project-1"}])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/datasources/ds-1/metadata/schemas",
        headers=_auth_headers(),
        params={"refresh": "1"},
    )

    assert response.status_code == 503
    assert response.json()["error"] == "metadata_probe_failed"


def test_sql_format_endpoint_returns_formatted_sql() -> None:
    app = create_app(services=cast(ApiServices, _Services(_FakeEngine([]))))

    response = AsgiClient(app).post(
        "/api/sql/format",
        headers=_auth_headers(),
        json_body={"sql": "select 1 as n", "dialect": "mysql"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert set(payload) == {"formatted_sql"}
    assert "SELECT" in payload["formatted_sql"]
    assert "select 1 as n" not in payload["formatted_sql"]


def test_sql_expand_star_uses_metadata_cache_contract_fields() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(),
            {"id": "project-1"},
            {
                "payload": [
                    {
                        "name": "id",
                        "type": "integer",
                        "driver_type": "INT",
                        "nullable": False,
                        "primary_key": True,
                        "comment": "primary key",
                    },
                    {
                        "name": "name",
                        "type": "string",
                        "driver_type": "VARCHAR(64)",
                        "nullable": True,
                        "primary_key": False,
                        "comment": None,
                    },
                ],
                "expires_at": datetime.now(UTC) + timedelta(days=1),
            },
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/sql/expand-star",
        headers=_auth_headers(),
        json_body={"sql": "SELECT * FROM users", "datasource_id": "ds-1"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert set(payload) == {"expanded_sql"}
    assert "id" in payload["expanded_sql"]
    assert "name" in payload["expanded_sql"]
    assert "*" not in payload["expanded_sql"]


def test_sql_expand_star_cache_miss_is_structured_error() -> None:
    engine = _FakeEngine([_datasource_row(), {"id": "project-1"}, None])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/sql/expand-star",
        headers=_auth_headers(),
        json_body={"sql": "SELECT * FROM users", "datasource_id": "ds-1"},
    )

    assert response.status_code == 409
    assert response.json()["error"] == "metadata_cache_missing"


def test_sql_explain_enqueues_explain_job_with_result_set() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(operation_policy=_policy(allow_explain=True)),
            {"id": "project-1"},
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/sql/explain",
        headers=_auth_headers(),
        json_body={"sql": "SELECT * FROM users", "datasource_id": "ds-1"},
    )

    assert response.status_code == 202
    payload = response.json()
    assert set(payload) == {"job_id", "result_set_id"}
    job = services.job_backend.enqueued[0]
    assert job.kind is JobKind.SQL_EXPLAIN
    assert job.payload["result_set_id"] == payload["result_set_id"]
    assert job.payload["datasource_id"] == "ds-1"
    assert "params" not in job.payload
    assert any(audit["action"] == "sql_explain" for audit in services.audits)


def test_sql_explain_rejects_params_instead_of_ignoring_them() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(operation_policy=_policy(allow_explain=True)),
            {"id": "project-1"},
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/sql/explain",
        headers=_auth_headers(),
        json_body={
            "sql": "SELECT * FROM users",
            "datasource_id": "ds-1",
            "params": {"id": 1},
        },
    )

    assert response.status_code == 400
    assert response.json()["error"] == "unsupported_sql_params"
    assert services.job_backend.enqueued == []


def test_compare_task_create_persists_explicit_rules_contract() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            [
                {"id": "ds-source", "project_id": "project-1"},
                {"id": "ds-target", "project_id": "project-1"},
            ],
            _compare_task_row(),
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/compare/tasks",
        headers=_auth_headers(),
        json_body={
            "project_id": "project-1",
            "name": "orders",
            "source_id": "ds-source",
            "target_id": "ds-target",
            "source_ref": {"kind": "table", "schema_name": "app", "table_name": "orders_a"},
            "target_ref": {"kind": "table", "schema_name": "app", "table_name": "orders_b"},
            "columns": [
                {"name": "id", "type": "integer"},
                {"name": "amount", "type": "decimal"},
            ],
            "compare_rules": {
                "key_columns": ["id"],
                "numeric_tolerance": 0.01,
                "schema_policy": "strict",
            },
            "run_limits": {"recursive_checksum": True, "bisection_factor": 8},
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["id"] == "task-1"
    assert payload["compare_rules"]["key_columns"] == ["id"]
    assert payload["compare_rules"]["schema_policy"] == "strict"
    assert payload["run_limits"]["bisection_factor"] == 8
    assert any("INSERT INTO compare_tasks" in statement for statement in engine.statements)
    assert any(audit["action"] == "compare_task_create" for audit in services.audits)


def test_compare_task_clone_copies_config_with_deduped_name() -> None:
    cloned_row = _compare_task_row()
    cloned_row["id"] = "task-2"
    cloned_row["name"] = "orders (copy)"
    engine = _FakeEngine(
        [
            _compare_task_row(),  # 源任务(归属校验前查询)
            {"id": "project-1"},  # 项目授权
            [{"name": "orders"}],  # 项目内已有任务名(去重用)
            cloned_row,  # 插入后回读
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/compare/tasks/task-1/clone",
        headers=_auth_headers(),
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["id"] == "task-2"
    assert payload["name"] == "orders (copy)"
    assert payload["compare_rules"]["key_columns"] == ["id"]
    assert payload["run_limits"]["bisection_factor"] == 8
    insert_statement = next(
        statement for statement in engine.statements if "INSERT INTO compare_tasks" in statement
    )
    assert insert_statement
    assert any(audit["action"] == "compare_task_clone" for audit in services.audits)


def test_compare_task_runs_lists_history_with_pagination() -> None:
    engine = _FakeEngine(
        [
            _compare_task_row(),
            {"id": "project-1"},
            [
                {
                    "run_id": "run-2",
                    "job_id": "job-2",
                    "status": "success",
                    "created_at": _dt(2),
                    "finished_at": _dt(3),
                    "bucket_counts": {"diff": 2, "same": 10},
                    "sample_result": None,
                },
                {
                    "run_id": "run-1",
                    "job_id": "job-1",
                    "status": "failed",
                    "created_at": _dt(1),
                    "finished_at": None,
                    "bucket_counts": {},
                    "sample_result": {"mode": "quick_check"},
                },
            ],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/compare/tasks/task-1/runs",
        headers=_auth_headers(),
        params={"limit": 1, "offset": 0},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["task_id"] == "task-1"
    assert payload["has_more"] is True  # limit+1 探测:取回 2 行说明还有下一页
    assert len(payload["runs"]) == 1
    assert payload["runs"][0]["run_id"] == "run-2"
    assert payload["runs"][0]["bucket_counts"] == {"diff": 2, "same": 10}
    assert payload["runs"][0]["sampled"] is False


def test_compare_preview_returns_bounded_rows(monkeypatch: pytest.MonkeyPatch) -> None:
    captured_sql: list[str] = []

    class _PreviewAdapter:
        def __init__(self, column_sink: Callable[[list[Column]], None]) -> None:
            self._column_sink = column_sink

        def execute_select(self, sql: str, params: dict[str, object]) -> Iterator[Row]:
            del params
            captured_sql.append(sql)
            self._column_sink(
                [
                    Column(name="id", type=ColumnType.INTEGER),
                    Column(name="name", type=ColumnType.STRING),
                ]
            )
            yield Row(values=[1, "alice"])
            yield Row(values=[2, "bob"])
            yield Row(values=[3, "carol"])  # limit+1 行 → truncated

    monkeypatch.setattr(
        core_routes,
        "build_database_adapter",
        lambda conn_info, secret_store, **kwargs: _PreviewAdapter(kwargs["column_sink"]),
    )
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _datasource_row(operation_policy=_policy(allow_select=True)),
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/preview",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "ref": {"kind": "table", "schema_name": "app", "table_name": "orders"},
            "limit": 2,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["columns"] == ["id", "name"]
    assert payload["rows"] == [[1, "alice"], [2, "bob"]]
    assert payload["row_count"] == 2
    assert payload["truncated"] is True
    # mysql 反引号 + SQL 侧 limit+1
    assert captured_sql == ["SELECT * FROM `app`.`orders` LIMIT 3"]
    assert any(audit["action"] == "compare_preview" for audit in services.audits)


def test_compare_preview_generates_alias_metadata_for_unaliased_expression(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured_sql: list[str] = []

    class _PreviewAdapter:
        def __init__(self, column_sink: Callable[[list[Column]], None]) -> None:
            self._column_sink = column_sink

        def execute_select(self, sql: str, params: dict[str, object]) -> Iterator[Row]:
            del params
            captured_sql.append(sql)
            self._column_sink(
                [
                    Column(name="CUST_NO", type=ColumnType.STRING),
                    Column(name="RESULT_1", type=ColumnType.DECIMAL),
                ]
            )
            yield Row(values=["C1", 3])

    monkeypatch.setattr(
        core_routes,
        "build_database_adapter",
        lambda conn_info, secret_store, **kwargs: _PreviewAdapter(kwargs["column_sink"]),
    )
    datasource_row = _datasource_row(operation_policy=_policy(allow_select=True))
    datasource_row["db_type"] = "db2"
    datasource_row["port"] = 50000
    engine = _FakeEngine([{"id": "project-1"}, datasource_row])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/preview",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "ref": {
                "kind": "sql",
                "sql": "SELECT CUST_NO, SUM(AMOUNT) FROM T GROUP BY CUST_NO",
            },
        },
    )

    assert response.status_code == 200
    assert "AS RESULT_1" in captured_sql[0].upper()
    assert response.json()["columns"] == ["CUST_NO", "RESULT_1"]
    assert response.json()["column_details"][1] == {
        "name": "RESULT_1",
        "generated": True,
        "projection_index": 2,
        "expression": "SUM(AMOUNT)",
    }


def test_compare_preview_rejects_non_readonly_sql() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _datasource_row(operation_policy=_policy(allow_select=True)),
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/preview",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "ref": {"kind": "sql", "sql": "DELETE FROM app.orders"},
        },
    )

    assert response.status_code == 400
    assert response.json()["error"] == "invalid_sql"


def test_compare_preview_denied_without_allow_select() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _datasource_row(operation_policy=_policy(allow_select=False)),
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/preview",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "ref": {"kind": "table", "schema_name": "app", "table_name": "orders"},
        },
    )

    assert response.status_code == 403
    assert response.json()["error"] == "select_not_allowed"


def test_compare_pk_precheck_reports_duplicates_and_nulls(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured_sql: list[str] = []

    class _PrecheckAdapter:
        def execute_select(self, sql: str, params: dict[str, object]) -> Iterator[Row]:
            del params
            captured_sql.append(sql)
            if "DISTINCT" in sql:
                yield Row(values=[7])  # distinct_pk_rows
            else:
                yield Row(values=[10, 2])  # total_rows, null_pk_rows

    monkeypatch.setattr(
        core_routes,
        "build_database_adapter",
        lambda conn_info, secret_store, **kwargs: _PrecheckAdapter(),
    )
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _datasource_row(operation_policy=_policy(allow_select=True)),
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/pk-precheck",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "ref": {"kind": "table", "schema_name": "app", "table_name": "orders"},
            "key_columns": ["id"],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload == {
        "total_rows": 10,
        "distinct_pk_rows": 7,
        "duplicate_pk_rows": 3,
        "null_pk_rows": 2,
        "has_duplicates": True,
        "has_nulls": True,
    }
    # mysql 反引号 + 两条聚合(总数/空值、去重)
    assert any("COUNT(*)" in sql and "`id` IS NULL" in sql for sql in captured_sql)
    assert any("DISTINCT `id`" in sql for sql in captured_sql)
    assert any(audit["action"] == "compare_pk_precheck" for audit in services.audits)


def test_compare_pk_precheck_uses_generated_alias_for_sql_expression(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured_sql: list[str] = []

    class _PrecheckAdapter:
        def execute_select(self, sql: str, params: dict[str, object]) -> Iterator[Row]:
            del params
            captured_sql.append(sql)
            yield Row(values=[1]) if "DISTINCT" in sql else Row(values=[1, 0])

    monkeypatch.setattr(
        core_routes,
        "build_database_adapter",
        lambda conn_info, secret_store, **kwargs: _PrecheckAdapter(),
    )
    datasource_row = _datasource_row(operation_policy=_policy(allow_select=True))
    datasource_row["db_type"] = "db2"
    datasource_row["port"] = 50000
    app = create_app(
        services=cast(
            ApiServices,
            _Services(_FakeEngine([{"id": "project-1"}, datasource_row])),
        )
    )

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/pk-precheck",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "ref": {"kind": "sql", "sql": "SELECT SUM(AMOUNT) FROM T"},
            "key_columns": ["RESULT_1"],
        },
    )

    assert response.status_code == 200
    assert len(captured_sql) == 2
    assert all("SUM(AMOUNT) AS RESULT_1" in sql.upper() for sql in captured_sql)


def test_compare_pk_precheck_rejects_file_source() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _datasource_row(operation_policy=_policy(allow_select=True)),
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/pk-precheck",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "ref": {"kind": "file", "upload_id": "up-1", "file_format": "csv"},
            "key_columns": ["id"],
        },
    )

    assert response.status_code == 400
    assert response.json()["error"] == "precheck_unsupported_file"


def test_compare_diff_sql_contract_builds_where_in_for_both_sides() -> None:
    run_row = {
        "run_id": "run-1",
        "job_id": "job-1",
        "project_id": "project-1",
        "task_id": "task-1",
        "bucket_spools": {"diff": "rs-diff"},
    }
    result_store = _ResultStore(
        rows=[
            encode_compare_result_row(
                pk={"id": 3},
                source={"id": 3, "amount": "10.00"},
                target={"id": 3, "amount": "11.00"},
                cells=[{"column": "amount", "source": "10.00", "target": "11.00"}],
            ),
            encode_compare_result_row(
                pk={"id": 4},
                source={"id": 4},
                target={"id": 4},
                cells=[],
            ),
        ]
    )
    engine = _FakeEngine(
        [
            run_row,  # _compare_run_for_current_user -> run_index
            {"id": "project-1"},  # _require_project_access
            _compare_task_row(),  # _compare_task_for_current_user -> compare_tasks
            {"id": "project-1"},  # _require_project_access
            {"db_type": "mysql"},  # source db_type
            {"db_type": "mysql"},  # target db_type
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine, result_store=result_store)))

    response = AsgiClient(app).get(
        "/api/compare/runs/run-1/diff-sql",
        headers=_auth_headers(),
        params={"bucket": "diff"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["bucket"] == "diff"
    assert payload["key_columns"] == ["id"]
    assert payload["pk_count"] == 2
    assert payload["truncated"] is False
    assert payload["source"] == {
        "available": True,
        "sql": "SELECT * FROM `app`.`orders_a` WHERE `id` IN (3, 4)",
        "reason": None,
    }
    assert payload["target"] == {
        "available": True,
        "sql": "SELECT * FROM `app`.`orders_b` WHERE `id` IN (3, 4)",
        "reason": None,
    }


def test_compare_infer_contract_returns_mapping_confidence_and_pk_draft() -> None:
    source_row = _datasource_row()
    source_row["id"] = "ds-source"
    target_row = _datasource_row()
    target_row["id"] = "ds-target"
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            [source_row, target_row],
            _metadata_cache(
                [
                    {
                        "name": "id",
                        "type": "integer",
                        "driver_type": "INT",
                        "nullable": False,
                        "primary_key": True,
                        "comment": None,
                    },
                    {
                        "name": "amount",
                        "type": "decimal",
                        "driver_type": "DECIMAL(12,2)",
                        "nullable": True,
                        "primary_key": False,
                        "comment": None,
                    },
                ]
            ),
            _metadata_cache(
                [
                    {
                        "name": "ID",
                        "type": "integer",
                        "driver_type": "INT",
                        "nullable": False,
                        "primary_key": True,
                        "comment": None,
                    },
                    {
                        "name": "amount",
                        "type": "decimal",
                        "driver_type": "DECIMAL(12,2)",
                        "nullable": True,
                        "primary_key": False,
                        "comment": None,
                    },
                ]
            ),
            _metadata_cache(
                [{"name": "PRIMARY", "columns": ["id"], "is_unique": True, "is_primary": True}]
            ),
            _metadata_cache(
                [{"name": "PRIMARY", "columns": ["ID"], "is_unique": True, "is_primary": True}]
            ),
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/infer",
        headers=_auth_headers(),
        json_body={
            "source_id": "ds-source",
            "target_id": "ds-target",
            "source_table": {"schema_name": "app", "table_name": "orders"},
            "target_table": {"schema_name": "app", "table_name": "orders"},
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["needs_manual_pk"] is False
    assert payload["pk_candidates"][0]["source_columns"] == ["id"]
    assert payload["pk_candidates"][0]["target_columns"] == ["ID"]
    assert payload["pk_candidates"][0]["reason"] == "primary_key"
    id_mapping = next(item for item in payload["mappings"] if item["source_column"] == "id")
    assert id_mapping["target_column"] == "ID"
    assert id_mapping["confidence"] > 0.8
    assert id_mapping["reason"] == "exact"
    assert id_mapping["conflict"] is False
    assert payload["compare_rules"]["key_columns"] == ["id"]
    assert payload["compare_rules"]["column_mappings"]["id"] == "ID"
    assert payload["columns"] == [
        {
            "name": "amount",
            "type": "decimal",
            "driver_type": "DECIMAL(12,2)",
            "nullable": True,
            "primary_key": False,
            "comment": None,
        }
    ]


def test_compare_suggest_tasks_contract_returns_normalized_table_pairs() -> None:
    source_row = _datasource_row()
    source_row["id"] = "ds-source"
    target_row = _datasource_row()
    target_row["id"] = "ds-target"
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            [source_row, target_row],
            _metadata_cache([{"name": "src"}]),
            _metadata_cache(
                [
                    {"schema_name": "src", "name": "orders", "table_type": "BASE TABLE"},
                    {"schema_name": "src", "name": "src_customer", "table_type": "BASE TABLE"},
                ]
            ),
            _metadata_cache([{"name": "dst"}]),
            _metadata_cache(
                [
                    {"schema_name": "dst", "name": "orders", "table_type": "BASE TABLE"},
                    {"schema_name": "dst", "name": "customer", "table_type": "BASE TABLE"},
                ]
            ),
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/compare/suggest-tasks",
        headers=_auth_headers(),
        params={"source_id": "ds-source", "target_id": "ds-target"},
    )

    assert response.status_code == 200
    suggestions = response.json()["suggestions"]
    assert suggestions[0]["source_table"] == "orders"
    assert suggestions[0]["target_table"] == "orders"
    assert suggestions[0]["reason"] == "exact"
    assert any(
        item["source_table"] == "src_customer"
        and item["target_table"] == "customer"
        and item["reason"] == "normalized"
        for item in suggestions
    )


def test_compare_draft_task_contract_reuses_compare_task_create() -> None:
    source_row = _datasource_row()
    source_row["id"] = "ds-source"
    target_row = _datasource_row()
    target_row["id"] = "ds-target"
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            [source_row, target_row],
            _metadata_cache(
                [
                    {
                        "name": "id",
                        "type": "integer",
                        "driver_type": "INT",
                        "nullable": False,
                        "primary_key": True,
                        "comment": None,
                    },
                    {
                        "name": "amount",
                        "type": "decimal",
                        "driver_type": "DECIMAL(12,2)",
                        "nullable": True,
                        "primary_key": False,
                        "comment": None,
                    },
                ]
            ),
            _metadata_cache(
                [
                    {
                        "name": "ID",
                        "type": "integer",
                        "driver_type": "INT",
                        "nullable": False,
                        "primary_key": True,
                        "comment": None,
                    },
                    {
                        "name": "amount",
                        "type": "decimal",
                        "driver_type": "DECIMAL(12,2)",
                        "nullable": True,
                        "primary_key": False,
                        "comment": None,
                    },
                ]
            ),
            _metadata_cache(
                [{"name": "PRIMARY", "columns": ["id"], "is_unique": True, "is_primary": True}]
            ),
            _metadata_cache(
                [{"name": "PRIMARY", "columns": ["ID"], "is_unique": True, "is_primary": True}]
            ),
            {"id": "project-1"},
            [
                {"id": "ds-source", "project_id": "project-1"},
                {"id": "ds-target", "project_id": "project-1"},
            ],
            _compare_task_row(),
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/draft-task",
        headers=_auth_headers(),
        json_body={
            "name": "orders draft",
            "source_id": "ds-source",
            "target_id": "ds-target",
            "source_table": {"schema_name": "app", "table_name": "orders_a"},
            "target_table": {"schema_name": "app", "table_name": "orders_b"},
        },
    )

    assert response.status_code == 201
    assert response.json()["id"] == "task-1"
    assert any("INSERT INTO compare_tasks" in statement for statement in engine.statements)
    assert any(audit["action"] == "compare_task_draft_create" for audit in services.audits)


def test_compare_task_run_enqueues_compare_job_and_run_index() -> None:
    engine = _FakeEngine([_compare_task_row(), {"id": "project-1"}])
    job_backend = _JobBackend()
    services = _Services(engine, job_backend=job_backend)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/compare/tasks/task-1/run",
        headers=_auth_headers(),
    )

    assert response.status_code == 202
    payload = response.json()
    assert set(payload) == {"job_id", "run_id"}
    job = job_backend.enqueued[0]
    assert job.kind is JobKind.COMPARE_RUN
    assert job.timeout_seconds == 1800
    assert job.datasource_ids == ["ds-source", "ds-target"]
    assert set(job.payload["bucket_result_set_ids"]) == {
        "only_source",
        "only_target",
        "diff",
        "same",
    }
    assert any("INSERT INTO run_index" in statement for statement in engine.statements)
    assert any(audit["action"] == "compare_run" for audit in services.audits)


def test_compare_sql_task_response_includes_projection_details() -> None:
    task_row = _sql_compare_task_row()
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            [
                {"id": "ds-source", "project_id": "project-1"},
                {"id": "ds-target", "project_id": "project-1"},
            ],
            task_row,
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/compare/tasks",
        headers=_auth_headers(),
        json_body={
            "project_id": "project-1",
            "name": "aggregate compare",
            "source_id": "ds-source",
            "target_id": "ds-target",
            "source_ref": task_row["source_ref"],
            "target_ref": task_row["target_ref"],
            "columns": task_row["columns"],
            "compare_rules": task_row["compare_rules"],
        },
    )

    assert response.status_code == 201
    payload = response.json()
    source_ref = cast(dict[str, object], task_row["source_ref"])
    assert payload["source_ref"]["sql"] == source_ref["sql"]
    assert payload["source_projection_details"][-1]["name"] == "RESULT_1"
    assert payload["target_projection_details"][-1]["generated"] is True


def test_compare_task_run_rejects_proven_legacy_numeric_aliases() -> None:
    task_row = _sql_compare_task_row(legacy_numeric=True)
    job_backend = _JobBackend()
    app = create_app(
        services=cast(
            ApiServices,
            _Services(
                _FakeEngine([task_row, {"id": "project-1"}]),
                job_backend=job_backend,
            ),
        )
    )

    response = AsgiClient(app).post(
        "/api/compare/tasks/task-1/run",
        headers=_auth_headers(),
    )

    assert response.status_code == 409
    assert response.json()["error"] == "compare_sql_aliases_stale"
    assert job_backend.enqueued == []


def test_compare_task_run_allows_explicit_numeric_alias() -> None:
    task_row = _sql_compare_task_row(explicit_numeric=True)
    job_backend = _JobBackend()
    app = create_app(
        services=cast(
            ApiServices,
            _Services(
                _FakeEngine([task_row, {"id": "project-1"}]),
                job_backend=job_backend,
            ),
        )
    )

    response = AsgiClient(app).post(
        "/api/compare/tasks/task-1/run",
        headers=_auth_headers(),
    )

    assert response.status_code == 202
    assert len(job_backend.enqueued) == 1


def test_compare_run_result_contract_returns_bucket_counts_and_cell_diffs() -> None:
    diff_profile = _diff_profile_payload()
    sample_result = _sample_result_payload()
    run_row = {
        "run_id": "run-1",
        "job_id": "job-1",
        "project_id": "project-1",
        "bucket_spools": {"diff": "rs-diff"},
        "bucket_counts": {
            "only_source": 1,
            "only_target": 1,
            "diff": 1,
            "same": 7,
        },
        "progress": {
            "scanned_segments": 2,
            "skipped_segments": 1,
            "skipped_rows": 7,
            "row_mode_segments": 1,
        },
        "diff_profile": diff_profile,
        "sample_result": sample_result,
    }
    result_store = _ResultStore(
        rows=[
            encode_compare_result_row(
                pk={"id": 3},
                source={"id": 3, "amount": "10.00"},
                target={"id": 3, "amount": "11.00"},
                cells=[{"column": "amount", "source": "10.00", "target": "11.00"}],
            )
        ]
    )
    engine = _FakeEngine([run_row, {"id": "project-1"}])
    app = create_app(services=cast(ApiServices, _Services(engine, result_store=result_store)))

    response = AsgiClient(app).get(
        "/api/compare/runs/run-1/results",
        headers=_auth_headers(),
        params={"bucket": "diff", "offset": 0, "limit": 10},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["bucket_counts"] == {
        "only_source": 1,
        "only_target": 1,
        "diff": 1,
        "same": 7,
    }
    assert payload["progress"]["skipped_rows"] == 7
    assert payload["diff_profile"] == diff_profile
    assert payload["sample_result"] == sample_result
    assert payload["rows"] == [
        {
            "pk": {"id": 3},
            "source": {"id": 3, "amount": "10.00"},
            "target": {"id": 3, "amount": "11.00"},
            "cells": [{"column": "amount", "source": "10.00", "target": "11.00"}],
        }
    ]


def test_compare_run_profile_contract_returns_profile_without_rows() -> None:
    diff_profile = _diff_profile_payload()
    sample_result = _sample_result_payload()
    run_row = {
        "run_id": "run-1",
        "job_id": "job-1",
        "project_id": "project-1",
        "bucket_spools": {},
        "bucket_counts": {"only_source": 0, "only_target": 0, "diff": 1, "same": 10},
        "progress": {"row_mode_segments": 1},
        "diff_profile": diff_profile,
        "sample_result": sample_result,
    }
    app = create_app(
        services=cast(
            ApiServices,
            _Services(_FakeEngine([run_row, {"id": "project-1"}])),
        )
    )

    response = AsgiClient(app).get("/api/compare/runs/run-1/profile", headers=_auth_headers())

    assert response.status_code == 200
    payload = response.json()
    assert payload["run_id"] == "run-1"
    assert payload["diff_profile"] == diff_profile
    assert payload["sample_result"] == sample_result
    assert "rows" not in payload


def test_compare_ai_attribution_uses_gateway_l2_profile() -> None:
    run_row = {
        "run_id": "run-1",
        "job_id": "job-1",
        "project_id": "project-1",
        "bucket_spools": {},
        "bucket_counts": {"only_source": 0, "only_target": 0, "diff": 1, "same": 10},
        "progress": {},
        "diff_profile": _diff_profile_payload(),
        "sample_result": None,
    }
    ai_row = {
        "enabled": True,
        "provider": "mock",
        "model": "mock-model",
        "base_url": None,
        "api_key_secret_ref": None,
        "max_auto_egress_level": 2,
        "l4_requires_optin": True,
        "enable_inference": True,
    }
    services = _Services(_FakeEngine([run_row, {"id": "project-1"}, ai_row]))
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/compare/runs/run-1/ai-attribution",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["attribution"] == "ok"
    assert payload["egress_level"] == 2
    assert any(audit["action"] == "compare_diff_attribution" for audit in services.audits)


def test_compare_ai_attribution_disabled_degrades_without_profile_loss() -> None:
    run_row = {
        "run_id": "run-1",
        "job_id": "job-1",
        "project_id": "project-1",
        "bucket_spools": {},
        "bucket_counts": {"only_source": 0, "only_target": 0, "diff": 1, "same": 10},
        "progress": {},
        "diff_profile": _diff_profile_payload(),
        "sample_result": None,
    }
    services = _Services(_FakeEngine([run_row, {"id": "project-1"}, None]))
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/compare/runs/run-1/ai-attribution",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    assert response.json()["error"] == "ai_disabled"
    assert response.json()["ok"] is False


def _ai_map_ai_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "enabled": True,
        "provider": "mock",
        "model": "mock-model",
        "base_url": None,
        "api_key_secret_ref": None,
        "max_auto_egress_level": 2,
        "l4_requires_optin": True,
        "enable_inference": True,
    }
    row.update(overrides)
    return row


def _ai_map_columns_cache(names_types: list[tuple[str, str]]) -> dict[str, object]:
    return _metadata_cache(
        [
            {
                "name": name,
                "type": ctype,
                "driver_type": ctype.upper(),
                "nullable": True,
                "primary_key": name == "id",
                "comment": None,
            }
            for name, ctype in names_types
        ]
    )


def _ai_map_datasource_rows() -> tuple[dict[str, object], dict[str, object]]:
    source_row = _datasource_row()
    source_row["id"] = "ds-source"
    target_row = _datasource_row()
    target_row["id"] = "ds-target"
    return source_row, target_row


def _ai_map_body(*, include_samples: bool) -> dict[str, object]:
    return {
        "source_id": "ds-source",
        "target_id": "ds-target",
        "source_table": {"schema_name": "app", "table_name": "orders_a"},
        "target_table": {"schema_name": "app", "table_name": "orders_b"},
        "confirmed_mappings": {"id": "id", "amount": "amount"},
        "include_samples": include_samples,
    }


def test_compare_ai_map_suggest_l2_path_with_mock_provider_and_no_history() -> None:
    source_row, target_row = _ai_map_datasource_rows()
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            [source_row, target_row],
            _ai_map_ai_row(),
            _ai_map_columns_cache([("id", "integer"), ("amount", "decimal"), ("note", "string")]),
            _ai_map_columns_cache([("id", "integer"), ("amount", "decimal"), ("memo", "string")]),
            [],  # no prior compare tasks → no mapping history
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/ai-map-suggest",
        headers=_auth_headers(),
        json_body=_ai_map_body(include_samples=False),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["egress_level"] == 2  # L2 only, no samples
    assert payload["samples_included"] is False
    assert payload["history_available"] is False
    assert payload["residual_source_columns"] == ["note"]
    assert payload["residual_target_columns"] == ["memo"]
    # MockProvider 回 "ok" 不可解析为映射 JSON → 空建议但成功(优雅降级)。
    assert payload["suggestions"] == []
    audit = next(a for a in services.audits if a["action"] == "ai_assist_call")
    assert audit["result"] == "success"
    # ★ R5:审计 detail 只记计数/等级/是否含样本,绝不含建议内容 / 样本值 / prompt 原文。
    detail = cast(dict[str, object], audit["detail"])
    assert detail["samples_included"] is False
    assert detail["egress_level"] == 2
    assert "suggestions" not in detail
    assert "prompt" not in detail
    assert "content" not in detail


def test_compare_ai_map_suggest_disabled_returns_409() -> None:
    source_row, target_row = _ai_map_datasource_rows()
    engine = _FakeEngine([{"id": "project-1"}, [source_row, target_row], None])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/ai-map-suggest",
        headers=_auth_headers(),
        json_body=_ai_map_body(include_samples=False),
    )

    assert response.status_code == 409
    assert response.json()["error"] == "ai_disabled"


def test_compare_ai_map_suggest_samples_without_optin_rejected_403() -> None:
    source_row, target_row = _ai_map_datasource_rows()
    # l4_requires_optin=True(默认/portable) + include_samples=True → 拒发样本,取数之前即闸门。
    engine = _FakeEngine(
        [{"id": "project-1"}, [source_row, target_row], _ai_map_ai_row(l4_requires_optin=True)]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/ai-map-suggest",
        headers=_auth_headers(),
        json_body=_ai_map_body(include_samples=True),
    )

    assert response.status_code == 403
    assert response.json()["error"] == "l4_optin_required"
    audit = next(a for a in services.audits if a["action"] == "ai_assist_call")
    assert audit["result"] == "skipped"


class _SequenceGateway:
    name = "sequence"

    def __init__(
        self,
        responses: list[AiResponse] | None = None,
        error: Exception | None = None,
    ) -> None:
        self.responses = list(responses or [])
        self.error = error
        self.calls: list[tuple[str, AiContext, AiOptions]] = []

    def complete(self, prompt: str, context: AiContext, options: AiOptions) -> AiResponse:
        self.calls.append((prompt, context, options))
        if self.error is not None:
            raise self.error
        return self.responses.pop(0)


def test_ai_sql_generate_uses_l2_schema_context_and_audits(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    columns_cache = _metadata_cache(
        [
            {
                "name": "id",
                "type": "integer",
                "driver_type": "INT",
                "nullable": False,
                "primary_key": True,
                "comment": "pk",
            },
            {
                "name": "amount",
                "type": "decimal",
                "driver_type": "DECIMAL(10,2)",
                "nullable": True,
                "primary_key": False,
                "comment": None,
            },
        ]
    )
    engine = _FakeEngine(
        [
            _datasource_row(),  # _datasource_for_current_user: datasources select
            {"id": "project-1"},  # _require_project_access
            _ai_config_row(),  # _ai_config_row_or_none
            columns_cache,  # _metadata_columns_for_table cache read (no adapter)
        ]
    )
    services = _Services(engine)
    gateway = _SequenceGateway(
        [AiResponse(content="SELECT id, amount FROM app.users", provider="mock", model="m")]
    )
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers={**_auth_headers(), "X-Request-ID": "req-ai-sql-1"},
        json_body={
            "natural_language": "sum of amount",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["sql"] == "SELECT id, amount FROM app.users"
    assert payload["egress_level"] == 2
    assert payload["tables_used"] == ["app.users"]
    assert payload["truncated"] is False
    assert payload["request_id"] == "req-ai-sql-1"
    assert any(audit["action"] == "ai_copilot_run" for audit in services.audits)
    # No datasource adapter was built (metadata came from cache) and no row values leaked.
    body = response.body.decode("utf-8")
    assert "password" not in body


def test_ai_sql_generate_validates_preview_without_execution(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    gateway = _SequenceGateway(
        [AiResponse(content="SELECT id, name FROM app.users", provider="mock", model="m")]
    )
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list user names",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    payload = response.json()
    assert response.status_code == 200
    assert payload["ok"] is True
    assert payload["stage"] == "validated"
    assert payload["attempts"] == 1
    assert payload["reasoning_mode"] == "disabled"
    assert payload["validation"] == {
        "readonly": "passed",
        "tables": "passed",
        "columns": "passed",
        "warnings": [],
    }
    assert gateway.calls[0][1].items[0].egress_level is EgressLevel.L2
    assert gateway.calls[0][2].max_tokens == 1200
    assert not any(audit["action"] == "sql_execute" for audit in services.audits)
    assert not any("INSERT INTO jobs" in statement for statement in services.engine.statements)


def test_ai_sql_generate_repairs_truncated_output_once(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    gateway = _SequenceGateway(
        [
            AiResponse(
                content="",
                finish_reason="length",
                reasoning_chars=100,
                provider="first-provider",
                model="first-model",
                tokens_in=10,
                tokens_out=20,
                tokens_total=30,
                duration_ms=12,
            ),
            AiResponse(
                content="SELECT id FROM app.users",
                finish_reason="stop",
                provider="final-provider",
                model="final-model",
                tokens_in=2,
                tokens_out=3,
                tokens_total=5,
                duration_ms=8,
            ),
        ]
    )
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())
    response = AsgiClient(create_app(services=cast(ApiServices, services))).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list users",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.json()["ok"] is True
    assert response.json()["attempts"] == 2
    assert len(gateway.calls) == 2
    first_max_tokens = gateway.calls[0][2].max_tokens
    second_max_tokens = gateway.calls[1][2].max_tokens
    assert first_max_tokens is not None
    assert second_max_tokens is not None
    assert second_max_tokens > first_max_tokens
    audits = [audit for audit in services.audits if audit["action"] == "ai_copilot_run"]
    assert len(audits) == 1
    detail = audits[0]["detail"]
    assert isinstance(detail, dict)
    assert detail["provider_duration_ms"] == 20
    assert detail["tokens_in"] == 12
    assert detail["tokens_out"] == 23
    assert detail["tokens_total"] == 35
    assert detail["provider"] == "final-provider"
    assert detail["model"] == "final-model"
    assert detail["attempts"] == 2


@pytest.mark.parametrize(
    ("first_response", "diagnostic_code"),
    [
        (
            AiResponse(content="", finish_reason="length"),
            "provider_output_truncated",
        ),
        (
            AiResponse(content="", finish_reason="stop", reasoning_chars=10),
            "provider_reasoning_only",
        ),
        (
            AiResponse(content="", finish_reason="stop"),
            "provider_invalid_response",
        ),
    ],
)
def test_ai_sql_generate_empty_repair_retains_original_intent_without_l3_candidate(
    monkeypatch: pytest.MonkeyPatch,
    first_response: AiResponse,
    diagnostic_code: str,
) -> None:
    original_intent = "retain customer intent marker"
    gateway = _SequenceGateway([first_response, AiResponse(content="SELECT id FROM app.users")])
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())

    response = AsgiClient(create_app(services=cast(ApiServices, services))).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": original_intent,
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.json()["ok"] is True
    assert response.json()["attempts"] == 2
    assert len(gateway.calls) == 2
    repair_prompt, repair_context, _ = gateway.calls[1]
    assert original_intent in repair_prompt
    assert diagnostic_code in repair_prompt
    assert len(repair_context.items) == 1
    assert repair_context.items[0].egress_level is EgressLevel.L2


def test_ai_sql_generate_stops_after_one_failed_repair(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    gateway = _SequenceGateway(
        [
            AiResponse(content="", finish_reason="length"),
            AiResponse(content="", finish_reason="length"),
        ]
    )
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())

    response = AsgiClient(create_app(services=cast(ApiServices, services))).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list users",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.json()["ok"] is False
    assert response.json()["diagnostic_code"] == "provider_output_truncated"
    assert response.json()["attempts"] == 2
    assert len(gateway.calls) == 2


@pytest.mark.parametrize(
    "code",
    [
        "provider_auth_failed",
        "provider_rate_limited",
        "provider_timeout",
        "provider_unreachable",
    ],
)
def test_ai_sql_generate_does_not_retry_deterministic_provider_failures(
    monkeypatch: pytest.MonkeyPatch,
    code: str,
) -> None:
    gateway = _SequenceGateway(error=ProviderError(code))
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())
    response = AsgiClient(create_app(services=cast(ApiServices, services))).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list users",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.json()["diagnostic_code"] == code
    assert response.json()["attempts"] == 1
    assert len(gateway.calls) == 1


def test_ai_sql_generate_repairs_invalid_provider_response_once(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    gateway = _SequenceGateway(error=ProviderError("provider_invalid_response"))
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())

    response = AsgiClient(create_app(services=cast(ApiServices, services))).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list users",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.json()["diagnostic_code"] == "provider_invalid_response"
    assert response.json()["attempts"] == 2
    assert len(gateway.calls) == 2


@pytest.mark.parametrize(
    ("error", "diagnostic_code"),
    [
        (BudgetExceededError(), "ai_budget_exceeded"),
        (AiGatewayError("opaque gateway failure"), "ai_gateway_failed"),
    ],
)
def test_ai_sql_generate_does_not_retry_gateway_policy_failures(
    monkeypatch: pytest.MonkeyPatch,
    error: AiGatewayError,
    diagnostic_code: str,
) -> None:
    gateway = _SequenceGateway(error=error)
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())

    response = AsgiClient(create_app(services=cast(ApiServices, services))).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list users",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.json()["diagnostic_code"] == diagnostic_code
    assert response.json()["attempts"] == 1
    assert len(gateway.calls) == 1


def test_ai_sql_generate_does_not_retry_egress_failure(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    gateway = _SequenceGateway(error=EgressBlockedError(EgressLevel.L3, EgressLevel.L2))
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())

    response = AsgiClient(create_app(services=cast(ApiServices, services))).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list users",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.json()["diagnostic_code"] == "ai_egress_blocked"
    assert response.json()["attempts"] == 1
    assert len(gateway.calls) == 1


def test_ai_sql_generate_repair_keeps_candidate_out_of_prompt_and_in_l3_context(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    candidate = "SELECT missing FROM app.users"
    gateway = _SequenceGateway(
        [
            AiResponse(content=candidate),
            AiResponse(content="SELECT id FROM app.users"),
        ]
    )
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())

    response = AsgiClient(create_app(services=cast(ApiServices, services))).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list users",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.json()["ok"] is True
    assert len(gateway.calls) == 2
    assert candidate not in gateway.calls[1][0]
    assert [item.egress_level for item in gateway.calls[1][1].items] == [
        EgressLevel.L2,
        EgressLevel.L3,
    ]
    assert gateway.calls[1][1].items[1].content == candidate


def test_ai_sql_generate_repair_l3_context_is_blocked_before_second_provider_call(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    candidate = "SELECT missing FROM app.users"
    provider = _SequenceGateway(
        [
            AiResponse(content=candidate),
            AiResponse(content="SELECT id FROM app.users"),
        ]
    )
    gateway = DefaultAiGateway(provider, max_auto_egress_level=EgressLevel.L2)
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())

    response = AsgiClient(create_app(services=cast(ApiServices, services))).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list users",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.json()["diagnostic_code"] == "ai_egress_blocked"
    assert response.json()["attempts"] == 2
    assert len(provider.calls) == 1


def test_ai_sql_generate_revision_uses_separate_l3_candidate_context(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    candidate = "SELECT id FROM app.users"
    gateway = _SequenceGateway(
        [AiResponse(content="SELECT id, name FROM app.users", provider="mock", model="m")]
    )
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())

    response = AsgiClient(create_app(services=cast(ApiServices, services))).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list users",
            "schema_name": "app",
            "table_names": ["users"],
            "candidate_sql": candidate,
            "revision_instruction": "include names",
        },
    )

    assert response.json()["ok"] is True
    assert response.json()["egress_level"] == 3
    assert [item.egress_level for item in gateway.calls[0][1].items] == [
        EgressLevel.L2,
        EgressLevel.L3,
    ]
    assert gateway.calls[0][1].items[1].content == candidate
    assert candidate not in gateway.calls[0][0]


def test_ai_sql_audit_contains_only_allowlisted_diagnostics(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    gateway = _SequenceGateway(
        [AiResponse(content="SELECT id FROM app.users", provider="mock", model="m")]
    )
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    services = _Services(_ai_sql_engine())
    AsgiClient(create_app(services=cast(ApiServices, services))).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "private prompt marker",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    audits = [audit for audit in services.audits if audit["action"] == "ai_copilot_run"]
    assert len(audits) == 1
    detail = audits[0]["detail"]
    assert isinstance(detail, dict)
    assert set(detail) <= {
        "diagnostic_code",
        "stage",
        "duration_ms",
        "provider_duration_ms",
        "attempts",
        "provider",
        "model",
        "reasoning_mode",
        "table_count",
        "tokens_in",
        "tokens_out",
        "tokens_total",
        "egress_level",
    }
    assert "private prompt marker" not in str(detail)
    assert "users" not in str(detail)
    assert "SELECT" not in str(detail)


def test_ai_sql_table_candidates_rank_without_calling_gateway(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from types import SimpleNamespace

    calls = {"gateway": 0}

    def fail_gateway(*args: object, **kwargs: object) -> None:
        calls["gateway"] += 1
        raise AssertionError("candidate ranking must not call AI")

    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", fail_gateway)
    monkeypatch.setattr(
        core_routes,
        "_metadata_all_tables",
        lambda services, row, refresh: [
            SimpleNamespace(schema_name="app", name="orders"),
            SimpleNamespace(schema_name="app", name="users"),
        ],
    )
    monkeypatch.setattr(
        core_routes,
        "_metadata_columns_for_table",
        lambda services, row, schema_name, table_name, refresh: [
            Column(name="customer_id", type=ColumnType.INTEGER),
            Column(
                name="amount" if table_name == "orders" else "name",
                type=ColumnType.STRING,
            ),
        ],
    )
    engine = _FakeEngine([_datasource_row(), {"id": "project-1"}])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/sql-table-candidates",
        headers=_auth_headers(),
        json_body={
            "natural_language": "customer order amount",
            "schema_name": None,
            "editor_sql": "SELECT * FROM app.users",
        },
    )

    assert response.status_code == 200
    assert response.json()["candidates"][0]["table_name"] == "users"
    assert calls["gateway"] == 0


def test_ai_sql_table_candidates_schema_scope_avoids_all_schema_enumeration(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    metadata_calls: list[tuple[str, str]] = []
    column_calls: list[tuple[str, str]] = []

    def fail_all_tables(*args: object, **kwargs: object) -> None:
        raise AssertionError("schema-scoped candidates must not enumerate all schemas")

    def schema_tables(*args: object, **kwargs: object) -> list[dict[str, object]]:
        metadata_calls.append((str(args[2]), str(kwargs["schema_name"])))
        return [
            {"schema_name": "app", "name": "users", "table_type": "BASE TABLE"},
            {"schema_name": "app", "name": "orders", "table_type": "BASE TABLE"},
        ]

    def columns(
        services: object,
        row: object,
        schema_name: str,
        table_name: str,
        refresh: bool,
    ) -> list[Column]:
        del services, row, refresh
        column_calls.append((schema_name, table_name))
        return [Column(name="id", type=ColumnType.INTEGER)]

    monkeypatch.setattr(core_routes, "_metadata_all_tables", fail_all_tables)
    monkeypatch.setattr(core_routes, "_metadata_payload", schema_tables)
    monkeypatch.setattr(core_routes, "_metadata_columns_for_table", columns)
    engine = _FakeEngine([_datasource_row(), {"id": "project-1"}])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/sql-table-candidates",
        headers=_auth_headers(),
        json_body={"natural_language": "users", "schema_name": "app"},
    )

    assert response.status_code == 200
    assert metadata_calls == [("tables", "app")]
    assert all(item["schema_name"] == "app" for item in response.json()["candidates"])
    assert column_calls == [("app", "users"), ("app", "orders")]


def test_ai_sql_table_candidates_reports_truncated_schema_scope(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    column_calls: list[str] = []

    def fail_all_tables(*args: object, **kwargs: object) -> None:
        raise AssertionError("schema-scoped candidates must not enumerate all schemas")

    monkeypatch.setattr(core_routes, "_metadata_all_tables", fail_all_tables)
    monkeypatch.setattr(
        core_routes,
        "_metadata_payload",
        lambda *args, **kwargs: [
            {
                "schema_name": "app",
                "name": f"table_{index}",
                "table_type": "BASE TABLE",
            }
            for index in range(MAX_TABLES + 1)
        ],
    )

    def columns(
        services: object,
        row: object,
        schema_name: str,
        table_name: str,
        refresh: bool,
    ) -> list[Column]:
        del services, row, schema_name, refresh
        column_calls.append(table_name)
        return [Column(name="id", type=ColumnType.INTEGER)]

    monkeypatch.setattr(core_routes, "_metadata_columns_for_table", columns)
    engine = _FakeEngine([_datasource_row(), {"id": "project-1"}])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/sql-table-candidates",
        headers=_auth_headers(),
        json_body={"natural_language": "table", "schema_name": "app"},
    )

    assert response.status_code == 200
    assert response.json()["truncated"] is True
    assert len(response.json()["candidates"]) == 8
    assert len(column_calls) == MAX_TABLES


def test_ai_sql_table_candidates_rejects_unauthorized_datasource_before_metadata(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_metadata(*args: object, **kwargs: object) -> None:
        raise AssertionError("unauthorized datasource must not read metadata")

    monkeypatch.setattr(core_routes, "_metadata_all_tables", fail_metadata)
    monkeypatch.setattr(core_routes, "_metadata_payload", fail_metadata)
    engine = _FakeEngine([_datasource_row(), None])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/sql-table-candidates",
        headers=_auth_headers(),
        json_body={"natural_language": "users"},
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"


def test_ai_sql_generate_without_schema_uses_only_unique_confirmed_table(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from types import SimpleNamespace

    column_calls: list[tuple[str, str]] = []
    monkeypatch.setattr(
        core_routes,
        "_metadata_all_tables",
        lambda services, row, refresh: [
            SimpleNamespace(schema_name="audit", name="events"),
            SimpleNamespace(schema_name="app", name="users"),
        ],
    )

    def columns(
        services: object,
        row: object,
        schema_name: str,
        table_name: str,
        refresh: bool,
    ) -> list[Column]:
        del services, row, refresh
        column_calls.append((schema_name, table_name))
        return [Column(name="id", type=ColumnType.INTEGER)]

    monkeypatch.setattr(core_routes, "_metadata_columns_for_table", columns)
    gateway = _SequenceGateway([AiResponse(content="SELECT id FROM app.users")])
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", lambda runtime: gateway)
    engine = _FakeEngine([_datasource_row(), {"id": "project-1"}, _ai_config_row()])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={"natural_language": "list users", "table_names": ["users"]},
    )

    assert response.status_code == 200
    assert response.json()["tables_used"] == ["app.users"]
    assert column_calls == [("app", "users")]


def test_ai_sql_generate_without_schema_rejects_ambiguous_confirmed_table(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from types import SimpleNamespace

    def fail_after_scope(*args: object, **kwargs: object) -> None:
        raise AssertionError("ambiguous table scope must stop before columns or gateway")

    monkeypatch.setattr(
        core_routes,
        "_metadata_all_tables",
        lambda services, row, refresh: [
            SimpleNamespace(schema_name="app", name="users"),
            SimpleNamespace(schema_name="audit", name="users"),
        ],
    )
    monkeypatch.setattr(core_routes, "_metadata_columns_for_table", fail_after_scope)
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", fail_after_scope)
    engine = _FakeEngine([_datasource_row(), {"id": "project-1"}, _ai_config_row()])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={"natural_language": "list users", "table_names": ["users"]},
    )

    assert response.status_code == 422
    assert response.json()["error"] == "ai_table_scope_ambiguous"


def test_ai_sql_generate_rejects_empty_confirmed_table_scope(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_after_scope(*args: object, **kwargs: object) -> None:
        raise AssertionError("empty table scope must stop before all downstream work")

    engine = _FakeEngine([_datasource_row(), {"id": "project-1"}, _ai_config_row()])
    services = _Services(engine)
    monkeypatch.setattr(core_routes, "_ai_config_row_or_none", fail_after_scope)
    monkeypatch.setattr(core_routes, "_metadata_all_tables", fail_after_scope)
    monkeypatch.setattr(core_routes, "_metadata_columns_for_table", fail_after_scope)
    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", fail_after_scope)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={"natural_language": "list rows", "schema_name": "app", "table_names": []},
    )

    assert response.status_code == 422
    assert response.json()["error"] == "ai_table_scope_required"
    assert len(engine.results) == 3
    assert not any(audit["action"] == "ai_copilot_run" for audit in services.audits)


def test_sql_generate_response_additive_defaults_preserve_old_construction() -> None:
    payload = SqlGenerateResponse(ok=False, error="ProviderError").model_dump(mode="json")
    assert payload["stage"] == "failed"
    assert payload["diagnostic_code"] is None
    assert payload["attempts"] == 0
    assert payload["validation"] is None


def test_ai_sql_generate_metadata_failure_is_audited(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_gateway(*args: object, **kwargs: object) -> None:
        raise AssertionError("metadata failure must stop before provider access")

    monkeypatch.setattr(core_routes, "build_gateway_from_runtime_config", fail_gateway)
    monkeypatch.setattr(
        core_routes,
        "build_database_adapter",
        lambda conn_info, secret_store, **kwargs: _FailingMetadataAdapter(),
    )
    engine = _FakeEngine(
        [
            _datasource_row(),
            {"id": "project-1"},
            _ai_config_row(),
            None,
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list rows",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.status_code == 503
    assert response.json()["error"] == "metadata_probe_failed"
    audit = next(a for a in services.audits if a["action"] == "ai_copilot_run")
    assert audit["result"] == "failed"
    assert audit["detail"] == {
        "diagnostic_code": "metadata_probe_failed",
        "stage": "failed",
    }


def test_ai_sql_generate_disabled_returns_structured_409() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(),  # datasources select
            {"id": "project-1"},  # _require_project_access
            None,  # no ai_configs row -> AI disabled
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/sql-generate",
        headers=_auth_headers(),
        json_body={
            "natural_language": "list rows",
            "schema_name": "app",
            "table_names": ["users"],
        },
    )

    assert response.status_code == 409
    assert response.json()["error"] == "ai_disabled"
    audit = next(audit for audit in services.audits if audit["action"] == "ai_copilot_run")
    assert audit["result"] == "skipped"
    assert audit["detail"] == {"diagnostic_code": "ai_disabled", "stage": "failed"}


def _ai_config_row_l3() -> dict[str, object]:
    # C4 sends masked SQL (L3); requires admin to allow max_auto_egress_level>=3.
    row = _ai_config_row()
    row["max_auto_egress_level"] = 3
    return row


def _slowsql_columns_cache() -> dict[str, object]:
    return _metadata_cache(
        [
            {
                "name": "id",
                "type": "integer",
                "driver_type": "INT",
                "nullable": False,
                "primary_key": True,
                "comment": "pk",
            },
            {
                "name": "amount",
                "type": "decimal",
                "driver_type": "DECIMAL(10,2)",
                "nullable": True,
                "primary_key": False,
                "comment": None,
            },
        ]
    )


def _slowsql_indexes_cache() -> dict[str, object]:
    return _metadata_cache(
        [{"name": "idx_amount", "columns": ["amount"], "is_unique": False, "is_primary": False}]
    )


def test_slow_sql_diagnose_assembles_l3_context_with_baseline_and_audits() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(),  # _datasource_for_current_user: datasources select
            {"id": "project-1"},  # _require_project_access
            _ai_config_row_l3(),  # _ai_config_row_or_none (max_auto L3)
            _slowsql_columns_cache(),  # _metadata_columns_for_table cache read
            _slowsql_indexes_cache(),  # _metadata_indexes_for_table cache read
            {  # historical baseline aggregate (same sql_hash)
                "runs": 3,
                "avg_seconds": 1.5,
                "min_seconds": 0.9,
                "max_seconds": 3.0,
                "p95_seconds": 2.8,
            },
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/slow-sql-diagnose",
        headers=_auth_headers(),
        json_body={"sql": "SELECT amount FROM orders WHERE name = 'alice' AND amount > 100"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["diagnosis"] == "ok"  # MockProvider echoes "ok"
    assert payload["egress_level"] == 3
    assert payload["plan_included"] is False  # no explain_job_id given -> degrade
    assert payload["tables_analyzed"] == ["app.orders"]
    assert payload["baseline_available"] is True
    assert payload["baseline_runs"] == 3
    assert any(
        audit["action"] == "ai_assist_call" and audit["result"] == "success"
        for audit in services.audits
    )
    # R5 / egress: masked literals must not appear in the response body.
    body = response.body.decode("utf-8")
    assert "alice" not in body


def test_slow_sql_diagnose_degrades_gracefully_without_baseline() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(),
            {"id": "project-1"},
            _ai_config_row_l3(),
            _slowsql_columns_cache(),
            _slowsql_indexes_cache(),
            {  # no historical runs -> early-delivery graceful degradation
                "runs": 0,
                "avg_seconds": None,
                "min_seconds": None,
                "max_seconds": None,
                "p95_seconds": None,
            },
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/slow-sql-diagnose",
        headers=_auth_headers(),
        json_body={"sql": "SELECT amount FROM orders"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["baseline_available"] is False
    assert payload["baseline_runs"] == 0


def test_slow_sql_diagnose_disabled_returns_structured_409() -> None:
    engine = _FakeEngine(
        [
            _datasource_row(),
            {"id": "project-1"},
            None,  # no ai_configs row -> AI disabled
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/datasources/ds-1/ai/slow-sql-diagnose",
        headers=_auth_headers(),
        json_body={"sql": "SELECT amount FROM orders"},
    )

    assert response.status_code == 409
    assert response.json()["error"] == "ai_disabled"
    assert any(
        audit["action"] == "ai_assist_call" and audit["result"] == "skipped"
        for audit in services.audits
    )


def test_create_compare_export_enqueues_four_bucket_export_and_token() -> None:
    bucket_spools: dict[str, str] = {
        "only_source": "rs-only-source",
        "only_target": "rs-only-target",
        "diff": "rs-diff",
        "same": "rs-same",
    }
    run_row: dict[str, object] = {
        "run_id": "run-1",
        "job_id": "job-compare",
        "project_id": "project-1",
        "status": "success",
        "bucket_spools": bucket_spools,
        "bucket_counts": {"only_source": 1, "only_target": 0, "diff": 1, "same": 7},
        "progress": {},
        "diff_profile": {},
        "sample_result": None,
    }
    engine = _FakeEngine([run_row, {"id": "project-1"}, 0])
    job_backend = _JobBackend()
    services = _Services(engine, job_backend=job_backend)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/compare/runs/run-1/export",
        headers=_auth_headers(),
        json_body={"format": "excel"},
    )

    assert response.status_code == 202
    payload = response.json()
    assert payload["job_id"]
    assert payload["download_token"]
    assert payload["format"] == "excel"
    assert payload["filename"] == "run-1.xlsx"
    job = job_backend.enqueued[0]
    assert job.kind is JobKind.RESULT_EXPORT
    assert job.payload["compare_run_id"] == "run-1"
    assert job.payload["bucket_result_set_ids"] == bucket_spools
    assert job.payload["format"] == "excel"
    assert job.payload["filename"] == "run-1.xlsx"
    assert any(audit["action"] == "compare_export" for audit in services.audits)
    assert "download_token" not in str(services.audits[-1])


def test_create_compare_export_requires_successful_run() -> None:
    run_row: dict[str, object] = {
        "run_id": "run-1",
        "job_id": "job-compare",
        "project_id": "project-1",
        "status": "running",
        "bucket_spools": {},
        "bucket_counts": {},
        "progress": {},
        "diff_profile": {},
        "sample_result": None,
    }
    services = _Services(_FakeEngine([run_row, {"id": "project-1"}]))
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/compare/runs/run-1/export",
        headers=_auth_headers(),
        json_body={"format": "excel"},
    )

    assert response.status_code == 409
    assert response.json()["error"] == "compare_run_not_successful"
    assert services.job_backend.enqueued == []


def test_create_export_enqueues_result_export_job_and_returns_one_time_token() -> None:
    engine = _FakeEngine(
        [
            _source_job_row(),
            {"id": "project-1"},
            0,
            {"db_type": "mysql"},
        ]
    )
    job_backend = _JobBackend()
    services = _Services(engine, job_backend=job_backend)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/jobs/job-source/export",
        headers=_auth_headers(),
        json_body={"format": "csv"},
    )

    assert response.status_code == 202
    payload = response.json()
    assert payload["job_id"]
    assert payload["download_token"]
    assert payload["format"] == "csv"
    assert payload["filename"] == "job-source.csv"
    job = job_backend.enqueued[0]
    assert job.kind is JobKind.RESULT_EXPORT
    assert job.payload["source_job_id"] == "job-source"
    assert job.payload["source_result_set_id"] == "rs-source"
    assert job.payload["format"] == "csv"
    assert job.payload["source_db_type"] == "mysql"
    assert any(audit["action"] == "sql_export" for audit in services.audits)
    assert "download_token" not in str(services.audits[-1])


def test_create_export_rate_limit_returns_429_without_enqueue() -> None:
    engine = _FakeEngine(
        [
            _source_job_row(),
            {"id": "project-1"},
            10,
        ]
    )
    job_backend = _JobBackend()
    services = _Services(engine, job_backend=job_backend)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/jobs/job-source/export",
        headers=_auth_headers(),
        json_body={"format": "json"},
    )

    assert response.status_code == 429
    assert response.json()["error"] == "export_rate_limited"
    assert job_backend.enqueued == []


def test_create_export_requires_source_spool_to_still_exist() -> None:
    engine = _FakeEngine([_source_job_row(), {"id": "project-1"}])
    job_backend = _JobBackend()
    result_store = _ResultStore(spool_exists=False)
    services = _Services(engine, job_backend=job_backend, result_store=result_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/jobs/job-source/export",
        headers=_auth_headers(),
        json_body={"format": "csv"},
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"
    assert job_backend.enqueued == []


def test_export_download_token_is_one_time() -> None:
    token_row = {
        "token_hash": "not-read-by-fake-engine",
        "owner_user_id": "user-1",
        "filename": "job-source.csv",
        "content_type": "text/csv; charset=utf-8",
        "expires_at": datetime.now(UTC) + timedelta(minutes=5),
        "consumed_at": None,
        "job_id": "job-export",
        "status": "success",
        "finished_at": datetime.now(UTC),
        "result_ref": {
            "backend": "local_fs",
            "uri": "exports/job-export/job-source.csv",
            "metadata": {},
        },
    }
    engine = _FakeEngine([token_row, token_row | {"consumed_at": _dt(1)}])
    result_store = _ResultStore(downloads={"exports/job-export/job-source.csv": b"value\n1\n"})
    app = create_app(services=cast(ApiServices, _Services(engine, result_store=result_store)))

    first = AsgiClient(app).get("/api/exports/download-token", headers=_auth_headers())
    second = AsgiClient(app).get("/api/exports/download-token", headers=_auth_headers())

    assert first.status_code == 200
    assert first.body == b"value\n1\n"
    assert first.headers["content-disposition"] == 'attachment; filename="job-source.csv"'
    assert second.status_code == 410
    assert second.json()["error"] == "download_token_consumed"


def test_export_download_token_valid_after_slow_job_outlives_creation_ttl() -> None:
    # The export job took longer than download_url_ttl_seconds (300s) to finish,
    # so the token's creation-relative expires_at is already in the past. The TTL
    # clock must run from job completion (finished_at), so the download still works
    # (backlog #96 #3).
    now = datetime.now(UTC)
    token_row = {
        "token_hash": "not-read-by-fake-engine",
        "owner_user_id": "user-1",
        "filename": "job-source.csv",
        "content_type": "text/csv; charset=utf-8",
        # created ~10 min ago → creation+300s window already expired
        "expires_at": now - timedelta(seconds=300),
        "consumed_at": None,
        "job_id": "job-export",
        "status": "success",
        # job finished just now, i.e. well after the creation-relative window
        "finished_at": now,
        "result_ref": {
            "backend": "local_fs",
            "uri": "exports/job-export/job-source.csv",
            "metadata": {},
        },
    }
    engine = _FakeEngine([token_row])
    result_store = _ResultStore(downloads={"exports/job-export/job-source.csv": b"value\n1\n"})
    app = create_app(services=cast(ApiServices, _Services(engine, result_store=result_store)))

    response = AsgiClient(app).get("/api/exports/download-token", headers=_auth_headers())

    assert response.status_code == 200
    assert response.body == b"value\n1\n"


def test_export_download_token_expired_after_completion_grace_window() -> None:
    # Job finished long ago (beyond download_url_ttl_seconds after completion) and
    # the token was never consumed → the completion-relative window has closed.
    now = datetime.now(UTC)
    token_row = {
        "token_hash": "not-read-by-fake-engine",
        "owner_user_id": "user-1",
        "filename": "job-source.csv",
        "content_type": "text/csv; charset=utf-8",
        "expires_at": now - timedelta(seconds=1200),
        "consumed_at": None,
        "job_id": "job-export",
        "status": "success",
        "finished_at": now - timedelta(seconds=600),
        "result_ref": {
            "backend": "local_fs",
            "uri": "exports/job-export/job-source.csv",
            "metadata": {},
        },
    }
    engine = _FakeEngine([token_row])
    result_store = _ResultStore(downloads={"exports/job-export/job-source.csv": b"value\n1\n"})
    app = create_app(services=cast(ApiServices, _Services(engine, result_store=result_store)))

    response = AsgiClient(app).get("/api/exports/download-token", headers=_auth_headers())

    assert response.status_code == 410
    assert response.json()["error"] == "download_token_expired"


def test_lineage_analyze_persists_edges_and_returns_cache_contract() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _datasource_row(),
            [
                {
                    "cache_level": "columns",
                    "schema_name": "app",
                    "table_name": "src",
                    "payload": [{"name": "id", "type": "integer"}],
                },
                {
                    "cache_level": "columns",
                    "schema_name": "app",
                    "table_name": "tgt",
                    "payload": [{"name": "id", "type": "integer"}],
                },
            ],
            None,
            {
                "id": "run-1",
                "project_id": "project-1",
                "datasource_id": "ds-1",
                "dialect": "mysql",
                "source_ref": "script.sql",
                "sql_hash": "hash-1",
                "parser_version": "sqlglot-w1-v1",
                "status": "success",
                "parse_summary": {
                    "table_edge_count": 1,
                    "column_mapping_count": 1,
                    "parse_error_count": 0,
                },
            },
            1,
            1,
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/analyze",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "source_ref": "script.sql",
            "default_schema": "app",
            "sql_text": "INSERT INTO app.tgt (id) SELECT id FROM app.src",
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["run_id"] == "run-1"
    assert payload["project_id"] == "project-1"
    assert payload["cached"] is False
    assert payload["parser_version"] == "sqlglot-w1-v1"
    assert payload["table_edge_count"] == 1
    assert payload["column_edge_count"] == 1
    assert any("INSERT INTO lineage_runs" in statement for statement in engine.statements)
    assert any("INSERT INTO lineage_edges" in statement for statement in engine.statements)
    assert any("INSERT INTO lineage_column_edges" in statement for statement in engine.statements)


def test_lineage_analyze_cache_hit_does_not_reinsert_edges() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _datasource_row(),
            [],
            {
                "id": "run-cached",
                "project_id": "project-1",
                "datasource_id": "ds-1",
                "dialect": "mysql",
                "source_ref": "script.sql",
                "sql_hash": "hash-1",
                "parser_version": "sqlglot-w1-v1",
                "status": "success",
                "parse_summary": {
                    "table_edge_count": 1,
                    "column_mapping_count": 0,
                    "parse_error_count": 1,
                    "parse_errors": [
                        {
                            "statement_index": 0,
                            "error_type": "unsupported_schema",
                            "message": "metadata cache missing for table app.src",
                            "unsupported": True,
                            "statement_type": "INSERT",
                        }
                    ],
                },
            },
            1,
            0,
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/analyze",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "source_ref": "script.sql",
            "sql_text": "INSERT INTO tgt (id) SELECT id FROM src",
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["cached"] is True
    assert payload["run_id"] == "run-cached"
    assert payload["table_edge_count"] == 1
    assert payload["parse_summary"]["parse_error_count"] == 1
    assert payload["parse_summary"]["parse_errors"] == [
        {
            "statement_index": 0,
            "error_type": "unsupported_schema",
            "message": "metadata cache missing for table app.src",
            "unsupported": True,
            "statement_type": "INSERT",
        }
    ]
    assert not any("INSERT INTO lineage_edges" in statement for statement in engine.statements)


def test_lineage_subgraph_contract_uses_recursive_cte() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            [
                {
                    "edge_id": "edge-1",
                    "source_table": "app.src",
                    "target_table": "app.mid",
                    "source_column": None,
                    "target_column": None,
                    "source": "app.src",
                    "target": "app.mid",
                    "edge_kind": "table",
                    "transformation": None,
                    "transformation_subtype": None,
                    "inferred": False,
                    "inference_status": "confirmed",
                    "confidence": 1,
                    "depth": 1,
                    "path": ["app.src", "app.mid"],
                    "direction": "downstream",
                }
            ],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/lineage/subgraph",
        headers=_auth_headers(),
        params={"focus": "app.src", "direction": "downstream", "max_depth": 3},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["node_count"] == 2
    assert payload["edge_count"] == 1
    assert payload["truncated"] is False
    assert payload["nodes"][0] == {
        "id": "app.src",
        "label": "app.src",
        "kind": "table",
        "table": "app.src",
        "column": None,
        "depth": 0,
    }
    assert payload["edges"][0]["source"] == "app.src"
    assert payload["edges"][0]["target"] == "app.mid"
    assert any("WITH RECURSIVE walk" in statement for statement in engine.statements)


def test_lineage_subgraph_include_columns_expands_bounded_column_edges() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            "edge-exists",
            [
                {
                    "edge_id": "edge-1",
                    "source_table": "app.src",
                    "target_table": "app.mid",
                    "source_column": None,
                    "target_column": None,
                    "source": "app.src",
                    "target": "app.mid",
                    "edge_kind": "table",
                    "transformation": None,
                    "transformation_subtype": None,
                    "inferred": False,
                    "inference_status": "confirmed",
                    "confidence": 1,
                    "depth": 1,
                    "path": ["app.src", "app.mid"],
                    "direction": "downstream",
                }
            ],
            [
                {
                    "id": "col-edge-1",
                    "source_table": "app.src",
                    "source_column": "id",
                    "target_table": "app.mid",
                    "target_column": "id",
                    "transformation": "DIRECT",
                    "transformation_subtype": "DIRECT",
                    "inferred": False,
                    "inference_status": "confirmed",
                    "confidence": 1,
                }
            ],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/lineage/subgraph",
        headers=_auth_headers(),
        params={
            "focus": "app.src",
            "direction": "downstream",
            "max_depth": 3,
            "include_columns": True,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert any(node["id"] == "app.src.id" and node["kind"] == "column" for node in payload["nodes"])
    assert any(
        edge["id"] == "col-edge-1" and edge["edge_kind"] == "column" for edge in payload["edges"]
    )


def test_lineage_impact_contract_returns_downstream_paths() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            "edge-exists",
            [
                {
                    "edge_id": "edge-1",
                    "source_table": "app.src",
                    "target_table": "app.mid",
                    "source_column": None,
                    "target_column": None,
                    "source": "app.src",
                    "target": "app.mid",
                    "edge_kind": "table",
                    "transformation": None,
                    "transformation_subtype": None,
                    "inferred": False,
                    "inference_status": "confirmed",
                    "confidence": 1,
                    "depth": 1,
                    "path": ["app.src", "app.mid"],
                    "direction": "downstream",
                }
            ],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/lineage/impact",
        headers=_auth_headers(),
        params={"focus": "app.src", "max_depth": 3},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["impact_count"] == 1
    assert payload["impacts"][0] == {
        "node": "app.mid",
        "table": "app.mid",
        "column": None,
        "depth": 1,
        "paths": [["app.src", "app.mid"]],
    }


def _impact_walk_row() -> dict[str, object]:
    return {
        "edge_id": "edge-1",
        "source_table": "app.src",
        "target_table": "app.mid",
        "source_column": None,
        "target_column": None,
        "source": "app.src",
        "target": "app.mid",
        "edge_kind": "table",
        "transformation": None,
        "transformation_subtype": None,
        "inferred": False,
        "inference_status": "confirmed",
        "confidence": 1,
        "depth": 1,
        "path": ["app.src", "app.mid"],
        "direction": "downstream",
    }


def _patch_gateway(monkeypatch: pytest.MonkeyPatch, captured: dict[str, Any]) -> None:
    class _Gateway:
        def complete(self, prompt: str, context: Any, options: Any) -> Any:
            captured["prompt"] = prompt
            captured["context"] = context
            captured["options"] = options
            return AiResponse(content="ranked", provider="mock", model="mock-model")

    monkeypatch.setattr(
        core_routes,
        "build_gateway_from_runtime_config",
        lambda runtime, **kwargs: _Gateway(),
    )


def test_lineage_ai_impact_weights_frequency_and_owner_via_gateway_l2(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},  # project access
            _ai_config_row(),  # ai config
            "edge-exists",  # focus_ref table exists
            [_impact_walk_row()],  # downstream walk
            {"username": "alice"},  # project owner
            [
                {
                    "table_name": "app.mid",
                    "ref_count": 4,
                    "source_ref_count": 2,
                    "last_referenced_at": datetime.now(UTC),
                }
            ],  # frequency
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))
    captured: dict[str, Any] = {}
    _patch_gateway(monkeypatch, captured)

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/ai-impact",
        headers=_auth_headers(),
        json_body={"focus": "app.src", "max_depth": 3, "window_days": 90},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["assessment"] == "ranked"
    assert payload["egress_level"] == 2
    assert payload["degraded"] is False
    assert payload["impact_count"] == 1
    assert payload["project_owner"] == "alice"
    assert payload["signals"][0]["ref_count"] == 4
    assert payload["signals"][0]["source_ref_count"] == 2
    assert payload["signals"][0]["last_referenced_days_ago"] == 0
    # L2 出站 + 无行值 / 无 SQL 原文:只送名字 + 聚合计数。
    context = captured["context"]
    assert len(context.items) == 1
    assert context.items[0].egress_level == 2
    content = context.items[0].content
    assert "app.mid" in content
    assert "alice" in content
    assert "SELECT" not in content.upper()
    assert "INSERT" not in content.upper()
    assert captured["options"].purpose == "lineage_ai_impact"
    assert any(
        audit["action"] == "lineage_ai_impact" and audit["result"] == "success"
        for audit in services.audits
    )


def test_lineage_ai_impact_degrades_gracefully_without_frequency(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _ai_config_row(),
            "edge-exists",
            [_impact_walk_row()],
            {"username": "alice"},
            [],  # no frequency rows within window → thin data
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))
    captured: dict[str, Any] = {}
    _patch_gateway(monkeypatch, captured)

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/ai-impact",
        headers=_auth_headers(),
        json_body={"focus": "app.src"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["degraded"] is True
    assert payload["signals"][0]["ref_count"] == 0
    assert payload["signals"][0]["last_referenced_days_ago"] is None
    # 稀薄频率仍出站,但标注 frequency_available=false 供 AI 降级加权。
    assert '"frequency_available":false' in captured["context"].items[0].content


def test_lineage_ai_impact_disabled_returns_409(monkeypatch: pytest.MonkeyPatch) -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            None,  # no ai_configs row → disabled
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))
    called: dict[str, Any] = {}
    monkeypatch.setattr(
        core_routes,
        "build_gateway_from_runtime_config",
        lambda *a, **k: called.setdefault("built", True),
    )

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/ai-impact",
        headers=_auth_headers(),
        json_body={"focus": "app.src"},
    )

    assert response.status_code == 409
    assert response.json()["error"] == "ai_disabled"
    assert "built" not in called  # gateway never constructed when disabled
    assert any(
        audit["action"] == "lineage_ai_impact" and audit["result"] == "skipped"
        for audit in services.audits
    )


def test_lineage_export_writes_artifact_token_and_audit_synchronously() -> None:
    walk_row = {
        "edge_id": "edge-1",
        "source_table": "app.src",
        "target_table": "app.mid",
        "source_column": None,
        "target_column": None,
        "source": "app.src",
        "target": "app.mid",
        "edge_kind": "table",
        "transformation": None,
        "transformation_subtype": None,
        "inferred": False,
        "inference_status": "confirmed",
        "confidence": 1,
        "depth": 1,
        "path": ["app.src", "app.mid"],
        "direction": "downstream",
    }
    engine = _FakeEngine(
        [
            {"id": "project-1"},  # project access
            0,  # export rate limit count
            [dict(walk_row)],  # subgraph downstream walk
            "edge-exists",  # impact focus_ref table exists
            [dict(walk_row)],  # impact downstream walk
        ]
    )
    job_backend = _JobBackend()
    result_store = _ResultStore()
    services = _Services(engine, job_backend=job_backend, result_store=result_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/export",
        headers=_auth_headers(),
        json_body={"focus": "app.src", "direction": "downstream", "max_depth": 3},
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["job_id"]
    assert payload["download_token"]
    assert payload["format"] == "excel"
    assert payload["filename"] == f"lineage-{payload['job_id']}.xlsx"
    # 同步生成:artifact 已落 result_store,不入 job 队列
    assert job_backend.enqueued == []
    export_id, name, content = result_store.export_artifacts[0]
    assert export_id == payload["job_id"]
    assert name == payload["filename"]
    assert content.startswith(b"PK")  # xlsx = zip
    # jobs 落 success 行 + token 落库(下载端点 join jobs 依赖)
    assert any(statement.startswith("INSERT INTO jobs") for statement in engine.statements)
    assert any(
        statement.startswith("INSERT INTO export_download_tokens")
        for statement in engine.statements
    )
    audit = next(item for item in services.audits if item["action"] == "lineage_export")
    detail = cast(dict[str, object], audit["detail"])
    assert detail["focus"] == "app.src"
    assert detail["direction"] == "downstream"
    assert detail["table_edge_count"] == 1
    assert detail["impact_count"] == 1
    assert "download_token" not in str(services.audits[-1])


def test_lineage_export_rate_limit_returns_429_without_artifact() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            10,  # 已到每小时上限
        ]
    )
    result_store = _ResultStore()
    services = _Services(engine, result_store=result_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/export",
        headers=_auth_headers(),
        json_body={"focus": "app.src"},
    )

    assert response.status_code == 429
    assert response.json()["error"] == "export_rate_limited"
    assert result_store.export_artifacts == []
    assert not any(statement.startswith("INSERT INTO") for statement in engine.statements)


def test_upload_stores_artifact_and_registers_row() -> None:
    engine = _FakeEngine([{"id": "project-1"}])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))
    content = b"PK\x03\x04-demo-zip-bytes"

    response = AsgiClient(app).post(
        "/api/projects/project-1/uploads",
        headers={**_auth_headers(), "content-type": "application/zip"},
        params={"purpose": "lineage_batch", "filename": "scripts.zip"},
        raw_body=content,
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["purpose"] == "lineage_batch"
    assert payload["filename"] == "scripts.zip"
    assert payload["bytes"] == len(content)
    upload_id, name, stored = services.result_store.upload_artifacts[0]
    assert upload_id == payload["upload_id"]
    assert name == "scripts.zip"
    assert stored == content
    assert any(statement.startswith("INSERT INTO uploads") for statement in engine.statements)
    assert any(audit["action"] == "upload_create" for audit in services.audits)


def test_upload_too_large_returns_413_without_side_effects() -> None:
    engine = _FakeEngine([{"id": "project-1"}])
    services = _Services(engine, upload_max_mb=1)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/uploads",
        headers=_auth_headers(),
        params={"purpose": "compare_source", "filename": "big.csv"},
        raw_body=b"x" * (1024 * 1024 + 1),
    )

    assert response.status_code == 413
    assert response.json()["error"] == "upload_too_large"
    assert services.result_store.upload_artifacts == []
    assert not any(statement.startswith("INSERT INTO") for statement in engine.statements)


def test_upload_empty_body_rejected() -> None:
    engine = _FakeEngine([{"id": "project-1"}])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/uploads",
        headers=_auth_headers(),
        params={"purpose": "lineage_batch", "filename": "empty.zip"},
        raw_body=b"",
    )

    assert response.status_code == 400
    assert response.json()["error"] == "empty_upload"


def _compare_upload_row(
    *,
    upload_id: str = "up-1",
    purpose: str = "compare_source",
    storage_uri: str = "uploads/up-1/data.csv",
    filename: str = "data.csv",
) -> dict[str, Any]:
    return {
        "id": upload_id,
        "project_id": "project-1",
        "owner_user_id": "user-1",
        "purpose": purpose,
        "filename": filename,
        "content_type": "application/octet-stream",
        "bytes": 64,
        "storage_uri": storage_uri,
        "created_at": _dt(1),
    }


def _xlsx_bytes(sheets: dict[str, list[list[object]]]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(title=sheet_name)
        for row in rows:
            worksheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_compare_file_preview_csv_returns_bounded_rows() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _compare_upload_row()])
    result_store = _ResultStore(
        downloads={"uploads/up-1/data.csv": b"id,name\n1,alice\n2,bob\n3,carol\n"}
    )
    services = _Services(engine, result_store=result_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/file-preview",
        headers=_auth_headers(),
        json_body={
            "ref": {"kind": "file", "upload_id": "up-1", "file_format": "csv"},
            "limit": 2,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["sheets"] == []
    assert payload["columns"] == ["id", "name"]
    assert payload["rows"] == [["1", "alice"], ["2", "bob"]]
    assert payload["row_count"] == 2
    assert payload["truncated"] is True  # 第 3 行探到 → 截断
    assert payload["detected_encoding"] == "utf-8-sig"
    assert any(audit["action"] == "compare_file_preview" for audit in services.audits)


def test_compare_file_preview_excel_lists_sheets_and_selects() -> None:
    xlsx = _xlsx_bytes(
        {
            "First": [["a"], [1]],
            "Second": [["b", "c"], [2, 3]],
        }
    )
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _compare_upload_row(storage_uri="uploads/up-1/book.xlsx", filename="book.xlsx"),
        ]
    )
    result_store = _ResultStore(downloads={"uploads/up-1/book.xlsx": xlsx})
    services = _Services(engine, result_store=result_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/file-preview",
        headers=_auth_headers(),
        json_body={
            "ref": {
                "kind": "file",
                "upload_id": "up-1",
                "file_format": "excel",
                "sheet": "Second",
            },
            "limit": 50,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["sheets"] == ["First", "Second"]
    assert payload["columns"] == ["b", "c"]
    assert payload["rows"] == [[2, 3]]
    assert payload["row_count"] == 1
    assert payload["truncated"] is False
    assert payload["detected_encoding"] is None  # excel 无编码回显


def test_compare_file_preview_rejects_non_file_ref() -> None:
    engine = _FakeEngine([{"id": "project-1"}])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/file-preview",
        headers=_auth_headers(),
        json_body={
            "ref": {"kind": "table", "table_name": "orders"},
            "limit": 10,
        },
    )

    assert response.status_code == 400
    assert response.json()["error"] == "invalid_ref_kind"


def test_compare_file_preview_upload_not_found() -> None:
    engine = _FakeEngine([{"id": "project-1"}, None])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/file-preview",
        headers=_auth_headers(),
        json_body={
            "ref": {"kind": "file", "upload_id": "missing", "file_format": "csv"},
        },
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"


def test_compare_file_preview_rejects_wrong_purpose() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _compare_upload_row(purpose="lineage_batch")])
    result_store = _ResultStore(downloads={"uploads/up-1/data.csv": b"id\n1\n"})
    app = create_app(services=cast(ApiServices, _Services(engine, result_store=result_store)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/file-preview",
        headers=_auth_headers(),
        json_body={
            "ref": {"kind": "file", "upload_id": "up-1", "file_format": "csv"},
        },
    )

    assert response.status_code == 400
    assert response.json()["error"] == "invalid_upload_purpose"


def test_compare_file_preview_reader_failure_returns_400() -> None:
    # 把非 xlsx 字节当 excel 读 → ReaderError,R5 下不透传内容,只回 preview_failed。
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _compare_upload_row(storage_uri="uploads/up-1/book.xlsx", filename="book.xlsx"),
        ]
    )
    result_store = _ResultStore(downloads={"uploads/up-1/book.xlsx": b"not a real workbook"})
    app = create_app(services=cast(ApiServices, _Services(engine, result_store=result_store)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/compare/file-preview",
        headers=_auth_headers(),
        json_body={
            "ref": {"kind": "file", "upload_id": "up-1", "file_format": "excel"},
        },
    )

    assert response.status_code == 400
    assert response.json()["error"] == "preview_failed"


def test_lineage_batch_create_enqueues_job_with_upload_ref() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            {
                "id": "up-1",
                "project_id": "project-1",
                "owner_user_id": "user-1",
                "purpose": "lineage_batch",
                "filename": "scripts.zip",
                "content_type": "application/zip",
                "bytes": 2048,
                "storage_uri": "uploads/up-1/scripts.zip",
                "created_at": _dt(1),
            },
            _datasource_row(),
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/batch",
        headers=_auth_headers(),
        json_body={"upload_id": "up-1", "datasource_id": "ds-1", "default_schema": "dw"},
    )

    assert response.status_code == 202
    payload = response.json()
    backend = services.job_backend
    assert len(backend.enqueued) == 1
    job = backend.enqueued[0]
    assert job.id == payload["job_id"]
    assert job.kind.value == "lineage_batch"
    assert job.payload["storage_uri"] == "uploads/up-1/scripts.zip"
    assert job.payload["default_schema"] == "dw"
    assert job.datasource_ids == ["ds-1"]
    assert any(audit["action"] == "lineage_batch_create" for audit in services.audits)


def test_lineage_batch_without_datasource_requires_dialect_and_skips_ds_lookup() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            {
                "id": "up-1",
                "project_id": "project-1",
                "owner_user_id": "user-1",
                "purpose": "lineage_batch",
                "filename": "scripts.zip",
                "content_type": "application/zip",
                "bytes": 2048,
                "storage_uri": "uploads/up-1/scripts.zip",
                "created_at": _dt(1),
            },
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/batch",
        headers=_auth_headers(),
        json_body={"upload_id": "up-1", "dialect": "oracle"},
    )

    assert response.status_code == 202
    job = services.job_backend.enqueued[0]
    assert job.payload["datasource_id"] is None
    assert job.payload["dialect"] == "oracle"
    assert job.datasource_ids == []
    # 未做数据源查询(引擎序列只喂了 project + upload 两行)
    assert not any("FROM datasources" in statement for statement in engine.statements)


def test_lineage_batch_without_datasource_missing_dialect_is_422() -> None:
    engine = _FakeEngine([{"id": "project-1"}])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/batch",
        headers=_auth_headers(),
        json_body={"upload_id": "up-1"},
    )

    # pydantic 校验:无数据源必须给 dialect
    assert response.status_code == 422


def test_lineage_batch_rejects_wrong_purpose_upload() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            {
                "id": "up-1",
                "project_id": "project-1",
                "owner_user_id": "user-1",
                "purpose": "compare_source",
                "filename": "data.csv",
                "content_type": "text/csv",
                "bytes": 100,
                "storage_uri": "uploads/up-1/data.csv",
                "created_at": _dt(1),
            },
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/batch",
        headers=_auth_headers(),
        json_body={"upload_id": "up-1", "datasource_id": "ds-1"},
    )

    assert response.status_code == 400
    assert response.json()["error"] == "invalid_upload_purpose"


def test_lineage_batch_status_returns_report_on_success() -> None:
    report_json = json.dumps(
        {
            "file_count": 2,
            "parsed": 2,
            "failed": 0,
            "files": [],
            "script_edges": [],
        }
    ).encode("utf-8")
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            {
                "id": "job-1",
                "kind": "lineage_batch",
                "project_id": "project-1",
                "status": "success",
                "error": None,
                "result_ref": {
                    "backend": "lineage_batch",
                    "uri": "lineage-batch/job-1",
                    "metadata": {"report_uri": "runs/job-1/artifacts/lineage_batch_report.json"},
                },
            },
        ]
    )
    result_store = _ResultStore(
        downloads={"runs/job-1/artifacts/lineage_batch_report.json": report_json}
    )
    app = create_app(services=cast(ApiServices, _Services(engine, result_store=result_store)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/lineage/batch/job-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "success"
    assert payload["report"]["file_count"] == 2


def test_lineage_batch_export_writes_xlsx_token_and_audit() -> None:
    from openpyxl import load_workbook

    report_uri = "runs/job-1/artifacts/lineage_batch_report.json"
    report_json = json.dumps(
        {
            "file_count": 2,
            "parsed": 2,
            "failed": 0,
            "skipped": {"non_sql": 0, "too_large": 0, "over_file_limit": 0},
            "table_edge_total": 1,
            "column_mapping_total": 0,
            "files": [
                {
                    "source_ref": "etl/a.sql",
                    "status": "parsed",
                    "run_id": "run-1",
                    "table_edge_count": 1,
                    "column_mapping_count": 0,
                    "parse_error_count": 0,
                    "lenient_statement_count": 0,
                    "tables_read": ["app.src"],
                    "tables_written": ["app.mid"],
                }
            ],
            "script_edges": [
                {"producer_file": "etl/a.sql", "consumer_file": "etl/b.sql", "table": "app.mid"}
            ],
            "semantic_view": {
                "targets": [
                    {
                        "table": "app.mid",
                        "primary_role": "target",
                        "roles": ["target"],
                        "refresh_mode": "append",
                        "counts": {"insert": 1},
                        "source_tables": ["app.src"],
                        "source_files": ["etl/a.sql"],
                        "titles": [],
                    }
                ],
                "table_roles": [
                    {"table": "app.src", "primary_role": "source_fact", "roles": ["source_fact"]}
                ],
                "observations": ["app.mid 追加写入"],
                "risks": [],
            },
        }
    ).encode("utf-8")
    engine = _FakeEngine(
        [
            {"id": "project-1"},  # project access
            {
                "id": "job-1",
                "kind": "lineage_batch",
                "project_id": "project-1",
                "status": "success",
                "error": None,
                "result_ref": {
                    "backend": "lineage_batch",
                    "uri": "lineage-batch/job-1",
                    "metadata": {"report_uri": report_uri},
                },
            },
            0,  # export rate limit count
        ]
    )
    result_store = _ResultStore(downloads={report_uri: report_json})
    job_backend = _JobBackend()
    services = _Services(engine, job_backend=job_backend, result_store=result_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/batch/job-1/export",
        headers=_auth_headers(),
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["format"] == "excel"
    assert payload["filename"] == f"lineage-batch-{payload['job_id']}.xlsx"
    assert payload["download_token"]
    # 同步生成:xlsx artifact 已落 result_store,不入 job 队列
    assert job_backend.enqueued == []
    export_id, name, content = result_store.export_artifacts[0]
    assert export_id == payload["job_id"]
    assert name == payload["filename"]
    assert content.startswith(b"PK")  # xlsx = zip

    workbook = load_workbook(io.BytesIO(content), read_only=True)
    assert workbook.sheetnames == [
        # A 组
        "流程总览",
        "脚本清单",
        "跨脚本依赖",
        "风险提示",
        "流程图分组",
        "表角色",
        "目标整合",
        # B 组(逐文件明细,PR2)
        "过程段明细",
        "语句明细",
        "解析失败明细",
        "动态SQL明细",
        "脚本警告",
        "字段映射明细",
        "表级数据流",
    ]

    assert any(statement.startswith("INSERT INTO jobs") for statement in engine.statements)
    assert any(
        statement.startswith("INSERT INTO export_download_tokens")
        for statement in engine.statements
    )
    audit = next(item for item in services.audits if item["action"] == "lineage_batch_export")
    detail = cast(dict[str, object], audit["detail"])
    assert detail["batch_job_id"] == "job-1"
    assert detail["file_count"] == 2
    assert "download_token" not in str(services.audits[-1])


def test_lineage_batch_export_missing_job_returns_404() -> None:
    engine = _FakeEngine([{"id": "project-1"}, None])
    result_store = _ResultStore()
    services = _Services(engine, result_store=result_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/batch/missing/export",
        headers=_auth_headers(),
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"
    assert result_store.export_artifacts == []
    assert not any(statement.startswith("INSERT INTO") for statement in engine.statements)


def test_lineage_batch_export_incomplete_job_returns_409() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            {
                "id": "job-1",
                "kind": "lineage_batch",
                "project_id": "project-1",
                "status": "running",
                "error": None,
                "result_ref": None,
            },
        ]
    )
    result_store = _ResultStore()
    services = _Services(engine, result_store=result_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/batch/job-1/export",
        headers=_auth_headers(),
    )

    assert response.status_code == 409
    assert response.json()["error"] == "batch_not_ready"
    assert result_store.export_artifacts == []
    assert not any(statement.startswith("INSERT INTO") for statement in engine.statements)


def test_lineage_edge_detail_returns_run_provenance() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            {
                "id": "edge-1",
                "run_id": "run-1",
                "project_id": "project-1",
                "source_table": "app.src",
                "target_table": "app.mid",
                "edge_kind": "table",
                "inferred": False,
                "inference_status": "confirmed",
                "confidence": 1.0,
                "sql_hash": "hash-1",
            },
            {
                "id": "run-1",
                "project_id": "project-1",
                "datasource_id": "ds-1",
                "dialect": "mysql",
                "source_ref": "etl/orders.sql",
                "sql_hash": "hash-1",
                "sql_text": "INSERT INTO app.mid SELECT * FROM app.src",
                "parser_version": "sqlglot-w1-v2",
                "status": "success",
                "parse_summary": {},
                "created_at": _dt(1),
                "updated_at": _dt(1),
            },
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/lineage/edges/edge-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["edge_id"] == "edge-1"
    assert payload["edge_kind"] == "table"
    assert payload["source_table"] == "app.src"
    assert payload["target_table"] == "app.mid"
    assert payload["run"]["run_id"] == "run-1"
    assert payload["run"]["source_ref"] == "etl/orders.sql"
    assert payload["run"]["sql_text"] == "INSERT INTO app.mid SELECT * FROM app.src"


def test_lineage_edge_detail_null_sql_text_for_legacy_runs() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            {
                "id": "edge-1",
                "run_id": "run-1",
                "project_id": "project-1",
                "source_table": "app.src",
                "target_table": "app.mid",
                "edge_kind": "table",
                "inferred": True,
                "inference_status": "inferred",
                "confidence": 0.8,
                "sql_hash": "hash-1",
            },
            {
                "id": "run-1",
                "project_id": "project-1",
                "datasource_id": "ds-1",
                "dialect": "mysql",
                "source_ref": "etl/orders.sql",
                "sql_hash": "hash-1",
                "sql_text": None,  # 0018 前的旧 run
                "parser_version": "sqlglot-w1-v2",
                "status": "success",
                "parse_summary": {},
                "created_at": _dt(1),
                "updated_at": _dt(1),
            },
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/lineage/edges/edge-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["inferred"] is True
    assert payload["confidence"] == 0.8
    assert payload["run"]["sql_text"] is None


def test_lineage_column_trace_contract_returns_hop_chain() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            [
                {
                    "edge_id": "col-edge-1",
                    "source_table": "app.src",
                    "target_table": "app.mid",
                    "source_column": "id",
                    "target_column": "id",
                    "source": "app.src.id",
                    "target": "app.mid.id",
                    "edge_kind": "column",
                    "transformation": "DIRECT",
                    "transformation_subtype": "DIRECT",
                    "inferred": False,
                    "inference_status": "confirmed",
                    "confidence": 1,
                    "depth": 1,
                    "path": ["app.mid.id", "app.src.id"],
                    "direction": "upstream",
                },
                {
                    "edge_id": "col-edge-2",
                    "source_table": "app.origin",
                    "target_table": "app.src",
                    "source_column": "id",
                    "target_column": "id",
                    "source": "app.origin.id",
                    "target": "app.src.id",
                    "edge_kind": "column",
                    "transformation": "TRANSFORMATION",
                    "transformation_subtype": "CAST",
                    "inferred": True,
                    "inference_status": "inferred",
                    "confidence": 0.8,
                    "depth": 2,
                    "path": ["app.mid.id", "app.src.id", "app.origin.id"],
                    "direction": "upstream",
                },
            ],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/lineage/column-trace",
        headers=_auth_headers(),
        params={"focus": "app.mid.id", "direction": "upstream", "max_depth": 3},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["focus"] == "app.mid.id"
    assert payload["hop_count"] == 2
    hop1, hop2 = payload["hops"]
    assert hop1["depth"] == 1
    # 链式语义:上游走向的本跳节点是 source,经由端(from)是 target
    assert hop1["items"][0]["node"] == "app.src.id"
    assert hop1["items"][0]["from_node"] == "app.mid.id"
    assert hop1["items"][0]["table"] == "app.src"
    assert hop1["items"][0]["column"] == "id"
    assert hop2["items"][0]["node"] == "app.origin.id"
    assert hop2["items"][0]["from_node"] == "app.src.id"
    assert hop2["items"][0]["inferred"] is True
    assert hop2["items"][0]["confidence"] == 0.8
    assert any("WITH RECURSIVE walk" in statement for statement in engine.statements)


def test_lineage_column_trace_requires_column_focus() -> None:
    engine = _FakeEngine([{"id": "project-1"}])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/lineage/column-trace",
        headers=_auth_headers(),
        params={"focus": "plaintable", "direction": "upstream"},
    )

    assert response.status_code == 400
    assert response.json()["error"] == "column_focus_required"


def _trace_compare_walk_rows() -> list[dict[str, object]]:
    """上游 walk 结果:焦点 app.orders.amount ← app.stg_orders.amt(1 跳)。"""
    return [
        {
            "edge_id": "col-edge-1",
            "source_table": "app.stg_orders",
            "target_table": "app.orders",
            "source_column": "amt",
            "target_column": "amount",
            "source": "app.stg_orders.amt",
            "target": "app.orders.amount",
            "edge_kind": "column",
            "transformation": "DIRECT",
            "transformation_subtype": "DIRECT",
            "inferred": False,
            "inference_status": "confirmed",
            "confidence": 1,
            "depth": 1,
            "path": ["app.orders.amount", "app.stg_orders.amt"],
            "direction": "upstream",
        }
    ]


def test_trace_compare_creates_workflow_of_compare_nodes() -> None:
    # project access -> upstream walk -> name available(None)-> INSERT workflows
    engine = _FakeEngine([{"id": "project-1"}, _trace_compare_walk_rows(), None])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/trace-compare",
        headers=_auth_headers(),
        json_body={
            "focus": "app.orders.amount",
            "source_id": "ds-prod",
            "target_id": "ds-test",
            "key_columns": ["id"],
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["workflow_id"] is not None
    assert payload["source_id"] == "ds-prod"
    assert payload["target_id"] == "ds-test"
    # 焦点表(hop0)+ 上游 1 跳(hop1)= 2 个 compare 节点
    assert payload["hop_count"] == 2
    assert [hop["node_id"] for hop in payload["hops"]] == ["hop0", "hop1"]
    assert payload["hops"][0]["table"] == "app.orders"
    assert payload["hops"][1]["table"] == "app.stg_orders"
    assert payload["hops"][1]["column"] == "amt"
    insert_statement = next(
        statement
        for statement in engine.statements
        if statement.startswith("INSERT INTO workflows")
    )
    assert "dag_jsonb" in insert_statement
    assert any(audit["action"] == "lineage_trace_compare" for audit in services.audits)


def test_trace_compare_dry_run_previews_without_persisting() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _trace_compare_walk_rows()])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/trace-compare",
        headers=_auth_headers(),
        json_body={
            "focus": "app.orders.amount",
            "source_id": "ds-prod",
            "target_id": "ds-test",
            "key_columns": ["id"],
            "dry_run": True,
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["workflow_id"] is None
    assert payload["hop_count"] == 2
    assert not any(statement.startswith("INSERT INTO workflows") for statement in engine.statements)


def test_trace_compare_no_upstream_returns_400() -> None:
    # 空 walk → 焦点无上游 → 400 no_upstream_lineage,不落库
    engine = _FakeEngine([{"id": "project-1"}, []])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/trace-compare",
        headers=_auth_headers(),
        json_body={
            "focus": "app.orders.amount",
            "source_id": "ds-prod",
            "target_id": "ds-test",
            "key_columns": ["id"],
        },
    )

    assert response.status_code == 400
    assert response.json()["error"] == "no_upstream_lineage"
    assert not any(statement.startswith("INSERT INTO workflows") for statement in engine.statements)


def test_trace_compare_requires_column_focus() -> None:
    engine = _FakeEngine([])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/trace-compare",
        headers=_auth_headers(),
        json_body={
            "focus": "plaintable",
            "source_id": "ds-prod",
            "target_id": "ds-test",
            "key_columns": ["id"],
        },
    )

    assert response.status_code == 400
    assert response.json()["error"] == "column_focus_required"
    assert engine.statements == []


def _lineage_run_row(parse_error_count: int = 2) -> dict[str, object]:
    return {
        "id": "run-1",
        "project_id": "project-1",
        "datasource_id": "ds-1",
        "dialect": "mysql",
        "source_ref": "script.sql",
        "sql_hash": "hash-1",
        "parser_version": "sqlglot-w1-v1",
        "status": "success",
        "parse_summary": {
            "table_edge_count": 0,
            "column_mapping_count": 0,
            "parse_error_count": parse_error_count,
        },
    }


def _ai_config_row() -> dict[str, object]:
    return {
        "enabled": True,
        "provider": "mock",
        "model": "mock-model",
        "base_url": None,
        "api_key_secret_ref": None,
        "max_auto_egress_level": 2,
        "l4_requires_optin": True,
        "enable_inference": True,
    }


def _ai_sql_engine() -> _FakeEngine:
    return _FakeEngine(
        [
            _datasource_row(),
            {"id": "project-1"},
            _ai_config_row(),
            _metadata_cache(
                [
                    {
                        "name": "id",
                        "type": "integer",
                        "driver_type": "INT",
                        "nullable": False,
                        "primary_key": True,
                        "comment": None,
                    },
                    {
                        "name": "name",
                        "type": "string",
                        "driver_type": "VARCHAR(64)",
                        "nullable": True,
                        "primary_key": False,
                        "comment": None,
                    },
                ]
            ),
        ]
    )


def test_lineage_analyze_ai_fallback_writes_inferred_edges_via_gateway(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _datasource_row(),
            [],  # empty metadata cache → unsupported_schema parse errors
            None,  # no cached run
            _lineage_run_row(),
            0,
            0,
            _ai_config_row(),
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))
    captured: dict[str, Any] = {}

    class _Gateway:
        def complete(self, prompt: str, context: Any, options: Any) -> Any:
            captured["prompt"] = prompt
            captured["context"] = context
            captured["options"] = options
            return AiResponse(
                content=json.dumps(
                    {
                        "table_edges": [
                            {
                                "source_table": "app.src",
                                "target_table": "app.tgt",
                                "confidence": 0.62,
                            }
                        ],
                        "column_edges": [
                            {
                                "source_table": "app.src",
                                "source_column": "id",
                                "target_table": "app.tgt",
                                "target_column": "id",
                                "confidence": 0.55,
                            }
                        ],
                    }
                ),
                provider="mock",
                model="mock-model",
            )

    monkeypatch.setattr(
        core_routes,
        "build_gateway_from_runtime_config",
        lambda runtime, **kwargs: _Gateway(),
    )

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/analyze",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "source_ref": "script.sql",
            "sql_text": (
                "INSERT INTO app.tgt (id, name) SELECT id, name FROM app.src "
                "WHERE name = 'top-secret-value'"
            ),
            "ai_fallback": True,
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["ai_fallback"] == {
        "requested": True,
        "executed": True,
        "reason": None,
        "inferred_table_edge_count": 1,
        "inferred_column_edge_count": 1,
        "provider": "mock",
        "model": "mock-model",
    }
    # R-AI:AI 只收脱敏片段(字面量不出境),L2 出站。
    context = captured["context"]
    assert len(context.items) == 1
    assert context.items[0].egress_level == 2
    assert context.items[0].redacted is True
    assert "top-secret-value" not in context.items[0].content
    assert "app.src" in context.items[0].content
    assert captured["options"].purpose == "lineage_ai_fallback"
    assert any("INSERT INTO lineage_edges" in statement for statement in engine.statements)
    assert any("INSERT INTO lineage_column_edges" in statement for statement in engine.statements)
    assert any("UPDATE lineage_runs" in statement for statement in engine.statements)
    assert any(
        audit["action"] == "lineage_ai_fallback" and audit["result"] == "success"
        for audit in services.audits
    )
    assert "top-secret-value" not in str(services.audits)


def test_lineage_analyze_ai_fallback_disabled_degrades_to_deterministic() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _datasource_row(),
            [],
            None,
            _lineage_run_row(),
            0,
            0,
            None,  # no ai_configs row → ai disabled
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/analyze",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "source_ref": "script.sql",
            "sql_text": "INSERT INTO app.tgt (id) SELECT id FROM app.src",
            "ai_fallback": True,
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["run_id"] == "run-1"
    assert payload["ai_fallback"]["requested"] is True
    assert payload["ai_fallback"]["executed"] is False
    assert payload["ai_fallback"]["reason"] == "ai_disabled"
    assert not any("INSERT INTO lineage_edges" in statement for statement in engine.statements)
    assert any(
        audit["action"] == "lineage_ai_fallback" and audit["result"] == "skipped"
        for audit in services.audits
    )


def _lineage_run_row_with_sql(
    sql_text: str = (
        "INSERT INTO app.tgt (id) SELECT id FROM app.src WHERE app.src.amount > 42999777"
    ),
) -> dict[str, object]:
    row = _lineage_run_row(parse_error_count=0)
    row["sql_text"] = sql_text
    return row


def _lineage_schema_cache_rows() -> list[dict[str, object]]:
    return [
        {
            "cache_level": "columns",
            "schema_name": "app",
            "table_name": "src",
            "payload": [{"name": "id", "type": "integer"}, {"name": "amount", "type": "integer"}],
        },
        {
            "cache_level": "columns",
            "schema_name": "app",
            "table_name": "tgt",
            "payload": [{"name": "id", "type": "integer"}],
        },
    ]


def test_lineage_ai_enrich_appends_interpretation_via_gateway(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},  # _require_project_access
            _lineage_run_row_with_sql(),  # lineage_runs
            _ai_config_row(),  # ai_configs
            _lineage_schema_cache_rows(),  # metadata_caches (schema context)
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))
    captured: dict[str, Any] = {}

    class _Gateway:
        def complete(self, prompt: str, context: Any, options: Any) -> Any:
            captured["prompt"] = prompt
            captured["context"] = context
            captured["options"] = options
            return AiResponse(
                content="全量重刷:app.tgt 采用 TRUNCATE+INSERT,建议核对源表完整性。",
                provider="mock",
                model="mock-model",
            )

    monkeypatch.setattr(
        core_routes,
        "build_gateway_from_runtime_config",
        lambda runtime, **kwargs: _Gateway(),
    )

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/runs/run-1/ai-enrich",
        headers=_auth_headers(),
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["ok"] is True
    assert payload["interpretation"].startswith("全量重刷")
    assert payload["provider"] == "mock"
    assert payload["egress_level"] == 2
    assert payload["created_at"] is not None
    # 出站只送 L2 结构画像,行值(字面量)绝不出境。
    context = captured["context"]
    assert len(context.items) == 1
    assert context.items[0].egress_level == 2
    # 只送结构画像:表名在,行值字面量(WHERE 数字)绝不出境。
    assert "42999777" not in context.items[0].content
    assert "app.src" in context.items[0].content
    assert captured["options"].purpose == "lineage_ai_enrich"
    # 只增不改:只写新表,绝不回写 lineage_runs / 确定性边表。
    assert any("INSERT INTO lineage_ai_enrichments" in statement for statement in engine.statements)
    assert not any("UPDATE lineage_runs" in statement for statement in engine.statements)
    assert not any("INSERT INTO lineage_edges" in statement for statement in engine.statements)
    assert any(
        audit["action"] == "lineage_ai_enrich" and audit["result"] == "success"
        for audit in services.audits
    )
    assert "top-secret-value" not in str(services.audits)


def test_lineage_ai_enrich_disabled_returns_structured_409() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _lineage_run_row_with_sql(),
            None,  # no ai_configs row → ai disabled
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/runs/run-1/ai-enrich",
        headers=_auth_headers(),
    )

    assert response.status_code == 409
    assert response.json()["error"] == "ai_disabled"
    assert not any(
        "INSERT INTO lineage_ai_enrichments" in statement for statement in engine.statements
    )
    assert any(
        audit["action"] == "lineage_ai_enrich" and audit["result"] == "skipped"
        for audit in services.audits
    )


def test_lineage_ai_enrich_get_returns_latest_or_none() -> None:
    stored = {
        "run_id": "run-1",
        "interpretation": "AI 解读文本",
        "provider": "mock",
        "model": "mock-model",
        "egress_level": 2,
        "created_at": datetime(2026, 7, 8, tzinfo=UTC),
    }
    engine = _FakeEngine([{"id": "project-1"}, _lineage_run_row(), stored])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/lineage/runs/run-1/ai-enrich",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["interpretation"] == "AI 解读文本"

    empty_engine = _FakeEngine([{"id": "project-1"}, _lineage_run_row(), None])
    empty_app = create_app(services=cast(ApiServices, _Services(empty_engine)))
    empty = AsgiClient(empty_app).get(
        "/api/projects/project-1/lineage/runs/run-1/ai-enrich",
        headers=_auth_headers(),
    )
    assert empty.status_code == 200
    assert empty.json()["ok"] is False
    assert empty.json()["error"] == "no_enrichment"


def test_lineage_analyze_ai_fallback_invalid_ai_response_degrades() -> None:
    # 真 DefaultAiGateway + MockProvider(回 "ok",不可解析为边 JSON)→ 优雅降级。
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _datasource_row(),
            [],
            None,
            _lineage_run_row(),
            0,
            0,
            _ai_config_row(),
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/analyze",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "source_ref": "script.sql",
            "sql_text": "INSERT INTO app.tgt (id) SELECT id FROM app.src",
            "ai_fallback": True,
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["ai_fallback"]["executed"] is False
    assert payload["ai_fallback"]["reason"] == "invalid_ai_response"
    assert payload["table_edge_count"] == 0
    assert not any("INSERT INTO lineage_edges" in statement for statement in engine.statements)
    assert not any("UPDATE lineage_runs" in statement for statement in engine.statements)


def test_lineage_analyze_without_ai_fallback_keeps_response_shape() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _datasource_row(),
            [],
            None,
            _lineage_run_row(),
            0,
            0,
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/lineage/analyze",
        headers=_auth_headers(),
        json_body={
            "datasource_id": "ds-1",
            "source_ref": "script.sql",
            "sql_text": "INSERT INTO app.tgt (id) SELECT id FROM app.src",
        },
    )

    assert response.status_code == 201
    assert response.json()["ai_fallback"] is None
    assert not any(audit["action"] == "lineage_ai_fallback" for audit in services.audits)


def _inferred_edge_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "id": "edge-1",
        "run_id": "run-1",
        "project_id": "project-1",
        "source_table": "app.src",
        "target_table": "app.tgt",
        "edge_kind": "table",
        "inferred": True,
        "inference_status": "inferred",
        "confidence": 0.66,
        "sql_hash": "hash-1",
    }
    row.update(overrides)
    return row


def test_lineage_edge_inference_confirm_table_edge() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _inferred_edge_row()])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PATCH",
        "/api/projects/project-1/lineage/edges/edge-1",
        headers=_auth_headers(),
        json_body={"inference_status": "confirmed"},
    )

    assert response.status_code == 200
    assert response.json() == {
        "edge_id": "edge-1",
        "edge_kind": "table",
        "inference_status": "confirmed",
        "inferred": True,
        "confidence": 0.66,
    }
    assert any(statement.startswith("UPDATE lineage_edges") for statement in engine.statements)
    assert any(audit["action"] == "lineage_edge_inference" for audit in services.audits)


def test_lineage_edge_inference_reject_column_edge() -> None:
    column_row = _inferred_edge_row(
        id="col-edge-1",
        source_column="id",
        target_column="id",
        transformation="DIRECT",
        transformation_subtype="DIRECT",
    )
    engine = _FakeEngine([{"id": "project-1"}, None, column_row])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).request(
        "PATCH",
        "/api/projects/project-1/lineage/edges/col-edge-1",
        headers=_auth_headers(),
        json_body={"inference_status": "rejected"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["edge_kind"] == "column"
    assert payload["inference_status"] == "rejected"
    assert any(
        statement.startswith("UPDATE lineage_column_edges") for statement in engine.statements
    )


def test_lineage_edge_inference_rejects_non_inferred_transition() -> None:
    engine = _FakeEngine(
        [{"id": "project-1"}, _inferred_edge_row(inference_status="confirmed", inferred=False)]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).request(
        "PATCH",
        "/api/projects/project-1/lineage/edges/edge-1",
        headers=_auth_headers(),
        json_body={"inference_status": "rejected"},
    )

    assert response.status_code == 409
    assert response.json()["error"] == "invalid_inference_transition"
    assert not any(statement.startswith("UPDATE lineage_edges") for statement in engine.statements)


def test_lineage_edge_inference_rejects_invalid_status_value() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _inferred_edge_row()])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).request(
        "PATCH",
        "/api/projects/project-1/lineage/edges/edge-1",
        headers=_auth_headers(),
        json_body={"inference_status": "inferred"},
    )

    assert response.status_code == 422


def test_lineage_edge_inference_cross_project_edge_is_not_found() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _inferred_edge_row(project_id="project-other")])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).request(
        "PATCH",
        "/api/projects/project-1/lineage/edges/edge-1",
        headers=_auth_headers(),
        json_body={"inference_status": "confirmed"},
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"
    assert not any(statement.startswith("UPDATE lineage_edges") for statement in engine.statements)


def test_workflow_create_contract_persists_spec_and_schedule_redundant_columns() -> None:
    engine = _FakeEngine([{"id": "project-1"}, None, _workflow_row()])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows",
        headers=_auth_headers(),
        json_body={
            "name": "nightly-report",
            "spec": _workflow_spec_payload(schedule={"cron": "0 2 * * *", "enabled": True}),
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["id"] == "wf-1"
    assert payload["name"] == "nightly-report"
    assert payload["enabled"] is True
    assert payload["schedule_cron"] == "0 2 * * *"
    assert payload["schedule_enabled"] is True
    assert payload["spec"]["nodes"][0]["job_kind"] == "sql_query"
    insert_statement = next(
        statement
        for statement in engine.statements
        if statement.startswith("INSERT INTO workflows")
    )
    assert "dag_jsonb" in insert_statement
    assert "schedule_cron" in insert_statement
    assert "schedule_enabled" in insert_statement
    assert any(audit["action"] == "workflow_create" for audit in services.audits)


def test_workflow_create_explicit_false_persists_disabled() -> None:
    disabled_row = _workflow_row()
    disabled_row["enabled"] = False
    engine = _FakeEngine([{"id": "project-1"}, None, disabled_row])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows",
        headers=_auth_headers(),
        json_body={
            "name": "disabled-report",
            "spec": _workflow_spec_payload(),
            "enabled": False,
        },
    )

    assert response.status_code == 201
    assert response.json()["enabled"] is False
    insert_statement = next(
        statement
        for statement in engine.executed
        if str(statement).lstrip().upper().startswith("INSERT INTO WORKFLOWS")
    )
    assert insert_statement.compile().params["enabled"] is False


def test_workflow_create_rejects_forbidden_node_kind_before_any_db_write() -> None:
    engine = _FakeEngine([])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows",
        headers=_auth_headers(),
        json_body={"name": "bad", "spec": _workflow_spec_payload(job_kind="shell")},
    )

    assert response.status_code == 400
    assert response.json()["error"] == "forbidden_node_kind"
    assert engine.statements == []


def test_workflow_create_rejects_unsupported_node_kind_distinct_from_forbidden() -> None:
    engine = _FakeEngine([])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows",
        headers=_auth_headers(),
        json_body={
            "name": "bad",
            "spec": _workflow_spec_payload(job_kind="scenario_materialize"),
        },
    )

    assert response.status_code == 400
    assert response.json()["error"] == "unsupported_node_kind"
    assert engine.statements == []


def test_workflow_create_returns_unknown_notify_target_error_code() -> None:
    spec = _workflow_spec_payload(job_kind="notify")
    nodes = cast(list[dict[str, object]], spec["nodes"])
    nodes[0]["payload"] = {"target_ids": ["missing-target"]}
    engine = _FakeEngine([])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows",
        headers=_auth_headers(),
        json_body={"name": "bad-notify", "spec": spec},
    )

    assert response.status_code == 400
    assert response.json()["error"] == "unknown_notify_target"
    assert engine.statements == []


def test_workflow_create_returns_code_only_node_output_reference_error() -> None:
    raw_expression = "nodes.query.bad-field"
    spec = _workflow_spec_payload()
    nodes = cast(list[dict[str, object]], spec["nodes"])
    nodes[0]["payload"] = {"sql": f"SELECT '${{{raw_expression}}}'"}
    engine = _FakeEngine([])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows",
        headers=_auth_headers(),
        json_body={"name": "bad-output-ref", "spec": spec},
    )

    assert response.status_code == 400
    assert response.json()["error"] == "invalid_node_output_reference"
    assert raw_expression not in response.body.decode("utf-8")
    assert engine.statements == []


def test_workflow_create_rejects_cyclic_dag() -> None:
    engine = _FakeEngine([])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows",
        headers=_auth_headers(),
        json_body={
            "name": "cyclic",
            "spec": {
                "nodes": [
                    {"id": "a", "job_kind": "sql_query", "timeout_seconds": 60},
                    {"id": "b", "job_kind": "sql_query", "timeout_seconds": 60},
                ],
                "edges": [
                    {"source": "a", "target": "b"},
                    {"source": "b", "target": "a"},
                ],
            },
        },
    )

    assert response.status_code == 400
    assert response.json()["error"] == "cycle_detected"
    assert engine.statements == []


def test_workflow_create_conflicting_name_returns_409() -> None:
    engine = _FakeEngine([{"id": "project-1"}, {"id": "wf-existing"}])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows",
        headers=_auth_headers(),
        json_body={"name": "nightly-report", "spec": _workflow_spec_payload()},
    )

    assert response.status_code == 409
    assert response.json()["error"] == "workflow_name_conflict"
    assert not any(statement.startswith("INSERT INTO workflows") for statement in engine.statements)


def test_workflow_create_without_project_access_returns_404() -> None:
    engine = _FakeEngine([None])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).post(
        "/api/projects/project-other/workflows",
        headers=_auth_headers(),
        json_body={"name": "nightly-report", "spec": _workflow_spec_payload()},
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"
    assert not any(statement.startswith("INSERT INTO workflows") for statement in engine.statements)


def test_workflow_list_contract_scopes_to_project_and_paginates() -> None:
    engine = _FakeEngine([{"id": "project-1"}, [_workflow_row()]])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflows",
        headers=_auth_headers(),
        params={"limit": 10, "offset": 0},
    )

    assert response.status_code == 200
    payload = response.json()
    assert [item["id"] for item in payload] == ["wf-1"]
    assert payload[0]["node_count"] == 1
    assert payload[0]["schedule_cron"] == "0 2 * * *"
    assert "spec" not in payload[0]
    statement = engine.statements[-1]
    assert "workflows.project_id" in statement
    assert "LIMIT" in statement
    assert "OFFSET" in statement


def test_workflow_detail_contract_returns_full_spec() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row()])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["id"] == "wf-1"
    assert payload["spec"]["nodes"][0]["id"] == "n1"
    assert payload["spec"]["schedule"] == {"cron": "0 2 * * *", "enabled": True}


def test_workflow_detail_outside_project_returns_404() -> None:
    engine = _FakeEngine([{"id": "project-1"}, None])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflows/wf-missing",
        headers=_auth_headers(),
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"


def test_workflow_update_contract_revalidates_spec_and_syncs_schedule_columns() -> None:
    updated_row = _workflow_row()
    updated_row["name"] = "renamed"
    updated_row["dag_jsonb"] = _workflow_spec_payload()
    updated_row["schedule_cron"] = None
    updated_row["schedule_enabled"] = False
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row(), None, updated_row])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
        json_body={"name": "renamed", "spec": _workflow_spec_payload()},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["name"] == "renamed"
    assert payload["schedule_cron"] is None
    assert payload["schedule_enabled"] is False
    update_statement = next(
        statement for statement in engine.statements if statement.startswith("UPDATE workflows")
    )
    assert "updated_at" in update_statement
    assert "schedule_cron" in update_statement
    assert any(audit["action"] == "workflow_update" for audit in services.audits)


def test_workflow_update_omitted_enabled_preserves_disabled() -> None:
    disabled_row = _workflow_row()
    disabled_row["enabled"] = False
    engine = _FakeEngine([{"id": "project-1"}, disabled_row, None, disabled_row])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
        json_body={"name": "nightly-report", "spec": _workflow_spec_payload()},
    )

    assert response.status_code == 200
    assert response.json()["enabled"] is False
    update_statement = next(
        statement
        for statement in engine.executed
        if str(statement).lstrip().upper().startswith("UPDATE WORKFLOWS")
    )
    assert "enabled" not in update_statement.compile().params


@pytest.mark.parametrize(("existing", "requested"), [(False, True), (True, False)])
def test_workflow_update_explicit_enabled_toggles(
    existing: bool,
    requested: bool,
) -> None:
    existing_row = _workflow_row()
    existing_row["enabled"] = existing
    updated_row = _workflow_row()
    updated_row["enabled"] = requested
    engine = _FakeEngine([{"id": "project-1"}, existing_row, None, updated_row])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
        json_body={
            "name": "nightly-report",
            "spec": _workflow_spec_payload(),
            "enabled": requested,
        },
    )

    assert response.status_code == 200
    assert response.json()["enabled"] is requested
    update_statement = next(
        statement
        for statement in engine.executed
        if str(statement).lstrip().upper().startswith("UPDATE WORKFLOWS")
    )
    assert update_statement.compile().params["enabled"] is requested


def test_workflow_update_rejects_forbidden_node_kind_before_any_db_write() -> None:
    engine = _FakeEngine([])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
        json_body={"name": "bad", "spec": _workflow_spec_payload(job_kind="shell")},
    )

    assert response.status_code == 400
    assert response.json()["error"] == "forbidden_node_kind"
    assert engine.statements == []


def _workflow_row_with_variables(**variables: object) -> dict[str, object]:
    row = _workflow_row()
    dag = dict(cast(dict[str, Any], row["dag_jsonb"]))
    dag["variables"] = dict(variables)
    row["dag_jsonb"] = dag
    return row


def test_workflow_update_omitting_notifications_preserves_existing_targets() -> None:
    # GET-then-PUT 少带 notifications:既有 notify 目标 + 其 SecretStore ref 必须存活,
    # 不能被静默清空(避免孤儿化 url_secret_ref)。
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_row_with_notification(),
            None,
            _workflow_row_with_notification(),
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
        json_body={"name": "nightly-report", "spec": _workflow_spec_payload()},
    )

    assert response.status_code == 200
    dag = _persisted_workflow_dag(engine)
    assert len(dag["notifications"]) == 1
    assert dag["notifications"][0]["id"] == "nt-1"
    assert dag["notifications"][0]["url_secret_ref"] == "secret-existing"
    # PUT 必须与通知子资源写入串行化,否则两边从同一旧 dag_jsonb 回写会丢目标。
    _assert_workflow_row_read_uses_for_update(engine)


def test_workflow_update_omitting_notifications_validates_after_preserving_notify_refs() -> None:
    existing = _workflow_row_with_notify_node()
    request_spec = dict(cast(dict[str, Any], existing["dag_jsonb"]))
    request_spec.pop("notifications")
    engine = _FakeEngine([{"id": "project-1"}, existing, None, existing])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
        json_body={"name": "nightly-report", "spec": request_spec},
    )

    assert response.status_code == 200
    dag = _persisted_workflow_dag(engine)
    assert dag["nodes"][0]["payload"]["target_ids"] == ["nt-1"]
    assert dag["notifications"][0]["id"] == "nt-1"


def test_workflow_update_explicit_empty_notifications_rejects_remaining_notify_refs() -> None:
    request_spec = dict(cast(dict[str, Any], _workflow_row_with_notify_node()["dag_jsonb"]))
    request_spec["notifications"] = []
    engine = _FakeEngine([])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
        json_body={"name": "nightly-report", "spec": request_spec},
    )

    assert response.status_code == 400


def test_workflow_update_explicit_empty_notifications_clears_targets() -> None:
    # 显式 notifications: [] 表达"我要清空",按原样处理(尊重显式意图)。
    engine = _FakeEngine(
        [{"id": "project-1"}, _workflow_row_with_notification(), None, _workflow_row()]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))
    spec = _workflow_spec_payload()
    spec["notifications"] = []

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
        json_body={"name": "nightly-report", "spec": spec},
    )

    assert response.status_code == 200
    assert _persisted_workflow_dag(engine)["notifications"] == []


def test_workflow_update_omitting_variables_preserves_existing() -> None:
    existing = _workflow_row_with_variables(region="us-east")
    engine = _FakeEngine([{"id": "project-1"}, existing, None, existing])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
        json_body={"name": "nightly-report", "spec": _workflow_spec_payload()},
    )

    assert response.status_code == 200
    assert _persisted_workflow_dag(engine)["variables"] == {"region": "us-east"}


def test_workflow_update_explicit_empty_variables_clears() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_row_with_variables(region="us-east"),
            None,
            _workflow_row(),
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))
    spec = _workflow_spec_payload()
    spec["variables"] = {}

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
        json_body={"name": "nightly-report", "spec": spec},
    )

    assert response.status_code == 200
    assert _persisted_workflow_dag(engine)["variables"] == {}


def test_workflow_update_still_full_replaces_nodes() -> None:
    # 控制用例:nodes/edges 仍是全量替换(preserve 只作用于 notifications/variables)。
    engine = _FakeEngine(
        [{"id": "project-1"}, _workflow_row_with_notification(), None, _workflow_row()]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))
    spec = _workflow_spec_payload()
    spec["nodes"] = [{"id": "n2", "job_kind": "sql_query", "timeout_seconds": 30}]

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
        json_body={"name": "nightly-report", "spec": spec},
    )

    assert response.status_code == 200
    dag = _persisted_workflow_dag(engine)
    assert [node["id"] for node in dag["nodes"]] == ["n2"]
    assert dag["nodes"][0]["timeout_seconds"] == 30


def test_workflow_delete_contract_removes_definition() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row()])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "DELETE",
        "/api/projects/project-1/workflows/wf-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 204
    assert any(statement.startswith("DELETE FROM workflows") for statement in engine.statements)
    assert any(audit["action"] == "workflow_delete" for audit in services.audits)


# ── Workflow run 通知目标配置(C-9 PR4)────────────────────────────────────────
# ★ 明文 webhook / 企微 url(含 ?key=token)只能进 SecretStore,配置 / 响应 / 日志
#   永远只留 SecretRef(R2/R5)。含 token 的固定 url 供各测试断言"明文不外泄"。
_NOTIFY_URL = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=secret-token-xyz"
_NOTIFY_URL_ROTATED = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=rotated-token-abc"
_SMTP_PASSWORD = "smtp-plaintext-pw-789"  # 断言绝不外泄(R2/R5)


def _workflow_row_with_notification(
    *,
    target_id: str = "nt-1",
    url_secret_ref: str = "secret-existing",
) -> dict[str, object]:
    row = _workflow_row()
    dag = dict(cast(dict[str, Any], row["dag_jsonb"]))
    dag["notifications"] = [
        {
            "id": target_id,
            "channel": "wecom",
            "url_secret_ref": url_secret_ref,
            "events": ["failed"],
            "enabled": True,
            "timeout_seconds": 5,
        }
    ]
    row["dag_jsonb"] = dag
    return row


def _workflow_row_with_notify_node() -> dict[str, object]:
    row = _workflow_row_with_notification()
    dag = dict(cast(dict[str, Any], row["dag_jsonb"]))
    dag["nodes"] = [
        {
            "id": "notify-1",
            "job_kind": "notify",
            "payload": {"target_ids": ["nt-1"]},
            "timeout_seconds": 60,
        }
    ]
    row["dag_jsonb"] = dag
    return row


def _workflow_run_referencing_notification(
    *,
    target_id: str = "nt-1",
    status: str = "running",
) -> dict[str, object]:
    return {
        "status": status,
        "payload": {
            "workflow_id": "wf-1",
            "spec": {
                "nodes": [
                    {
                        "id": "notify-1",
                        "job_kind": "notify",
                        "payload": {"target_ids": [target_id]},
                        "timeout_seconds": 60,
                    }
                ],
                "edges": [],
                "notifications": [
                    {
                        "id": target_id,
                        "channel": "wecom",
                        "url_secret_ref": "secret-frozen",
                    }
                ],
            },
        },
    }


def _persisted_workflow_dag(engine: _FakeEngine) -> dict[str, Any]:
    """从最近一条 UPDATE workflows 语句里 compile 出持久化的 dag_jsonb。"""
    statement = next(
        stmt
        for stmt in reversed(engine.executed)
        if str(stmt).lstrip().upper().startswith("UPDATE WORKFLOWS")
    )
    return dict(statement.compile().params["dag_jsonb"])


def _assert_workflow_row_read_uses_for_update(engine: _FakeEngine) -> None:
    workflow_reads = [
        str(statement).upper()
        for statement in engine.executed
        if str(statement).lstrip().upper().startswith("SELECT")
        and "FROM WORKFLOWS" in str(statement).upper()
    ]
    assert workflow_reads
    assert any("FOR UPDATE" in statement for statement in workflow_reads)


def test_workflow_notify_create_stores_ref_and_hides_url() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row()])
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows/wf-1/notifications",
        headers=_auth_headers(),
        json_body={"channel": "wecom", "url": _NOTIFY_URL, "events": ["failed", "success"]},
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["channel"] == "wecom"
    assert payload["events"] == ["failed", "success"]
    assert payload["enabled"] is True
    assert "url" not in payload  # R5:响应绝不回显 url
    # R2:明文 url 进了 SecretStore,配置里只留返回的 ref
    assert secret_store.stored == [_NOTIFY_URL]
    # R5:明文 url / token 不出现在响应体或任何 SQL 语句文本
    assert _NOTIFY_URL not in response.body.decode("utf-8")
    assert "secret-token" not in response.body.decode("utf-8")
    assert not any(_NOTIFY_URL in statement for statement in engine.statements)
    # 持久化的 spec.notifications 只带 url_secret_ref,绝无明文 url
    dag = _persisted_workflow_dag(engine)
    assert len(dag["notifications"]) == 1
    target = dag["notifications"][0]
    assert target["url_secret_ref"] == "secret-new"
    assert "url" not in target
    assert _NOTIFY_URL not in json.dumps(dag)
    assert any(audit["action"] == "workflow_notify_add" for audit in services.audits)
    # 先锁 workflow 行再读取 spec,避免并发 PUT 用旧快照覆盖新 target、孤儿化 ref。
    _assert_workflow_row_read_uses_for_update(engine)


def test_workflow_notify_create_unknown_workflow_returns_404_without_orphan_secret() -> None:
    engine = _FakeEngine([{"id": "project-1"}, None])
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows/wf-missing/notifications",
        headers=_auth_headers(),
        json_body={"channel": "webhook", "url": _NOTIFY_URL},
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"
    # ★ 校验失败先于 store_secret:不留孤儿 secret
    assert secret_store.stored == []


def test_workflow_notify_create_email_stores_password_ref_and_hides_it() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row()])
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows/wf-1/notifications",
        headers=_auth_headers(),
        json_body={
            "channel": "email",
            "smtp_host": "smtp.example.com",
            "smtp_port": 2525,
            "smtp_from": "ops@example.com",
            "smtp_to": ["team@example.com"],
            "smtp_user": "ops",
            "smtp_password": _SMTP_PASSWORD,
            "events": ["failed"],
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["channel"] == "email"
    assert payload["smtp_host"] == "smtp.example.com"
    assert payload["smtp_to"] == ["team@example.com"]
    # ★ R2/R5:响应绝不回显明文密码,也无任何 password 字段
    assert "smtp_password" not in payload
    assert "password" not in payload
    assert _SMTP_PASSWORD not in response.body.decode("utf-8")
    # R2:明文密码进 SecretStore;配置只留 password_secret_ref
    assert secret_store.stored == [_SMTP_PASSWORD]
    dag = _persisted_workflow_dag(engine)
    target = dag["notifications"][0]
    assert target["password_secret_ref"] == "secret-new"
    assert target["url_secret_ref"] is None
    assert "smtp_password" not in target
    assert _SMTP_PASSWORD not in json.dumps(dag)
    assert not any(_SMTP_PASSWORD in statement for statement in engine.statements)
    # 审计只带 target_id / channel,无密码
    add_audits = [a for a in services.audits if a["action"] == "workflow_notify_add"]
    assert add_audits
    assert _SMTP_PASSWORD not in json.dumps(add_audits, default=str)


def test_workflow_notify_update_replaces_ref_and_deletes_old() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row_with_notification(), []])
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1/notifications/nt-1",
        headers=_auth_headers(),
        json_body={"url": _NOTIFY_URL_ROTATED, "enabled": False, "events": ["all"]},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["id"] == "nt-1"
    assert payload["enabled"] is False
    assert payload["events"] == ["all"]
    assert "url" not in payload
    # R2:新明文 url 进 SecretStore;旧 secret 被清理
    assert secret_store.stored == [_NOTIFY_URL_ROTATED]
    assert secret_store.deleted == ["secret-existing"]
    dag = _persisted_workflow_dag(engine)
    assert dag["notifications"][0]["url_secret_ref"] == "secret-new"
    assert _NOTIFY_URL_ROTATED not in json.dumps(dag)
    assert any(audit["action"] == "workflow_notify_update" for audit in services.audits)
    _assert_workflow_row_read_uses_for_update(engine)


def test_workflow_notify_update_without_url_keeps_ref_and_stores_nothing() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row_with_notification(), []])
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1/notifications/nt-1",
        headers=_auth_headers(),
        json_body={"enabled": False},
    )

    assert response.status_code == 200
    assert secret_store.stored == []
    assert secret_store.deleted == []
    dag = _persisted_workflow_dag(engine)
    assert dag["notifications"][0]["url_secret_ref"] == "secret-existing"
    assert dag["notifications"][0]["enabled"] is False


def test_workflow_notify_update_unknown_target_returns_404() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row_with_notification()])
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1/notifications/nt-missing",
        headers=_auth_headers(),
        json_body={"enabled": False},
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"
    assert secret_store.stored == []


def test_workflow_notify_update_rejects_target_referenced_by_active_frozen_run() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_row_with_notification(),
            [_workflow_run_referencing_notification()],
        ]
    )
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1/notifications/nt-1",
        headers=_auth_headers(),
        json_body={"url": _NOTIFY_URL_ROTATED},
    )

    assert response.status_code == 409
    assert response.json()["error"] == "workflow_notify_target_active_run"
    assert secret_store.stored == []
    assert secret_store.deleted == []
    assert not any(
        str(statement).lstrip().upper().startswith("UPDATE WORKFLOWS")
        for statement in engine.executed
    )


def test_workflow_notify_update_rejects_target_referenced_by_pending_sensor_check() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_row_with_notification(),
            [_workflow_run_referencing_notification(status="pending")],
        ]
    )
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "PUT",
        "/api/projects/project-1/workflows/wf-1/notifications/nt-1",
        headers=_auth_headers(),
        json_body={"url": _NOTIFY_URL_ROTATED},
    )

    assert response.status_code == 409
    assert response.json()["error"] == "workflow_notify_target_active_run"
    job_read = next(
        statement
        for statement in engine.executed
        if str(statement).lstrip().upper().startswith("SELECT")
        and "FROM JOBS" in str(statement).upper()
    )
    kind_values = [
        value for key, value in job_read.compile().params.items() if key.startswith("kind_")
    ]
    assert kind_values == [["workflow_run", "workflow_sensor_check"]]
    assert secret_store.stored == []
    assert secret_store.deleted == []


def test_workflow_notify_delete_removes_target_and_secret() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row_with_notification(), []])
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "DELETE",
        "/api/projects/project-1/workflows/wf-1/notifications/nt-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 204
    assert secret_store.deleted == ["secret-existing"]
    assert _persisted_workflow_dag(engine)["notifications"] == []
    assert any(audit["action"] == "workflow_notify_remove" for audit in services.audits)
    _assert_workflow_row_read_uses_for_update(engine)


def test_workflow_notify_delete_rejects_target_referenced_by_current_dag() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row_with_notify_node()])
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "DELETE",
        "/api/projects/project-1/workflows/wf-1/notifications/nt-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 409
    assert response.json()["error"] == "workflow_notify_target_in_use"
    assert secret_store.deleted == []
    assert not any(
        str(statement).lstrip().upper().startswith("UPDATE WORKFLOWS")
        for statement in engine.executed
    )


def test_workflow_notify_delete_rejects_target_referenced_by_active_frozen_run() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_row_with_notification(),
            [_workflow_run_referencing_notification()],
        ]
    )
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "DELETE",
        "/api/projects/project-1/workflows/wf-1/notifications/nt-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 409
    assert response.json()["error"] == "workflow_notify_target_active_run"
    assert secret_store.deleted == []
    assert not any(
        str(statement).lstrip().upper().startswith("UPDATE WORKFLOWS")
        for statement in engine.executed
    )


def test_workflow_notify_delete_unknown_target_returns_404() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row_with_notification()])
    secret_store = _SecretStore()
    services = _Services(engine, secret_store=secret_store)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).request(
        "DELETE",
        "/api/projects/project-1/workflows/wf-1/notifications/nt-missing",
        headers=_auth_headers(),
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"
    assert secret_store.deleted == []


def test_workflow_run_trigger_contract_enqueues_workflow_run_job() -> None:
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row()])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows/wf-1/runs",
        headers=_auth_headers(),
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["workflow_id"] == "wf-1"
    assert payload["run_id"] == payload["job_id"]  # WorkflowRun 本身就是 job(ADR-0009)
    backend = services.job_backend
    assert len(backend.enqueued) == 1
    job = backend.enqueued[0]
    assert job.kind is JobKind.WORKFLOW_RUN
    assert job.id == payload["run_id"]
    assert job.project_id == "project-1"
    assert job.payload["workflow_id"] == "wf-1"
    assert job.payload["trigger"] == "manual"
    # spec 快照进 payload:执行期不回读 workflows 表
    spec = job.payload["spec"]
    assert isinstance(spec, dict)
    assert spec["nodes"][0]["id"] == "n1"
    # when 变量在触发时刻冻结进 payload:执行器与状态查询同源,决策不随时间漂移
    frozen = job.payload["when_variables"]
    assert isinstance(frozen, dict)
    assert set(frozen) == {"today", "now", "year", "month", "day"}
    assert all(isinstance(value, str) for value in frozen.values())
    assert any(audit["action"] == "workflow_run_trigger" for audit in services.audits)
    _assert_workflow_row_read_uses_for_update(engine)


def test_workflow_run_trigger_disabled_workflow_returns_409() -> None:
    row = _workflow_row()
    row["enabled"] = False
    engine = _FakeEngine([{"id": "project-1"}, row])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows/wf-1/runs",
        headers=_auth_headers(),
    )

    assert response.status_code == 409
    assert response.json()["error"] == "workflow_disabled"
    assert services.job_backend.enqueued == []


def test_workflow_run_trigger_missing_workflow_returns_404() -> None:
    engine = _FakeEngine([{"id": "project-1"}, None])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows/wf-missing/runs",
        headers=_auth_headers(),
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"
    assert services.job_backend.enqueued == []


def test_workflow_run_trigger_merges_variable_sources_into_snapshot() -> None:
    # C-7 PR2:合并优先级 builtin < spec.variables < 触发时;触发覆盖 spec,builtin 不被覆盖
    row = _workflow_row()
    cast(dict[str, object], row["dag_jsonb"])["variables"] = {
        "env": "staging",
        "region": "ap-east-1",
    }
    engine = _FakeEngine([{"id": "project-1"}, row])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows/wf-1/runs",
        headers=_auth_headers(),
        json_body={"variables": {"env": "prod"}},
    )

    assert response.status_code == 201
    job = services.job_backend.enqueued[0]
    frozen = job.payload["when_variables"]
    # 内置 5 变量 + spec/触发变量都进同一份快照(服务 when + ${var} 插值)
    assert {"today", "now", "year", "month", "day"} <= set(frozen)
    assert frozen["env"] == "prod"  # 触发 > spec
    assert frozen["region"] == "ap-east-1"  # spec 保留
    assert all(isinstance(value, str) for value in frozen.values())


def test_workflow_run_trigger_without_body_snapshot_is_builtins_only() -> None:
    # 无 body → 与既往一致:快照只含内置变量
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row()])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows/wf-1/runs",
        headers=_auth_headers(),
    )

    assert response.status_code == 201
    frozen = services.job_backend.enqueued[0].payload["when_variables"]
    assert set(frozen) == {"today", "now", "year", "month", "day"}


def test_workflow_run_trigger_rejects_unsafe_variable_value_name_only() -> None:
    # 危险取值(引号 + 分号)→ 400 unsafe_variable_value;不 enqueue;错误不回显取值(R5)
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row()])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows/wf-1/runs",
        headers=_auth_headers(),
        json_body={"variables": {"inj": "x'; DROP TABLE t"}},
    )

    assert response.status_code == 400
    body = response.json()
    assert body["error"] == "unsafe_variable_value"
    assert "inj" in body["message"]
    # ★ R5:响应绝不回显危险取值
    assert "DROP TABLE" not in response.body.decode("utf-8")
    assert services.job_backend.enqueued == []


def test_workflow_run_trigger_accepts_list_variable_into_snapshot() -> None:
    # C-7 PR3:list 型触发变量并入快照(供 ${var | sql_in} / csv 展开)
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row()])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows/wf-1/runs",
        headers=_auth_headers(),
        json_body={"variables": {"ids": ["1", "2", "3"], "env": "prod"}},
    )

    assert response.status_code == 201
    frozen = services.job_backend.enqueued[0].payload["when_variables"]
    assert frozen["ids"] == ["1", "2", "3"]
    assert frozen["env"] == "prod"


def test_workflow_run_trigger_rejects_unsafe_list_element_name_only() -> None:
    # list 元素含危险取值 → 400 unsafe_variable_value;不 enqueue;不回显取值(R5)
    engine = _FakeEngine([{"id": "project-1"}, _workflow_row()])
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflows/wf-1/runs",
        headers=_auth_headers(),
        json_body={"variables": {"ids": ["ok", "x'; DROP TABLE t"]}},
    )

    assert response.status_code == 400
    body = response.json()
    assert body["error"] == "unsafe_variable_value"
    assert "ids" in body["message"]
    assert "DROP TABLE" not in response.body.decode("utf-8")
    assert services.job_backend.enqueued == []


def test_workflow_run_status_contract_returns_run_and_node_states() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_run_job_row(status="running"),
            [_workflow_child_job_row("n1", status="success")],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflow-runs/run-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["run_id"] == "run-1"
    assert payload["workflow_id"] == "wf-1"
    assert payload["project_id"] == "project-1"
    assert payload["status"] == "running"
    assert len(payload["nodes"]) == 1
    node = payload["nodes"][0]
    assert node["node_id"] == "n1"
    assert node["job_kind"] == "sql_query"
    assert node["status"] == "success"
    assert node["job_id"] == "child-1"
    assert node["attempts"] == 0


@pytest.mark.parametrize(
    ("job_kind", "metadata", "expected_kind_outputs"),
    [
        (
            "sql_query",
            {"result_set_id": "rs-1", "loaded_rows": 7},
            {"result_set_id": "rs-1", "loaded_rows": 7},
        ),
        (
            "sql_explain",
            {"result_set_id": "rs-plan", "loaded_rows": 1},
            {"result_set_id": "rs-plan", "loaded_rows": 1},
        ),
        (
            "compare_run",
            {
                "run_id": "compare-1",
                "bucket_counts": {"same": 11, "only_source": 2, "only_target": 3, "diff": 4},
            },
            {
                "run_id": "compare-1",
                "same_count": 11,
                "only_source_count": 2,
                "only_target_count": 3,
                "diff_count": 4,
            },
        ),
        (
            "lineage_analyze",
            {
                "lineage_run_id": "lineage-1",
                "cached": False,
                "table_edge_count": 5,
                "column_edge_count": 6,
                "parse_error_count": 1,
            },
            {
                "run_id": "lineage-1",
                "cached": False,
                "table_edge_count": 5,
                "column_edge_count": 6,
                "parse_error_count": 1,
            },
        ),
        ("export_excel", {}, {}),
        ("notify", {"sent_count": 2}, {"sent_count": 2}),
        ("sleep", {"duration_seconds": 30}, {"duration_seconds": 30}),
        ("branch", {"selected_target": "n2"}, {"selected_target": "n2"}),
    ],
)
def test_workflow_run_status_returns_only_safe_scalar_node_outputs(
    job_kind: str,
    metadata: dict[str, object],
    expected_kind_outputs: dict[str, object],
) -> None:
    sentinels = {
        "rows": ["rows-sentinel"],
        "password": "password-sentinel",
        "token": "token-sentinel",
        "message": "message-sentinel",
        "nested": {"secret": "nested-sentinel"},
        "unknown": "unknown-sentinel",
    }
    status_spec = _workflow_status_spec(job_kind)
    WorkflowSpec.model_validate(status_spec)
    run_row = _workflow_run_job_row(status="running")
    cast(dict[str, Any], run_row["payload"])["spec"] = status_spec
    child = _workflow_child_job_row(
        "n1",
        status="success",
        job_kind=job_kind,
        result_ref={
            "backend": "test",
            "uri": "uri-sentinel",
            "metadata": {**metadata, **sentinels},
        },
    )
    engine = _FakeEngine([{"id": "project-1"}, run_row, [child]])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflow-runs/run-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    node = response.json()["nodes"][0]
    assert node["outputs"] == {
        **expected_kind_outputs,
        "status": "success",
        "job_id": "child-1",
        "error_code": None,
    }
    serialized = response.body.decode("utf-8")
    for sentinel in (
        "uri-sentinel",
        "rows-sentinel",
        "password-sentinel",
        "token-sentinel",
        "message-sentinel",
        "nested-sentinel",
        "unknown-sentinel",
    ):
        assert sentinel not in serialized


def test_workflow_run_status_maps_child_error_code_into_common_outputs() -> None:
    child = _workflow_child_job_row(
        "n1",
        status="failed",
        error_code="sql_failed",
    )
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_run_job_row(status="running"),
            [child],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflow-runs/run-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    assert response.json()["nodes"][0]["outputs"] == {
        "status": "failed",
        "job_id": "child-1",
        "error_code": "sql_failed",
    }


def test_workflow_run_status_malformed_result_ref_keeps_safe_common_outputs() -> None:
    child = _workflow_child_job_row(
        "n1",
        status="success",
        result_ref={
            "backend": "test",
            "uri": ["malformed-uri-sentinel"],
            "metadata": {"result_set_id": "metadata-sentinel"},
        },
    )
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_run_job_row(status="running"),
            [child],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflow-runs/run-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    assert response.json()["nodes"][0]["outputs"] == {
        "status": "success",
        "job_id": "child-1",
        "error_code": None,
    }
    assert "malformed-uri-sentinel" not in response.body.decode("utf-8")
    assert "metadata-sentinel" not in response.body.decode("utf-8")


def test_workflow_run_status_invalid_spec_fallback_uses_safe_projection() -> None:
    run_row = _workflow_run_job_row(status="running")
    cast(dict[str, Any], run_row["payload"])["spec"] = {"nodes": "invalid-spec"}
    child = _workflow_child_job_row(
        "n1",
        status="success",
        result_ref={
            "backend": "test",
            "uri": "fallback-uri-sentinel",
            "metadata": {
                "result_set_id": "rs-fallback",
                "loaded_rows": 3,
                "rows": ["fallback-row-sentinel"],
                "token": "fallback-token-sentinel",
            },
        },
    )
    engine = _FakeEngine([{"id": "project-1"}, run_row, [child]])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflow-runs/run-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    assert response.json()["nodes"][0]["outputs"] == {
        "result_set_id": "rs-fallback",
        "loaded_rows": 3,
        "status": "success",
        "job_id": "child-1",
        "error_code": None,
    }
    serialized = response.body.decode("utf-8")
    assert "fallback-uri-sentinel" not in serialized
    assert "fallback-row-sentinel" not in serialized
    assert "fallback-token-sentinel" not in serialized


def test_workflow_run_status_uses_configured_node_default_max_retries() -> None:
    # RetryPolicy=None 的失败节点:worker 若配 default_max_retries>=1 会重排(retry_wait)。
    # API 状态查询必须与 worker 同源读该配置,而非硬编码 0(否则展示为终态 failed,漂移)。
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_run_job_row(status="running"),
            [_workflow_child_job_row("n1", status="failed")],
        ]
    )
    services = _Services(engine, workflow_node_default_max_retries=1)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflow-runs/run-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    node = response.json()["nodes"][0]
    assert node["node_id"] == "n1"
    # 配置生效 → 继承重试 → retry_wait;若仍硬编码 0 则会是 failed
    assert node["status"] == "retry_wait"


def test_workflow_run_status_default_max_retries_zero_keeps_failed_terminal() -> None:
    # 默认口径 0:同一失败节点展示为终态 failed(回归护栏,确认非无条件 retry_wait)。
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_run_job_row(status="running"),
            [_workflow_child_job_row("n1", status="failed")],
        ]
    )
    services = _Services(engine, workflow_node_default_max_retries=0)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflow-runs/run-1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    assert response.json()["nodes"][0]["status"] == "failed"


def test_workflow_run_status_not_found_returns_404() -> None:
    engine = _FakeEngine([{"id": "project-1"}, None])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflow-runs/run-missing",
        headers=_auth_headers(),
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"


def test_workflow_runs_list_contract_returns_runs_desc() -> None:
    # PR0(workflow 前端前置):列某 workflow 的历史 run(倒序 + has_more 分页口径)。
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_row(),  # workflow 存在性检查
            [
                _workflow_run_job_row(status="success"),
                _workflow_run_job_row(status="running"),
            ],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflows/wf-1/runs",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["workflow_id"] == "wf-1"
    assert payload["has_more"] is False
    assert payload["limit"] == 20
    assert len(payload["runs"]) == 2
    first = payload["runs"][0]
    assert first["run_id"] == "run-1"
    assert first["job_id"] == first["run_id"]  # run 即 job(ADR-0009)
    assert first["status"] == "success"
    # 节点明细不在列表项里(走单 run 状态端点)
    assert "nodes" not in first
    assert "outputs" not in first


def test_workflow_runs_list_has_more_when_over_limit() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_row(),
            [  # limit=1 → 后端取 limit+1=2 行判 has_more(只回 1 行)
                _workflow_run_job_row(status="success"),
                _workflow_run_job_row(status="running"),
            ],
        ]
    )
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflows/wf-1/runs?limit=1",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["limit"] == 1
    assert payload["has_more"] is True
    assert len(payload["runs"]) == 1


def test_workflow_runs_list_missing_workflow_returns_404() -> None:
    engine = _FakeEngine([{"id": "project-1"}, None])
    app = create_app(services=cast(ApiServices, _Services(engine)))

    response = AsgiClient(app).get(
        "/api/projects/project-1/workflows/wf-missing/runs",
        headers=_auth_headers(),
    )

    assert response.status_code == 404
    assert response.json()["error"] == "not_found"


def test_workflow_run_cancel_contract_propagates_to_unfinished_children() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_run_job_row(status="running"),
            [
                _workflow_child_job_row("n1", status="running", job_id="child-1"),
                _workflow_child_job_row("n2", status="success", job_id="child-2"),
                _workflow_child_job_row("n3", status="pending", job_id="child-3"),
            ],
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflow-runs/run-1:cancel",
        headers=_auth_headers(),
    )

    assert response.status_code == 200
    assert response.json() == {"cancelled": True}
    backend = services.job_backend
    # run 软取消 + 未终态子 job 传播;已 success 的子 job 不动
    assert backend.cancel_requested == ["run-1", "child-1", "child-3"]
    assert backend.cancelled_pending == [("child-3", "workflow run cancelled")]
    assert any(audit["action"] == "workflow_run_cancel" for audit in services.audits)


def test_workflow_run_cancel_on_terminal_run_returns_409() -> None:
    engine = _FakeEngine(
        [
            {"id": "project-1"},
            _workflow_run_job_row(status="success"),
            [
                _workflow_child_job_row("n1", status="success", job_id="child-1"),
            ],
        ]
    )
    services = _Services(engine)
    app = create_app(services=cast(ApiServices, services))

    response = AsgiClient(app).post(
        "/api/projects/project-1/workflow-runs/run-1:cancel",
        headers=_auth_headers(),
    )

    assert response.status_code == 409
    assert response.json()["error"] == "workflow_run_terminal"
    # 已终态 run 不软取消,也不传播到子 job
    assert services.job_backend.cancel_requested == []
    assert services.job_backend.cancelled_pending == []
    assert not any(audit["action"] == "workflow_run_cancel" for audit in services.audits)


def _workflow_run_job_row(*, status: str = "running") -> dict[str, object]:
    return {
        "id": "run-1",
        "kind": "workflow_run",
        "status": status,
        "owner_user_id": "user-1",
        "project_id": "project-1",
        "payload": {
            "workflow_id": "wf-1",
            "workflow_name": "nightly-report",
            "trigger": "manual",
            "spec": _workflow_spec_payload(),
        },
        "error": None,
        "retry_count": 0,
        "created_at": _dt(1),
        "started_at": _dt(2),
        "finished_at": None,
    }


def _workflow_child_job_row(
    node_id: str,
    *,
    status: str,
    job_id: str = "child-1",
    job_kind: str = "sql_query",
    error_code: str | None = None,
    result_ref: object | None = None,
) -> dict[str, object]:
    return {
        "id": job_id,
        "kind": job_kind,
        "status": status,
        "owner_user_id": "user-1",
        "project_id": "project-1",
        "payload": {"sql": "SELECT 1", "workflow_node_id": node_id},
        "error": None,
        "error_code": error_code,
        "result_ref": result_ref,
        "retry_count": 0,
        "created_at": _dt(3),
        "started_at": _dt(4),
        "finished_at": _dt(5) if status in {"success", "failed", "cancelled"} else None,
        "parent_workflow_run_id": "run-1",
    }


def _workflow_spec_payload(
    *,
    job_kind: str = "sql_query",
    schedule: dict[str, object] | None = None,
) -> dict[str, object]:
    return {
        "nodes": [{"id": "n1", "job_kind": job_kind, "timeout_seconds": 60}],
        "edges": [],
        "schedule": schedule,
    }


def _workflow_status_spec(job_kind: str) -> dict[str, object]:
    payload_by_kind: dict[str, dict[str, object]] = {
        "notify": {"target_ids": ["nt-1"]},
        "sleep": {"duration_seconds": 30},
    }
    spec: dict[str, object] = {
        "nodes": [
            {
                "id": "n1",
                "job_kind": job_kind,
                "payload": payload_by_kind.get(job_kind, {}),
                "timeout_seconds": 60,
            }
        ],
        "edges": [],
    }
    if job_kind == "branch":
        cast(list[dict[str, object]], spec["nodes"]).extend(
            [
                {"id": "n2", "job_kind": "sql_query", "payload": {}, "timeout_seconds": 60},
                {"id": "n3", "job_kind": "sql_query", "payload": {}, "timeout_seconds": 60},
            ]
        )
        spec["edges"] = [
            {"source": "n1", "target": "n2", "when": "1 == 1"},
            {"source": "n1", "target": "n3", "is_default": True},
        ]
    if job_kind == "notify":
        spec["notifications"] = [
            {
                "id": "nt-1",
                "channel": "webhook",
                "url_secret_ref": "secret-ref-1",
            }
        ]
    return spec


def _workflow_row() -> dict[str, object]:
    return {
        "id": "wf-1",
        "project_id": "project-1",
        "name": "nightly-report",
        "dag_jsonb": {
            "nodes": [
                {
                    "id": "n1",
                    "job_kind": "sql_query",
                    "payload": {},
                    "retry_policy": None,
                    "timeout_seconds": 60,
                    "on_failure": "abort",
                    "when": None,
                }
            ],
            "edges": [],
            "schedule": {"cron": "0 2 * * *", "enabled": True},
        },
        "schedule_cron": "0 2 * * *",
        "schedule_enabled": True,
        "sensor_enabled": False,
        "enabled": True,
        "created_by": "user-1",
        "created_at": _dt(1),
        "updated_at": _dt(2),
    }


def _setting_row(key: str, value: str) -> dict[str, object]:
    return {"key": key, "value": value, "updated_at": _dt(1)}


def _user_row_for_login() -> dict[str, object]:
    return {
        "id": "user-1",
        "username": "admin",
        "role": "admin",
        "password_hash": "hash",
        "mfa_secret_ref": None,
    }


def _auth_headers(user_id: str = "user-1") -> dict[str, str]:
    token = create_access_token(user_id=user_id, role="admin", secret=_Services.jwt_secret)
    return {"Authorization": f"Bearer {token}"}


def _dt(second: int) -> datetime:
    return datetime(2026, 1, 1, 0, 0, second, tzinfo=UTC)


def _metadata_cache(payload: list[dict[str, object]]) -> dict[str, object]:
    return {"payload": payload, "expires_at": datetime.now(UTC) + timedelta(days=1)}


def _policy(**overrides: bool) -> dict[str, bool]:
    values = {
        "allow_select": True,
        "allow_explain": False,
        "allow_dm_explain": False,
        "allow_oracle_plan_table": False,
        "allow_schema_import": False,
        "allow_schema_save": False,
        "allow_scenario_write": False,
        "allow_record_task": False,
    }
    values.update(overrides)
    return values


def _datasource_row(
    *,
    name: str = "warehouse",
    password_secret_ref: str = "secret-old",
    operation_policy: dict[str, bool] | None = None,
) -> dict[str, object]:
    return {
        "id": "ds-1",
        "project_id": "project-1",
        "name": name,
        "db_type": "mysql",
        "host": "mysql.internal",
        "port": 3306,
        "username": "dataops",
        "database_name": "app",
        "password_secret_ref": password_secret_ref,
        "environment": "sandbox",
        "environment_verified": False,
        "capability_profile": {},
        "operation_policy": operation_policy or _policy(),
        "created_at": _dt(1),
    }


def _source_job_row() -> dict[str, object]:
    return {
        "id": "job-source",
        "kind": "sql_query",
        "status": "success",
        "owner_user_id": "user-1",
        "project_id": "project-1",
        "datasource_ids": ["ds-1"],
        "created_at": _dt(1),
        "started_at": _dt(2),
        "finished_at": _dt(3),
        "error": None,
        "error_code": None,
        "payload": {"result_set_id": "rs-source"},
    }


def _compare_task_row() -> dict[str, object]:
    return {
        "id": "task-1",
        "project_id": "project-1",
        "name": "orders",
        "source_id": "ds-source",
        "target_id": "ds-target",
        "source_ref": {"kind": "table", "schema_name": "app", "table_name": "orders_a"},
        "target_ref": {"kind": "table", "schema_name": "app", "table_name": "orders_b"},
        "columns": [
            {"name": "id", "type": "integer"},
            {"name": "amount", "type": "decimal"},
        ],
        "compare_rules": {"key_columns": ["id"], "schema_policy": "strict"},
        "run_limits": {
            "recursive_checksum": True,
            "bisection_factor": 8,
            "query_timeout_seconds": 1800,
        },
        "created_by": "user-1",
        "created_at": _dt(1),
        "updated_at": _dt(2),
    }


def _sql_compare_task_row(
    *,
    legacy_numeric: bool = False,
    explicit_numeric: bool = False,
) -> dict[str, object]:
    row = _compare_task_row()
    if explicit_numeric:
        sql = 'SELECT 1 AS "6" FROM T'
        row["columns"] = [{"name": "6", "type": "integer"}]
        row["compare_rules"] = {"key_columns": ["6"], "schema_policy": "strict"}
    else:
        sql = "SELECT CUST_NO, SUM(AMOUNT) FROM T GROUP BY CUST_NO"
        result_name = "2" if legacy_numeric else "RESULT_1"
        row["columns"] = [
            {"name": "CUST_NO", "type": "string"},
            {"name": result_name, "type": "decimal"},
        ]
        row["compare_rules"] = {
            "key_columns": ["CUST_NO"],
            "schema_policy": "strict",
        }
    row["name"] = "aggregate compare"
    row["source_ref"] = {"kind": "sql", "sql": sql}
    row["target_ref"] = {"kind": "sql", "sql": sql}
    return row


def _diff_profile_payload() -> dict[str, object]:
    return {
        "version": 1,
        "generated": True,
        "summary": {"diff_rows": 1, "same_rows": 10, "paired_rows_observed": 11},
        "columns": {
            "amount": {
                "type": "decimal",
                "observed_rows": 11,
                "changed_rows": 1,
                "diff_rate": 1.0,
                "numeric_delta": {
                    "count": 1,
                    "constant_offset": "-1.00",
                    "systematic_offset": True,
                },
            }
        },
        "missing_key_ranges": {"only_source": [], "only_target": []},
    }


def _sample_result_payload() -> dict[str, object]:
    return {
        "enabled": True,
        "mode": "pk_random_anchor",
        "requested_rows": 300,
        "sampled_rows": 300,
        "observed_differences": 0,
        "all_sampled_equal": True,
        "confidence": 0.95,
        "difference_rate_upper_bound": 0.00994,
    }


def _pagination_root_row(*, pagination_mode: str) -> dict[str, object]:
    return {
        "id": "job-root",
        "kind": "sql_query",
        "status": "success",
        "owner_user_id": "user-1",
        "project_id": "project-1",
        "datasource_ids": ["ds-1"],
        "priority": 0,
        "timeout_seconds": 300,
        "resource_profile": {},
        "payload": {
            "datasource_id": "ds-1",
            "sql": "SELECT id FROM users ORDER BY id",
            "params": {"min_id": 1},
            "result_set_id": "rs-pages",
            "page_size": 100,
            "max_result_rows": 1000,
            "pagination_mode": pagination_mode,
            "pagination_reason": (
                "fresh_read_ordered_offset"
                if pagination_mode == "ordered_offset"
                else "top_level_order_by_required"
            ),
            "page_offset": 0,
        },
    }


class _NoopServices:
    pass


class _Services:
    jwt_secret = "jwt-secret"

    def __init__(
        self,
        engine: _FakeEngine,
        secret_store: _SecretStore | None = None,
        job_backend: _JobBackend | None = None,
        result_store: _ResultStore | None = None,
        metadata_probe_timeout_seconds: int = 5,
        export_per_user_per_hour: int = 10,
        download_url_ttl_seconds: int = 300,
        upload_max_mb: int = 100,
        workflow_node_default_max_retries: int = 0,
        access_token_ttl_seconds: int = 3600,
        license_enforcement_enabled: bool = True,
    ) -> None:
        self.engine = engine
        self.rate_limiter = _RateLimiter()
        self.secret_store = secret_store or _SecretStore()
        self.job_backend = job_backend or _JobBackend()
        self.result_store = result_store or _ResultStore()
        self.metadata_probe_timeout_seconds = metadata_probe_timeout_seconds
        self.export_per_user_per_hour = export_per_user_per_hour
        self.download_url_ttl_seconds = download_url_ttl_seconds
        self.upload_max_mb = upload_max_mb
        self.workflow_node_default_max_retries = workflow_node_default_max_retries
        self._access_token_ttl_seconds = access_token_ttl_seconds
        self._license_enforcement_enabled = license_enforcement_enabled
        self.audits: list[dict[str, object]] = []

    def access_token_ttl_seconds(self) -> int:
        return self._access_token_ttl_seconds

    def license_enforcement_enabled(self) -> bool:
        return self._license_enforcement_enabled

    def current_license_mode(self) -> LicenseMode:
        return LicenseMode.TRIAL

    def is_token_revoked(
        self,
        *,
        user_id: str,
        issued_at: int,
        expires_at: int,
        jti: str | None,
    ) -> bool:
        del user_id, issued_at, expires_at, jti
        return False

    def write_audit(self, **kwargs: object) -> None:
        self.audits.append(kwargs)


class _RateLimiter:
    def allow(self, key: str) -> bool:
        return True


class _FakeEngine:
    def __init__(self, results: list[object]) -> None:
        self.results = list(results)
        self.statements: list[str] = []
        # 原始 statement 对象(便于测试 compile 出绑定参数,如 UPDATE 的 dag_jsonb)。
        self.executed: list[Any] = []

    def connect(self) -> _FakeConnection:
        return _FakeConnection(self)

    def begin(self) -> _FakeConnection:
        return _FakeConnection(self)


class _FakeConnection:
    def __init__(self, engine: _FakeEngine) -> None:
        self._engine = engine

    def __enter__(self) -> _FakeConnection:
        return self

    def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
        return None

    def execute(self, statement: object, parameters: object | None = None) -> _FakeResult:
        del parameters
        statement_text = str(statement)
        self._engine.statements.append(statement_text)
        self._engine.executed.append(statement)
        command = statement_text.lstrip().upper()
        if command.startswith("SELECT") or command.startswith("WITH"):
            return _FakeResult(self._engine.results.pop(0))
        return _FakeResult(None)


class _FakeResult:
    def __init__(self, result: object) -> None:
        self._result = result
        self.rowcount = 1

    def mappings(self) -> _FakeResult:
        return self

    def all(self) -> list[dict[str, Any]]:
        if isinstance(self._result, list):
            return [dict(row) for row in self._result if isinstance(row, dict)]
        raise AssertionError(f"expected list result, got {type(self._result).__name__}")

    def one_or_none(self) -> dict[str, Any] | None:
        if self._result is None:
            return None
        if isinstance(self._result, dict):
            return dict(self._result)
        raise AssertionError(f"expected single row result, got {type(self._result).__name__}")

    def scalar_one(self) -> object:
        if isinstance(self._result, int):
            return self._result
        raise AssertionError(f"expected scalar result, got {type(self._result).__name__}")

    def scalar_one_or_none(self) -> object | None:
        if self._result is None or isinstance(self._result, (int, str)):
            return self._result
        raise AssertionError(f"expected optional scalar result, got {type(self._result).__name__}")


class _JobBackend:
    def __init__(self) -> None:
        self.enqueued: list[Job] = []
        self.cancel_requested: list[str] = []
        self.cancelled_pending: list[tuple[str, str]] = []
        self.jobs_by_id: dict[str, Job] = {}

    def enqueue(self, job: Job) -> None:
        self.enqueued.append(job)
        self.jobs_by_id[job.id] = job

    def get_job(self, job_id: str) -> Job | None:
        return self.jobs_by_id.get(job_id)

    def request_cancel(self, job_id: str) -> None:
        self.cancel_requested.append(job_id)

    def cancel_pending_job(self, job_id: str, reason: str = "cancelled") -> None:
        self.cancelled_pending.append((job_id, reason))


class _ResultStore:
    def __init__(
        self,
        downloads: dict[str, bytes] | None = None,
        spool_exists: bool = True,
        rows: list[Row] | None = None,
        manifest: dict[str, object] | None = None,
    ) -> None:
        self._downloads = downloads or {}
        self._spool_exists = spool_exists
        self._rows = rows or []
        self._manifest = manifest
        self.export_artifacts: list[tuple[str, str, bytes]] = []
        self.upload_artifacts: list[tuple[str, str, bytes]] = []

    def put_export_artifact(self, export_id: str, name: str, stream: Any) -> ResultRef:
        self.export_artifacts.append((export_id, name, stream.read()))
        return ResultRef(backend="local_fs", uri=f"exports/{export_id}/{name}")

    def put_upload_artifact(self, upload_id: str, name: str, stream: Any) -> ResultRef:
        self.upload_artifacts.append((upload_id, name, stream.read()))
        return ResultRef(backend="local_fs", uri=f"uploads/{upload_id}/{name}")

    def spool_ref(self, result_set_id: str) -> ResultRef:
        return ResultRef(backend="local_fs", uri=f"spool/{result_set_id}")

    def spool_exists(self, result_set_id: str) -> bool:
        del result_set_id
        return self._spool_exists

    def get_spool_manifest(self, result_set_id: str) -> dict[str, object]:
        del result_set_id
        if self._manifest is not None:
            return dict(self._manifest)
        return {
            "columns": [{"name": "value", "type": "string"}],
            "loaded_rows": 1,
            "truncated": False,
        }

    def fetch_range(self, result_set_id: str, offset: int, limit: int) -> list[Row]:
        del result_set_id
        return list(self._rows[offset : offset + limit])

    def open_download(self, ref: ResultRef) -> io.BytesIO:
        if ref.uri not in self._downloads:
            raise FileNotFoundError(ref.uri)
        return io.BytesIO(self._downloads[ref.uri])


class _MetadataAdapter:
    def list_schemas(self) -> list[Schema]:
        return [Schema(name="app")]

    def list_tables(self, schema: str) -> list[Table]:
        return [Table(schema_name=schema, name="users", table_type="BASE TABLE")]

    def list_columns(self, schema: str, table: str) -> list[Column]:
        del schema, table
        return [
            Column(
                name="id",
                type=ColumnType.INTEGER,
                driver_type="INT",
                nullable=False,
                primary_key=True,
                comment="primary key",
            ),
            Column(
                name="name",
                type=ColumnType.STRING,
                driver_type="VARCHAR(64)",
                nullable=True,
                primary_key=False,
            ),
        ]

    def list_indexes(self, schema: str, table: str) -> list[Index]:
        del schema, table
        return [Index(name="PRIMARY", columns=["id"], is_unique=True, is_primary=True)]


class _FailingMetadataAdapter:
    def list_schemas(self) -> list[Schema]:
        raise AdapterConnectionError("adapter connection failed")

    def list_columns(self, schema: str, table: str) -> list[Column]:
        del schema, table
        raise AdapterConnectionError("adapter connection failed")


class _SecretStore:
    def __init__(self) -> None:
        self.stored: list[str] = []
        self.deleted: list[str] = []

    def verify_password(self, plaintext: str, ref: object) -> bool:
        del ref
        return plaintext == "secret"

    def store_secret(self, plaintext: str, kind: object) -> object:
        self.stored.append(plaintext)
        return _SecretRef("secret-new")

    def delete_secret(self, ref: Any) -> None:
        self.deleted.append(str(ref.ref))


class _SecretRef:
    def __init__(self, ref: str) -> None:
        self.ref = ref
