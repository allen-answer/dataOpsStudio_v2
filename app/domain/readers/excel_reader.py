"""Excel(.xlsx / .xlsm)行读取器(C-1 PR1)。

- **openpyxl** `load_workbook(read_only=True, data_only=True)` 流式读:
  只读模式内存有界;`data_only` 取公式最后保存值(不重算)。
- 可配 `sheet`(空=第一个)、`header_row`(1-indexed);表头前的导言行被跳过("跳行")。
- 丢弃空表头列(`build_columns`),跳过全空数据行(`is_blank_row`)。
- 值保留 openpyxl 原生类型(int / float / datetime / bool / None);
  类型规范化留给对比内核 `CompareRules`,reader 只取原值(设计稿 §一.2)。

纯 domain,零 DB 依赖(R1)。openpyxl 非 DB 驱动,不受 R1 banned-api 约束。
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import load_workbook

from app.domain.readers.base import (
    ReaderError,
    RowReader,
    build_columns,
    is_blank_row,
    row_to_dict,
)

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook


def _load_workbook(path: Path) -> Workbook:
    try:
        return load_workbook(path, read_only=True, data_only=True)
    except Exception as error:  # openpyxl 抛多种异常;统一归一为 ReaderError
        raise ReaderError(f"could not open Excel workbook: {error}") from error


class ExcelReader(RowReader):
    def __init__(
        self,
        path: str | Path,
        *,
        sheet: str | None = None,
        header_row: int = 1,
    ) -> None:
        if header_row < 1:
            raise ValueError("header_row must be >= 1 (1-indexed)")
        self._path = Path(path)
        self._sheet = sheet
        self._header_row = header_row

    def _select_sheet_name(self, workbook: Workbook) -> str:
        sheet_names = workbook.sheetnames
        if not sheet_names:
            raise ReaderError("workbook has no sheets")
        if self._sheet is None:
            return str(sheet_names[0])
        if self._sheet not in sheet_names:
            raise ReaderError(f"sheet not found: {self._sheet!r}")
        return self._sheet

    def columns(self) -> list[str]:
        names, _ = self._header()
        return names

    def _header(self) -> tuple[list[str], list[int]]:
        workbook = _load_workbook(self._path)
        try:
            worksheet = workbook[self._select_sheet_name(workbook)]
            for index, raw in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if index == self._header_row:
                    return build_columns(list(raw))
            return [], []
        finally:
            workbook.close()

    def iter_rows(self) -> Iterator[dict[str, Any]]:
        workbook = _load_workbook(self._path)
        try:
            worksheet = workbook[self._select_sheet_name(workbook)]
            columns: list[str] = []
            kept_indices: list[int] = []
            for index, raw in enumerate(worksheet.iter_rows(values_only=True), start=1):
                values = list(raw)
                if index < self._header_row:
                    continue
                if index == self._header_row:
                    columns, kept_indices = build_columns(values)
                    if not columns:
                        return
                    continue
                if is_blank_row(values):
                    continue
                yield row_to_dict(columns, kept_indices, values)
        finally:
            workbook.close()


def list_sheets(path: str | Path) -> list[str]:
    """返回工作簿 sheet 名列表(给"上传后选 sheet"UI 用)。"""

    workbook = _load_workbook(Path(path))
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def list_columns(
    path: str | Path,
    *,
    sheet: str | None = None,
    header_row: int = 1,
) -> list[str]:
    """给"选主键 / 列映射"UI 用:只读指定 sheet 表头返回列名。"""

    return ExcelReader(path, sheet=sheet, header_row=header_row).columns()


__all__ = ["ExcelReader", "list_columns", "list_sheets"]
